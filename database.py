#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""数据库模块 - SQLite v1.5"""

import sqlite3
import os
import json
from datetime import datetime




def _get_data_dir(data_dir=None):
    """获取数据目录路径，支持自定义"""
    if data_dir is not None:
        return data_dir
    return os.path.join(os.path.expanduser("~"), "采购管理系统数据")


class Database:
    def __init__(self, data_dir=None):
        self.data_dir = _get_data_dir(data_dir)
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = os.path.join(self.data_dir, "procurement.db")
        # 使用 URI 模式 + rwc 确保可读写创建，防止 PyInstaller 环境下
        # "unable to open database file" 错误
        db_uri = self.db_path.replace("\\", "/")
        try:
            self.conn = sqlite3.connect(f"file:{db_uri}?mode=rwc", uri=True,
                                        check_same_thread=False)
        except Exception:
            # 回退到标准连接方式
            self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        # 使用内存日志，避免打包后磁盘 journal 权限问题
        self.conn.execute("PRAGMA journal_mode=MEMORY")
        self.conn.row_factory = sqlite3.Row
        self._init_tables()

    def _init_tables(self):
        c = self.conn.cursor()
        # 采购垫付主表
        c.execute("""
            CREATE TABLE IF NOT EXISTS purchase (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                project TEXT NOT NULL,
                handler TEXT NOT NULL,
                payment_method TEXT NOT NULL,
                invoice_status TEXT NOT NULL,
                reimbursement_status TEXT NOT NULL,
                remark TEXT,
                archived INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        # 采购物料明细
        c.execute("""
            CREATE TABLE IF NOT EXISTS purchase_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                purchase_id INTEGER NOT NULL,
                name TEXT,
                spec TEXT,
                quantity REAL,
                unit_price REAL,
                supplier TEXT,
                total REAL,
                FOREIGN KEY(purchase_id) REFERENCES purchase(id)
            )
        """)
        # 自定义项目
        c.execute("""
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        c.execute("INSERT OR IGNORE INTO projects(name) VALUES('默认项目')")
        c.execute("INSERT OR IGNORE INTO projects(name) VALUES('电商新品')")
        c.execute("INSERT OR IGNORE INTO projects(name) VALUES('传渠项目')")

        # 物料台账
        c.execute("""
            CREATE TABLE IF NOT EXISTS material_ledger (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contract_no TEXT,
                supplier TEXT,
                item_no TEXT,
                material_name TEXT,
                quantity REAL,
                unit TEXT,
                unit_price REAL,
                amount REAL,
                year TEXT,
                raw_data TEXT
            )
        """)

        # 包材下单表
        c.execute("""
            CREATE TABLE IF NOT EXISTS packaging_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                material_name TEXT NOT NULL,
                project TEXT NOT NULL,
                project_no TEXT,
                order_factory TEXT,
                compare_price REAL,
                compare_date TEXT,
                compare_remark TEXT,
                contract_status TEXT,
                contract_remark TEXT,
                notify_date TEXT,
                expected_delivery_date TEXT,
                notify_remark TEXT,
                production_cycle TEXT,
                expected_ship_date TEXT,
                production_remark TEXT,
                ship_date TEXT,
                ship_method TEXT,
                tracking_no TEXT,
                expected_arrival TEXT,
                notify_warehouse INTEGER DEFAULT 0,
                archived INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # 差旅主表
        c.execute("""
            CREATE TABLE IF NOT EXISTS travel (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reason TEXT NOT NULL,
                destination TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                duration INTEGER,
                handler TEXT,
                invoice_status TEXT NOT NULL,
                reimbursement_status TEXT NOT NULL,
                remark TEXT,
                archived INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        # 差旅交通明细
        c.execute("""
            CREATE TABLE IF NOT EXISTS travel_transport (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                travel_id INTEGER NOT NULL,
                transport_type TEXT,
                travel_date TEXT,
                departure TEXT,
                destination TEXT,
                amount REAL,
                FOREIGN KEY(travel_id) REFERENCES travel(id)
            )
        """)
        # 差旅住宿明细
        c.execute("""
            CREATE TABLE IF NOT EXISTS travel_hotel (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                travel_id INTEGER NOT NULL,
                checkin_date TEXT,
                checkout_date TEXT,
                room_count INTEGER,
                amount REAL,
                invoice_status TEXT,
                FOREIGN KEY(travel_id) REFERENCES travel(id)
            )
        """)

        # 供应商管理表 (v1.2 新增)
        c.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT,
                main_product TEXT,
                contact_person TEXT,
                phone TEXT,
                wechat TEXT,
                cooperation_status TEXT DEFAULT '接洽中',
                quote_status TEXT,
                sample_status TEXT,
                payment_method TEXT,
                invoice_type TEXT,
                tax_rate TEXT,
                remark TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # 催款记录表 (v1.2 新增)
        c.execute("""
            CREATE TABLE IF NOT EXISTS collection_reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_name TEXT NOT NULL,
                contact_person TEXT,
                wechat TEXT,
                reminder_date TEXT,
                amount_due REAL,
                notify_internal INTEGER DEFAULT 0,
                notify_manager INTEGER DEFAULT 0,
                remark TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # 备忘录表 (v1.3 新增)
        c.execute("""
            CREATE TABLE IF NOT EXISTS memos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                project TEXT,
                handler TEXT,
                content TEXT,
                deadline TEXT,
                status TEXT DEFAULT '待处理',
                remark TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # 报价单产品表 (v1.3 新增)
        c.execute("""
            CREATE TABLE IF NOT EXISTS quotation_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_no TEXT,
                product_name TEXT NOT NULL,
                product_size TEXT,
                material_process TEXT,
                supply_cycle TEXT,
                carton_spec TEXT,
                unit TEXT DEFAULT 'PCS',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # 报价单阶梯价格表 (v1.3 新增)
        c.execute("""
            CREATE TABLE IF NOT EXISTS quotation_tiers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                tier_name TEXT NOT NULL,
                min_qty INTEGER NOT NULL DEFAULT 0,
                max_qty INTEGER,
                unit_price REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(product_id) REFERENCES quotation_products(id) ON DELETE CASCADE
            )
        """)

        # 报价单配置表 (v1.3 新增) — 需方信息、条款等
        c.execute("""
            CREATE TABLE IF NOT EXISTS quotation_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                buyer_name TEXT DEFAULT '北京同仁堂健康药业（青海）有限公司',
                buyer_contact TEXT DEFAULT '王维',
                buyer_phone TEXT DEFAULT '13897764859',
                buyer_address TEXT DEFAULT '青海省海西州德令哈市同仁堂路1号',
                payment_terms TEXT DEFAULT '按协议条件付款',
                transport_method TEXT DEFAULT '物料或者专车请提前说明',
                delivery_docs TEXT DEFAULT '请随货放【发货单】【厂检报告】',
                quote_requirement TEXT DEFAULT '需含税含运',
                quote_template_note TEXT DEFAULT '报价单模板由需方提供',
                footer_note TEXT DEFAULT '请写明产品尺寸和详细的材质工艺、发货包装形式、箱规等信息',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        c.execute("INSERT OR IGNORE INTO quotation_config(id) VALUES(1)")
        
        # 报价单供方配置表 (v1.3 新增) — 供方信息
        c.execute("""
            CREATE TABLE IF NOT EXISTS quotation_supplier (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_name TEXT DEFAULT '',
                contact_person TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                address TEXT DEFAULT '',
                quote_date TEXT DEFAULT '',
                quote_validity TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        c.execute("INSERT OR IGNORE INTO quotation_supplier(id) VALUES(1)")

        # ====== 合同供应商表 (v1.9.0 从 JSON 迁移) ======
        c.execute("""
            CREATE TABLE IF NOT EXISTS contract_suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                short_name TEXT DEFAULT '',
                full_name TEXT DEFAULT '',
                legal_rep TEXT DEFAULT '',
                address TEXT DEFAULT '',
                contact TEXT DEFAULT '',
                auth_rep TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                fax TEXT DEFAULT '',
                payment_days TEXT DEFAULT '90',
                payment_method TEXT DEFAULT '电汇',
                account_name TEXT DEFAULT '',
                bank TEXT DEFAULT '',
                account TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # ====== 合同甲方配置表 (v1.9.0 从 JSON 迁移) ======
        c.execute("""
            CREATE TABLE IF NOT EXISTS contract_party_a (
                id INTEGER PRIMARY KEY DEFAULT 1,
                company_name TEXT DEFAULT '北京同仁堂健康药业（青海）有限公司',
                legal_rep TEXT DEFAULT '施能文',
                address TEXT DEFAULT '',
                contact TEXT DEFAULT '龙存英',
                phone TEXT DEFAULT '13897764859'
            )
        """)
        c.execute("INSERT OR IGNORE INTO contract_party_a(id) VALUES(1)")

        self.conn.commit()
        self._migrate_tables()
        self._import_json_data()

    def _migrate_tables(self):
        """数据库迁移：检测并修复所有表的列名和结构不匹配"""
        c = self.conn.cursor()

        def _rebuild_table(table, new_ddl, col_map=None):
            c.execute(f"PRAGMA table_info({table})")
            old_cols = [row[1] for row in c.fetchall()]
            has_data = bool(c.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])

            if col_map and has_data:
                try:
                    c.execute(f"CREATE TABLE IF NOT EXISTS {table}_backup AS SELECT * FROM {table}")
                    c.execute(f"DROP TABLE {table}")
                    c.execute(new_ddl)
                    old_common = [oc for oc in old_cols if oc in col_map]
                    new_common = [col_map[oc] for oc in old_common]
                    if old_common:
                        cols = ", ".join(new_common)
                        placeholders = ", ".join("?" * len(new_common))
                        c.execute(f"SELECT {', '.join(old_common)} FROM {table}_backup")
                        for row in c.fetchall():
                            c.execute(f"INSERT INTO {table}({cols}) VALUES({placeholders})", row)
                    c.execute(f"DROP TABLE {table}_backup")
                    return
                except Exception:
                    c.execute(f"DROP TABLE IF EXISTS {table}_backup")
            c.execute(f"DROP TABLE IF EXISTS {table}")
            c.execute(new_ddl)

        # material_ledger — 添加缺失列
        c.execute("PRAGMA table_info(material_ledger)")
        ml_cols = [row[1] for row in c.fetchall()]
        for col_name, col_type in [
            ("contract_no", "TEXT"), ("supplier", "TEXT"), ("item_no", "TEXT"),
            ("material_name", "TEXT"), ("quantity", "REAL"), ("unit", "TEXT"),
            ("unit_price", "REAL"), ("amount", "REAL"), ("year", "TEXT"), ("raw_data", "TEXT"),
        ]:
            if col_name not in ml_cols:
                try:
                    c.execute(f"ALTER TABLE material_ledger ADD COLUMN {col_name} {col_type}")
                except Exception:
                    pass

        # purchase — 添加缺失列
        c.execute("PRAGMA table_info(purchase)")
        p_cols = [row[1] for row in c.fetchall()]
        for col_name, col_type in [
            ("project", "TEXT NOT NULL DEFAULT '默认项目'"),
            ("archived", "INTEGER DEFAULT 0"),
        ]:
            if col_name not in p_cols:
                try:
                    c.execute(f"ALTER TABLE purchase ADD COLUMN {col_name} {col_type}")
                except Exception:
                    pass
        try:
            c.execute("UPDATE purchase SET project='默认项目' WHERE project IS NULL OR project=''")
        except Exception:
            pass

        # purchase_items — 列名重命名
        c.execute("PRAGMA table_info(purchase_items)")
        pi_cols = [row[1] for row in c.fetchall()]
        if "name" not in pi_cols or "spec" not in pi_cols or "total" not in pi_cols or "supplier" not in pi_cols:
            _rebuild_table("purchase_items", """CREATE TABLE purchase_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                purchase_id INTEGER NOT NULL,
                name TEXT, spec TEXT, quantity REAL, unit_price REAL,
                supplier TEXT, total REAL,
                FOREIGN KEY(purchase_id) REFERENCES purchase(id)
            )""", col_map={
                "purchase_id": "purchase_id", "material_name": "name",
                "specification": "spec", "quantity": "quantity",
                "unit_price": "unit_price", "total_price": "total",
            })

        # packaging_orders — 结构检查
        c.execute("PRAGMA table_info(packaging_orders)")
        po_cols = [row[1] for row in c.fetchall()]
        if "compare_price" not in po_cols or "contract_status" not in po_cols:
            _rebuild_table("packaging_orders", """
                CREATE TABLE packaging_orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    material_name TEXT NOT NULL, project TEXT NOT NULL,
                    project_no TEXT,
                    order_factory TEXT,
                    compare_price REAL, compare_date TEXT, compare_remark TEXT,
                    contract_status TEXT, contract_remark TEXT,
                    notify_date TEXT, expected_delivery_date TEXT, notify_remark TEXT,
                    production_cycle TEXT, expected_ship_date TEXT, production_remark TEXT,
                    ship_date TEXT, ship_method TEXT, tracking_no TEXT,
                    expected_arrival TEXT, notify_warehouse INTEGER DEFAULT 0,
                    archived INTEGER DEFAULT 0,
                    created_at TEXT DEFAULT (datetime('now','localtime'))
                )
            """, col_map=None)

        # packaging_orders — 添加缺失列（v1.5 新增）
        c.execute("PRAGMA table_info(packaging_orders)")
        po_cols2 = [row[1] for row in c.fetchall()]
        for col_name, col_type in [
            ("project_no", "TEXT"),
            ("order_factory", "TEXT"),
            ("order_quantity", "TEXT"),
        ]:
            if col_name not in po_cols2:
                try:
                    c.execute(f"ALTER TABLE packaging_orders ADD COLUMN {col_name} {col_type}")
                except Exception:
                    pass

        # travel — 添加缺失列
        c.execute("PRAGMA table_info(travel)")
        t_cols = [row[1] for row in c.fetchall()]
        for col_name, col_type in [
            ("duration", "INTEGER"), ("handler", "TEXT"), ("archived", "INTEGER DEFAULT 0"),
        ]:
            if col_name not in t_cols:
                try:
                    c.execute(f"ALTER TABLE travel ADD COLUMN {col_name} {col_type}")
                except Exception:
                    pass
        if "traveler" in t_cols:
            c.execute("PRAGMA table_info(travel)")
            travel_info = {row[1]: row[3] for row in c.fetchall()}
            if travel_info.get("traveler"):
                _rebuild_table("travel", """CREATE TABLE travel (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    reason TEXT NOT NULL, destination TEXT NOT NULL,
                    start_date TEXT NOT NULL, end_date TEXT NOT NULL,
                    duration INTEGER, handler TEXT,
                    invoice_status TEXT NOT NULL DEFAULT '未开票',
                    reimbursement_status TEXT NOT NULL DEFAULT '未报销',
                    remark TEXT, archived INTEGER DEFAULT 0,
                    created_at TEXT DEFAULT (datetime('now','localtime'))
                )""", col_map={
                    "id": "id", "reason": "reason", "destination": "destination",
                    "start_date": "start_date", "end_date": "end_date",
                    "duration": "duration", "handler": "handler", "traveler": "handler",
                    "invoice_status": "invoice_status", "reimbursement_status": "reimbursement_status",
                    "remark": "remark", "archived": "archived", "created_at": "created_at",
                })

        # travel_transport
        c.execute("PRAGMA table_info(travel_transport)")
        tt_cols = [row[1] for row in c.fetchall()]
        if "transport_date" in tt_cols and "travel_date" not in tt_cols:
            _rebuild_table("travel_transport", """CREATE TABLE travel_transport (
                id INTEGER PRIMARY KEY AUTOINCREMENT, travel_id INTEGER NOT NULL,
                transport_type TEXT, travel_date TEXT, departure TEXT,
                destination TEXT, amount REAL,
                FOREIGN KEY(travel_id) REFERENCES travel(id)
            )""", col_map={
                "travel_id": "travel_id", "transport_type": "transport_type",
                "transport_date": "travel_date", "departure": "departure",
                "destination": "destination", "amount": "amount",
            })

        # travel_hotel
        c.execute("PRAGMA table_info(travel_hotel)")
        th_cols = [row[1] for row in c.fetchall()]
        if "check_in_date" in th_cols or "check_out_date" in th_cols or "invoice_status" not in th_cols:
            _rebuild_table("travel_hotel", """CREATE TABLE travel_hotel (
                id INTEGER PRIMARY KEY AUTOINCREMENT, travel_id INTEGER NOT NULL,
                checkin_date TEXT, checkout_date TEXT, room_count INTEGER,
                amount REAL, invoice_status TEXT,
                FOREIGN KEY(travel_id) REFERENCES travel(id)
            )""", col_map={
                "travel_id": "travel_id", "check_in_date": "checkin_date",
                "check_out_date": "checkout_date", "room_count": "room_count", "amount": "amount",
            })

        # suppliers — 添加缺失列
        c.execute("PRAGMA table_info(suppliers)")
        s_cols = [row[1] for row in c.fetchall()]
        for col_name, col_type in [
            ("name", "TEXT NOT NULL DEFAULT ''"),
            ("category", "TEXT"), ("main_product", "TEXT"),
            ("contact_person", "TEXT"), ("phone", "TEXT"), ("wechat", "TEXT"),
            ("cooperation_status", "TEXT DEFAULT '接洽中'"),
            ("quote_status", "TEXT"), ("sample_status", "TEXT"),
            ("payment_method", "TEXT"), ("invoice_type", "TEXT"),
            ("tax_rate", "TEXT"), ("remark", "TEXT"),
            ("created_at", "TEXT"),
        ]:
            if col_name not in s_cols:
                try:
                    c.execute(f"ALTER TABLE suppliers ADD COLUMN {col_name} {col_type}")
                except Exception:
                    pass

        self.conn.commit()

    def _import_json_data(self):
        """首次启动时从旧版 JSON 文件导入数据到数据库（v1.9.0）"""
        c = self.conn.cursor()

        # 导入合同供应商
        c.execute("SELECT COUNT(*) FROM contract_suppliers")
        if c.fetchone()[0] == 0:
            # 尝试从 assets/suppliers.json 导入
            import sys
            json_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "suppliers.json"),
            ]
            if hasattr(sys, '_MEIPASS'):
                json_paths.insert(0, os.path.join(sys._MEIPASS, "assets", "suppliers.json"))
            for jp in json_paths:
                if os.path.exists(jp):
                    try:
                        with open(jp, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        suppliers = data.get("suppliers", [])
                        for s in suppliers:
                            c.execute("""
                                INSERT INTO contract_suppliers(short_name, full_name, legal_rep,
                                address, contact, auth_rep, phone, fax, payment_days,
                                payment_method, account_name, bank, account)
                                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """, (
                                s.get("short_name",""), s.get("full_name",""), s.get("legal_rep",""),
                                s.get("address",""), s.get("contact",""), s.get("auth_rep",""),
                                s.get("phone",""), s.get("fax",""), s.get("payment_days","90"),
                                s.get("payment_method","电汇"), s.get("account_name",""),
                                s.get("bank",""), s.get("account","")
                            ))
                        if suppliers:
                            self.conn.commit()
                    except Exception:
                        pass
                    break

        # 导入合同甲方配置
        c.execute("SELECT company_name, legal_rep, contact, phone FROM contract_party_a WHERE id=1")
        row = c.fetchone()
        # 如果还是默认值（未被用户编辑过），尝试从JSON导入旧数据
        if row and row[0] == "北京同仁堂健康药业（青海）有限公司":
            import sys
            json_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "party_a.json"),
            ]
            if hasattr(sys, '_MEIPASS'):
                json_paths.insert(0, os.path.join(sys._MEIPASS, "assets", "party_a.json"))
            for jp in json_paths:
                if os.path.exists(jp):
                    try:
                        with open(jp, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        c.execute("""
                            UPDATE contract_party_a SET company_name=?, legal_rep=?,
                            address=?, contact=?, phone=?
                            WHERE id=1
                        """, (
                            data.get("company_name","北京同仁堂健康药业（青海）有限公司"),
                            data.get("legal_rep","施能文"),
                            data.get("address",""),
                            data.get("contact","龙存英"),
                            data.get("phone","13897764859"),
                        ))
                        self.conn.commit()
                    except Exception:
                        pass
                    break

    # ====== 项目管理 ======
    def get_projects(self):
        c = self.conn.cursor()
        c.execute("SELECT name FROM projects ORDER BY id")
        return [r[0] for r in c.fetchall()]

    def add_project(self, name):
        c = self.conn.cursor()
        c.execute("INSERT OR IGNORE INTO projects(name) VALUES(?)", (name,))
        self.conn.commit()

    # ====== 采购垫付 ======
    def save_purchase(self, data, items):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO purchase(date,project,handler,payment_method,invoice_status,reimbursement_status,remark)
            VALUES(:date,:project,:handler,:payment_method,:invoice_status,:reimbursement_status,:remark)
        """, data)
        pid = c.lastrowid
        for item in items:
            item["purchase_id"] = pid
            c.execute("""
                INSERT INTO purchase_items(purchase_id,name,spec,quantity,unit_price,supplier,total)
                VALUES(:purchase_id,:name,:spec,:quantity,:unit_price,:supplier,:total)
            """, item)
        self.conn.commit()
        return pid

    def update_purchase(self, pid, data, items):
        c = self.conn.cursor()
        c.execute("""
            UPDATE purchase SET date=:date,project=:project,handler=:handler,
            payment_method=:payment_method,invoice_status=:invoice_status,
            reimbursement_status=:reimbursement_status,remark=:remark WHERE id=:id
        """, {**data, "id": pid})
        c.execute("DELETE FROM purchase_items WHERE purchase_id=?", (pid,))
        for item in items:
            item["purchase_id"] = pid
            c.execute("""
                INSERT INTO purchase_items(purchase_id,name,spec,quantity,unit_price,supplier,total)
                VALUES(:purchase_id,:name,:spec,:quantity,:unit_price,:supplier,:total)
            """, item)
        self.conn.commit()

    def get_purchases(self, archived=0):
        c = self.conn.cursor()
        c.execute("SELECT * FROM purchase WHERE archived=? ORDER BY date DESC", (archived,))
        rows = c.fetchall()
        result = []
        for r in rows:
            d = dict(r)
            c.execute("SELECT * FROM purchase_items WHERE purchase_id=?", (d["id"],))
            d["items"] = [dict(i) for i in c.fetchall()]
            d["total"] = sum(i["total"] or 0 for i in d["items"])
            result.append(d)
        return result

    def archive_purchase(self, pid):
        c = self.conn.cursor()
        c.execute("UPDATE purchase SET archived=1 WHERE id=?", (pid,))
        self.conn.commit()

    def delete_purchase(self, pid):
        c = self.conn.cursor()
        c.execute("DELETE FROM purchase_items WHERE purchase_id=?", (pid,))
        c.execute("DELETE FROM purchase WHERE id=?", (pid,))
        self.conn.commit()

    # ====== 差旅 ======
    def save_travel(self, data, transports, hotels):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO travel(reason,destination,start_date,end_date,duration,handler,
            invoice_status,reimbursement_status,remark)
            VALUES(:reason,:destination,:start_date,:end_date,:duration,:handler,
            :invoice_status,:reimbursement_status,:remark)
        """, data)
        tid = c.lastrowid
        for t in transports:
            t["travel_id"] = tid
            c.execute("""
                INSERT INTO travel_transport(travel_id,transport_type,travel_date,departure,destination,amount)
                VALUES(:travel_id,:transport_type,:travel_date,:departure,:destination,:amount)
            """, t)
        for h in hotels:
            h["travel_id"] = tid
            c.execute("""
                INSERT INTO travel_hotel(travel_id,checkin_date,checkout_date,room_count,amount,invoice_status)
                VALUES(:travel_id,:checkin_date,:checkout_date,:room_count,:amount,:invoice_status)
            """, h)
        self.conn.commit()
        return tid

    def update_travel(self, tid, data, transports, hotels):
        c = self.conn.cursor()
        c.execute("""
            UPDATE travel SET reason=:reason,destination=:destination,start_date=:start_date,
            end_date=:end_date,duration=:duration,handler=:handler,
            invoice_status=:invoice_status,reimbursement_status=:reimbursement_status,remark=:remark
            WHERE id=:id
        """, {**data, "id": tid})
        c.execute("DELETE FROM travel_transport WHERE travel_id=?", (tid,))
        c.execute("DELETE FROM travel_hotel WHERE travel_id=?", (tid,))
        for t in transports:
            t["travel_id"] = tid
            c.execute("""
                INSERT INTO travel_transport(travel_id,transport_type,travel_date,departure,destination,amount)
                VALUES(:travel_id,:transport_type,:travel_date,:departure,:destination,:amount)
            """, t)
        for h in hotels:
            h["travel_id"] = tid
            c.execute("""
                INSERT INTO travel_hotel(travel_id,checkin_date,checkout_date,room_count,amount,invoice_status)
                VALUES(:travel_id,:checkin_date,:checkout_date,:room_count,:amount,:invoice_status)
            """, h)
        self.conn.commit()

    def get_travels(self, archived=0):
        c = self.conn.cursor()
        c.execute("SELECT * FROM travel WHERE archived=? ORDER BY start_date DESC", (archived,))
        rows = c.fetchall()
        result = []
        for r in rows:
            d = dict(r)
            c.execute("SELECT * FROM travel_transport WHERE travel_id=?", (d["id"],))
            d["transports"] = [dict(i) for i in c.fetchall()]
            c.execute("SELECT * FROM travel_hotel WHERE travel_id=?", (d["id"],))
            d["hotels"] = [dict(i) for i in c.fetchall()]
            t_total = sum(i["amount"] or 0 for i in d["transports"])
            h_total = sum(i["amount"] or 0 for i in d["hotels"])
            d["total"] = t_total + h_total
            result.append(d)
        return result

    def archive_travel(self, tid):
        c = self.conn.cursor()
        c.execute("UPDATE travel SET archived=1 WHERE id=?", (tid,))
        self.conn.commit()

    def delete_travel(self, tid):
        c = self.conn.cursor()
        c.execute("DELETE FROM travel_transport WHERE travel_id=?", (tid,))
        c.execute("DELETE FROM travel_hotel WHERE travel_id=?", (tid,))
        c.execute("DELETE FROM travel WHERE id=?", (tid,))
        self.conn.commit()

    def close(self):
        self.conn.close()

    # ====== 备忘录 (v1.3) ======
    def save_memo(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO memos(date, project, handler, content, deadline, status, remark)
            VALUES(:date, :project, :handler, :content, :deadline, :status, :remark)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_memo(self, mid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE memos SET date=:date, project=:project, handler=:handler,
            content=:content, deadline=:deadline, status=:status, remark=:remark
            WHERE id=:id
        """, {**data, "id": mid})
        self.conn.commit()

    def get_memos(self, keyword=None, project=None, status=None):
        c = self.conn.cursor()
        sql = "SELECT * FROM memos WHERE 1=1"
        params = []
        if keyword:
            sql += " AND (content LIKE ? OR handler LIKE ? OR remark LIKE ?)"
            kw = f"%{keyword}%"
            params.extend([kw, kw, kw])
        if project and project != "全部":
            sql += " AND project=?"
            params.append(project)
        if status and status != "全部":
            sql += " AND status=?"
            params.append(status)
        sql += " ORDER BY date DESC"
        c.execute(sql, params)
        return [dict(r) for r in c.fetchall()]

    def get_memo(self, mid):
        c = self.conn.cursor()
        c.execute("SELECT * FROM memos WHERE id=?", (mid,))
        row = c.fetchone()
        return dict(row) if row else None

    def delete_memo(self, mid):
        c = self.conn.cursor()
        c.execute("DELETE FROM memos WHERE id=?", (mid,))
        self.conn.commit()

    # ====== 物料台账 ======
    def save_material_ledger(self, rows):
        c = self.conn.cursor()
        c.execute("DELETE FROM material_ledger")
        for r in rows:
            c.execute("""
                INSERT INTO material_ledger(contract_no,supplier,item_no,material_name,
                quantity,unit,unit_price,amount,year,raw_data)
                VALUES(?,?,?,?,?,?,?,?,?,?)
            """, (
                r.get("contract_no", ""), r.get("supplier", ""), r.get("item_no", ""),
                r.get("material_name", ""), r.get("quantity") or 0, r.get("unit", ""),
                r.get("unit_price") or 0, r.get("amount") or 0, r.get("year", ""),
                json.dumps(r.get("_raw", {}), ensure_ascii=False),
            ))
        self.conn.commit()

    def get_material_ledger(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM material_ledger ORDER BY id")
        rows = []
        for r in c.fetchall():
            d = dict(r)
            try:
                d["_raw"] = json.loads(d.get("raw_data") or "{}")
            except Exception:
                d["_raw"] = {}
            rows.append(d)
        return rows

    def clear_material_ledger(self):
        c = self.conn.cursor()
        c.execute("DELETE FROM material_ledger")
        self.conn.commit()

    # ====== 通用导出 ======
    @staticmethod
    def export_to_xlsx(filepath, sheet_name, headers, rows, col_widths=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            hdr_fill = PatternFill(start_color="C1816D", end_color="C1816D", fill_type="solid")
            hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border

            data_font = Font(name="微软雅黑", size=10)
            data_align = Alignment(vertical="center")
            alt_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

            for ri, row in enumerate(rows, 2):
                for ci, key in enumerate(headers, 1):
                    val = row.get(key, "") if isinstance(row, dict) else ""
                    cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
                    if ri % 2 == 0:
                        cell.fill = alt_fill

            if col_widths:
                for ci, w in enumerate(col_widths, 1):
                    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
            else:
                for ci in range(1, len(headers) + 1):
                    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 16

            wb.save(filepath)
            return True
        except Exception as e:
            raise e

    # ====== 包材下单 ======
    def save_packaging(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO packaging_orders(
                material_name, project, project_no, order_factory, order_quantity,
                compare_price, compare_date, compare_remark,
                contract_status, contract_remark,
                notify_date, expected_delivery_date, notify_remark,
                production_cycle, expected_ship_date, production_remark,
                ship_date, ship_method, tracking_no, expected_arrival, notify_warehouse,
                archived
            ) VALUES(
                :material_name, :project, :project_no, :order_factory, :order_quantity,
                :compare_price, :compare_date, :compare_remark,
                :contract_status, :contract_remark,
                :notify_date, :expected_delivery_date, :notify_remark,
                :production_cycle, :expected_ship_date, :production_remark,
                :ship_date, :ship_method, :tracking_no, :expected_arrival, :notify_warehouse,
                :archived
            )
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_packaging(self, oid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE packaging_orders SET
                material_name=:material_name, project=:project,
                project_no=:project_no, order_factory=:order_factory, order_quantity=:order_quantity,
                compare_price=:compare_price, compare_date=:compare_date, compare_remark=:compare_remark,
                contract_status=:contract_status, contract_remark=:contract_remark,
                notify_date=:notify_date, expected_delivery_date=:expected_delivery_date, notify_remark=:notify_remark,
                production_cycle=:production_cycle, expected_ship_date=:expected_ship_date, production_remark=:production_remark,
                ship_date=:ship_date, ship_method=:ship_method, tracking_no=:tracking_no,
                expected_arrival=:expected_arrival, notify_warehouse=:notify_warehouse,
                archived=:archived
            WHERE id=:id
        """, {**data, "id": oid})
        self.conn.commit()

    def get_packaging_factories(self):
        """获取所有不重复的下单厂家（供应商）"""
        c = self.conn.cursor()
        c.execute(
            "SELECT DISTINCT order_factory FROM packaging_orders "
            "WHERE order_factory IS NOT NULL AND order_factory != '' ORDER BY order_factory"
        )
        return [r[0] for r in c.fetchall()]

    def get_packagings(self, archived=0):
        c = self.conn.cursor()
        c.execute("SELECT * FROM packaging_orders WHERE archived=? ORDER BY created_at DESC", (archived,))
        return [dict(r) for r in c.fetchall()]

    def get_packaging(self, oid):
        c = self.conn.cursor()
        c.execute("SELECT * FROM packaging_orders WHERE id=?", (oid,))
        row = c.fetchone()
        return dict(row) if row else None

    def archive_packaging(self, oid):
        c = self.conn.cursor()
        c.execute("UPDATE packaging_orders SET archived=1 WHERE id=?", (oid,))
        self.conn.commit()

    def delete_packaging(self, oid):
        c = self.conn.cursor()
        c.execute("DELETE FROM packaging_orders WHERE id=?", (oid,))
        self.conn.commit()

    # ====== 供应商管理 (v1.2) ======
    def save_supplier(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO suppliers(name,category,main_product,contact_person,phone,wechat,
            cooperation_status,quote_status,sample_status,payment_method,invoice_type,tax_rate,remark)
            VALUES(:name,:category,:main_product,:contact_person,:phone,:wechat,
            :cooperation_status,:quote_status,:sample_status,:payment_method,:invoice_type,:tax_rate,:remark)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_supplier(self, sid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE suppliers SET name=:name,category=:category,main_product=:main_product,
            contact_person=:contact_person,phone=:phone,wechat=:wechat,
            cooperation_status=:cooperation_status,
            quote_status=:quote_status,sample_status=:sample_status,
            payment_method=:payment_method,invoice_type=:invoice_type,
            tax_rate=:tax_rate,remark=:remark
            WHERE id=:id
        """, {**data, "id": sid})
        self.conn.commit()

    def get_suppliers(self, category=None, keyword=None, cooperation_status=None):
        c = self.conn.cursor()
        sql = "SELECT * FROM suppliers WHERE 1=1"
        params = []
        if category and category != "全部":
            sql += " AND category=?"
            params.append(category)
        if keyword:
            sql += " AND name LIKE ?"
            params.append(f"%{keyword}%")
        if cooperation_status and cooperation_status != "全部":
            sql += " AND cooperation_status=?"
            params.append(cooperation_status)
        sql += " ORDER BY id DESC"
        c.execute(sql, params)
        return [dict(r) for r in c.fetchall()]

    def get_supplier(self, sid):
        c = self.conn.cursor()
        c.execute("SELECT * FROM suppliers WHERE id=?", (sid,))
        row = c.fetchone()
        return dict(row) if row else None

    def delete_supplier(self, sid):
        c = self.conn.cursor()
        c.execute("DELETE FROM suppliers WHERE id=?", (sid,))
        self.conn.commit()

    # ====== 催款记录 (v1.2) ======
    def save_collection(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO collection_reminders(supplier_name,contact_person,wechat,
            reminder_date,amount_due,notify_internal,notify_manager,remark)
            VALUES(:supplier_name,:contact_person,:wechat,
            :reminder_date,:amount_due,:notify_internal,:notify_manager,:remark)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_collection(self, cid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE collection_reminders SET supplier_name=:supplier_name,
            contact_person=:contact_person,wechat=:wechat,
            reminder_date=:reminder_date,amount_due=:amount_due,
            notify_internal=:notify_internal,notify_manager=:notify_manager,
            remark=:remark WHERE id=:id
        """, {**data, "id": cid})
        self.conn.commit()

    def get_collections(self, keyword=None, start_date=None, end_date=None):
        c = self.conn.cursor()
        sql = "SELECT * FROM collection_reminders WHERE 1=1"
        params = []
        if keyword:
            sql += " AND (supplier_name LIKE ? OR contact_person LIKE ? OR wechat LIKE ?)"
            kw = f"%{keyword}%"
            params.extend([kw, kw, kw])
        if start_date:
            sql += " AND reminder_date >= ?"
            params.append(start_date)
        if end_date:
            sql += " AND reminder_date <= ?"
            params.append(end_date)
        sql += " ORDER BY reminder_date DESC"
        c.execute(sql, params)
        return [dict(r) for r in c.fetchall()]

    def get_collection(self, cid):
        c = self.conn.cursor()
        c.execute("SELECT * FROM collection_reminders WHERE id=?", (cid,))
        row = c.fetchone()
        return dict(row) if row else None

    def delete_collection(self, cid):
        c = self.conn.cursor()
        c.execute("DELETE FROM collection_reminders WHERE id=?", (cid,))
        self.conn.commit()

    # ====== 报价单产品 (v1.3) ======
    def save_quotation_product(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO quotation_products(item_no, product_name, product_size,
            material_process, supply_cycle, carton_spec, unit)
            VALUES(:item_no, :product_name, :product_size,
            :material_process, :supply_cycle, :carton_spec, :unit)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_quotation_product(self, pid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE quotation_products SET item_no=:item_no, product_name=:product_name,
            product_size=:product_size, material_process=:material_process,
            supply_cycle=:supply_cycle, carton_spec=:carton_spec, unit=:unit
            WHERE id=:id
        """, {**data, "id": pid})
        self.conn.commit()

    def get_quotation_products(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM quotation_products ORDER BY id")
        products = []
        for r in c.fetchall():
            d = dict(r)
            c.execute("SELECT * FROM quotation_tiers WHERE product_id=? ORDER BY min_qty", (d["id"],))
            d["tiers"] = [dict(t) for t in c.fetchall()]
            products.append(d)
        return products

    def get_quotation_product(self, pid):
        c = self.conn.cursor()
        c.execute("SELECT * FROM quotation_products WHERE id=?", (pid,))
        row = c.fetchone()
        if not row:
            return None
        d = dict(row)
        c.execute("SELECT * FROM quotation_tiers WHERE product_id=? ORDER BY min_qty", (pid,))
        d["tiers"] = [dict(t) for t in c.fetchall()]
        return d

    def delete_quotation_product(self, pid):
        c = self.conn.cursor()
        c.execute("DELETE FROM quotation_tiers WHERE product_id=?", (pid,))
        c.execute("DELETE FROM quotation_products WHERE id=?", (pid,))
        self.conn.commit()

    # ====== 报价单阶梯价格 ======
    def save_quotation_tier(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO quotation_tiers(product_id, tier_name, min_qty, max_qty)
            VALUES(:product_id, '', :min_qty, :max_qty)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def delete_quotation_tiers(self, product_id):
        c = self.conn.cursor()
        c.execute("DELETE FROM quotation_tiers WHERE product_id=?", (product_id,))
        self.conn.commit()

    # ====== 报价单配置（需方信息） ======
    def get_quotation_config(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM quotation_config WHERE id=1")
        row = c.fetchone()
        return dict(row) if row else {}

    def update_quotation_config(self, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE quotation_config SET buyer_name=:buyer_name,
                buyer_contact=:buyer_contact, buyer_phone=:buyer_phone,
                buyer_address=:buyer_address, payment_terms=:payment_terms,
                transport_method=:transport_method, delivery_docs=:delivery_docs,
                quote_requirement=:quote_requirement,
                quote_template_note=:quote_template_note,
                footer_note=:footer_note
            WHERE id=1
        """, data)
        self.conn.commit()

    # ====== 报价单供方配置 ======
    def get_quotation_supplier(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM quotation_supplier WHERE id=1")
        row = c.fetchone()
        return dict(row) if row else {}

    def update_quotation_supplier(self, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE quotation_supplier SET supplier_name=:supplier_name,
                contact_person=:contact_person, phone=:phone,
                address=:address, quote_date=:quote_date,
                quote_validity=:quote_validity
            WHERE id=1
        """, data)
        self.conn.commit()

    # ====== 合同供应商 (v1.9.0 从 JSON 迁移) ======
    def get_contract_suppliers(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM contract_suppliers ORDER BY id")
        return [dict(r) for r in c.fetchall()]

    def save_contract_supplier(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO contract_suppliers(short_name, full_name, legal_rep,
            address, contact, auth_rep, phone, fax, payment_days,
            payment_method, account_name, bank, account, remark)
            VALUES(:short_name, :full_name, :legal_rep,
            :address, :contact, :auth_rep, :phone, :fax, :payment_days,
            :payment_method, :account_name, :bank, :account, :remark)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_contract_supplier(self, sid, data):
        c = self.conn.cursor()
        data["id"] = sid
        c.execute("""
            UPDATE contract_suppliers SET short_name=:short_name, full_name=:full_name,
            legal_rep=:legal_rep, address=:address, contact=:contact,
            auth_rep=:auth_rep, phone=:phone, fax=:fax,
            payment_days=:payment_days, payment_method=:payment_method,
            account_name=:account_name, bank=:bank, account=:account,
            remark=:remark
            WHERE id=:id
        """, data)
        self.conn.commit()

    def delete_contract_supplier(self, sid):
        c = self.conn.cursor()
        c.execute("DELETE FROM contract_suppliers WHERE id=?", (sid,))
        self.conn.commit()

    # ====== 合同甲方配置 (v1.9.0 从 JSON 迁移) ======
    def get_contract_party_a(self):
        c = self.conn.cursor()
        c.execute("SELECT company_name, legal_rep, address, contact, phone FROM contract_party_a WHERE id=1")
        row = c.fetchone()
        if row:
            return {
                "company_name": row[0], "legal_rep": row[1],
                "address": row[2], "contact": row[3], "phone": row[4],
            }
        return {
            "company_name": "北京同仁堂健康药业（青海）有限公司",
            "legal_rep": "施能文", "address": "",
            "contact": "龙存英", "phone": "13897764859",
        }

    def update_contract_party_a(self, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE contract_party_a SET company_name=:company_name,
            legal_rep=:legal_rep, address=:address,
            contact=:contact, phone=:phone
            WHERE id=1
        """, data)
        self.conn.commit()
