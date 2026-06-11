#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""三方比价页面 V1.9.9 - 厂家名称替换表头、三列价格、确认按钮、预览板块"""

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
from datetime import datetime
import os
import sys
from copy import copy

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import range_boundaries


def _get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", relative_path)


class ThirdPartyPage(ctk.CTkFrame):
    """三方比价主页面 V1.9.9 — 产品信息板块 + 预览板块"""

    def __init__(self, parent, db, C):
        super().__init__(parent, fg_color=C["bg"], corner_radius=0)
        self.db = db
        self.C = C
        self.records = []
        self.editing_oid = None
        self.qp_rows = []  # [(qty_entry, p1_entry, p2_entry, p3_entry, row_frame)]
        self._build_ui()
        self._load_data()

    # ═══════════════════════════════════════════════
    #  UI 构建
    # ═══════════════════════════════════════════════

    def _build_ui(self):
        C = self.C

        # ── 顶部工具栏 ─────────────────────────────
        toolbar = ctk.CTkFrame(self, fg_color="transparent", height=44)
        toolbar.pack(fill="x", padx=24, pady=(12, 4))
        toolbar.pack_propagate(False)

        ctk.CTkLabel(
            toolbar, text="📊  比价",
            font=ctk.CTkFont(family="Microsoft YaHei", size=20, weight="bold"),
            text_color=C["text"],
        ).pack(side="left", pady=6)

        self.stats_label = ctk.CTkLabel(
            toolbar, text="",
            font=ctk.CTkFont(size=12),
            text_color=C.get("text_secondary", "#5D5D5D"),
        )
        self.stats_label.pack(side="right", padx=(0, 12), pady=6)

        ctk.CTkButton(
            toolbar, text="📤 导出Excel", width=120, height=34,
            fg_color=C["success"], hover_color="#7A9A6E",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._export_excel,
        ).pack(side="right", pady=6)

        # ── 外层滚动容器 ───────────────────────────
        self.outer_scroll = ctk.CTkScrollableFrame(
            self, fg_color="transparent", corner_radius=0,
            scrollbar_button_color=C["border"],
            scrollbar_button_hover_color=C.get("sidebar_hover", "#E8DDD0"),
        )
        self.outer_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        # ═══════════════════════════════════════════
        #  产品信息板块
        # ═══════════════════════════════════════════

        # ── 申请时间 + 最终做货供应商 ──────────────
        time_frame = ctk.CTkFrame(self.outer_scroll, fg_color="transparent")
        time_frame.pack(fill="x", padx=24, pady=(4, 6))

        ctk.CTkLabel(time_frame, text="申请时间：", font=ctk.CTkFont(size=13),
                     text_color=C["text"]).pack(side="left")
        self.date_entry = ctk.CTkEntry(time_frame, width=150, height=32,
                                       font=ctk.CTkFont(size=13),
                                       placeholder_text="YYYY-MM-DD")
        self.date_entry.pack(side="left", padx=(8, 0))
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        ctk.CTkLabel(time_frame, text="  最终做货供应商：", font=ctk.CTkFont(size=13),
                     text_color=C["text"]).pack(side="left", padx=(20, 0))
        self.final_supplier_entry = ctk.CTkEntry(time_frame, width=180, height=32,
                                                  font=ctk.CTkFont(size=13),
                                                  placeholder_text="输入供应商名称")
        self.final_supplier_entry.pack(side="left", padx=(8, 0))

        # ── 产品信息卡片 ───────────────────────────
        form_card = ctk.CTkFrame(self.outer_scroll, fg_color=C["card"],
                                 corner_radius=C["radius_card"])
        form_card.pack(fill="x", padx=24, pady=(0, 8))

        ctk.CTkLabel(
            form_card, text="📦 产品信息",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            text_color=C["text"],
        ).pack(anchor="w", padx=16, pady=(12, 6))

        # 品名 + 项目号
        r1 = ctk.CTkFrame(form_card, fg_color="transparent")
        r1.pack(fill="x", padx=16, pady=2)
        ctk.CTkLabel(r1, text="品名*", width=60, anchor="w",
                     font=ctk.CTkFont(size=13)).pack(side="left")
        self.name_entry = ctk.CTkEntry(r1, height=32, font=ctk.CTkFont(size=13),
                                       placeholder_text="输入产品名称")
        self.name_entry.pack(side="left", fill="x", expand=True, padx=(8, 16))
        ctk.CTkLabel(r1, text="项目号", width=50, anchor="w",
                     font=ctk.CTkFont(size=13)).pack(side="left")
        self.item_no_entry = ctk.CTkEntry(r1, width=160, height=32, font=ctk.CTkFont(size=13))
        self.item_no_entry.pack(side="left", padx=(8, 0))

        # 材质结构 + 规格尺寸
        r2 = ctk.CTkFrame(form_card, fg_color="transparent")
        r2.pack(fill="x", padx=16, pady=2)
        ctk.CTkLabel(r2, text="材质结构", width=60, anchor="w",
                     font=ctk.CTkFont(size=13)).pack(side="left")
        self.material_entry = ctk.CTkEntry(r2, height=32, font=ctk.CTkFont(size=13),
                                           placeholder_text="输入材质结构")
        self.material_entry.pack(side="left", fill="x", expand=True, padx=(8, 16))
        ctk.CTkLabel(r2, text="规格尺寸", width=50, anchor="w",
                     font=ctk.CTkFont(size=13)).pack(side="left")
        self.spec_entry = ctk.CTkEntry(r2, width=160, height=32, font=ctk.CTkFont(size=13))
        self.spec_entry.pack(side="left", padx=(8, 0))

        # ── 供应商名称 ─────────────────────────────
        sup_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        sup_frame.pack(fill="x", padx=16, pady=(8, 2))

        ctk.CTkLabel(sup_frame, text="供应商", width=60, anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C.get("primary", "#D4917A")).pack(side="left")
        self.supplier1_entry = ctk.CTkEntry(sup_frame, height=32, font=ctk.CTkFont(size=13),
                                            placeholder_text="供应商1名称")
        self.supplier1_entry.pack(side="left", fill="x", expand=True, padx=(4, 8))

        ctk.CTkLabel(sup_frame, text="供应商", width=60, anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#6B8FA3").pack(side="left")
        self.supplier2_entry = ctk.CTkEntry(sup_frame, height=32, font=ctk.CTkFont(size=13),
                                            placeholder_text="供应商2名称")
        self.supplier2_entry.pack(side="left", fill="x", expand=True, padx=(4, 8))

        ctk.CTkLabel(sup_frame, text="供应商", width=60, anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#9B7EBD").pack(side="left")
        self.supplier3_entry = ctk.CTkEntry(sup_frame, height=32, font=ctk.CTkFont(size=13),
                                            placeholder_text="供应商3名称")
        self.supplier3_entry.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 绑定供应商名称实时更新表头
        self.supplier1_entry.bind("<KeyRelease>", lambda e: self._on_supplier_changed())
        self.supplier2_entry.bind("<KeyRelease>", lambda e: self._on_supplier_changed())
        self.supplier3_entry.bind("<KeyRelease>", lambda e: self._on_supplier_changed())

        # ── 数量 + 三列价格表头 ────────────────────
        qp_header = ctk.CTkFrame(form_card, fg_color="transparent")
        qp_header.pack(fill="x", padx=16, pady=(10, 0))

        ctk.CTkLabel(qp_header, text="数量", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C.get("text_secondary", "#8B7355"),
                     width=80, anchor="w").pack(side="left")
        ctk.CTkLabel(qp_header, text="供应商1价格", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C.get("primary", "#D4917A"),
                     width=100, anchor="center").pack(side="left")
        ctk.CTkLabel(qp_header, text="供应商2价格", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#6B8FA3",
                     width=100, anchor="center").pack(side="left")
        ctk.CTkLabel(qp_header, text="供应商3价格", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#9B7EBD",
                     width=100, anchor="center").pack(side="left")
        ctk.CTkLabel(qp_header, text="", width=44).pack(side="left")  # spacer for delete button

        # ── 数量+价格行容器 ────────────────────────
        self.qp_list_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        self.qp_list_frame.pack(fill="x", padx=16, pady=(4, 4))
        self._add_qp_row()

        # ── 添加数量按钮 + 确认/清空 ────────────────
        action_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        action_frame.pack(fill="x", padx=16, pady=(4, 12))

        ctk.CTkButton(
            action_frame, text="＋ 添加数量", width=110, height=30,
            font=ctk.CTkFont(size=12),
            fg_color="#6B7280", hover_color="#4B5563",
            command=self._add_qp_row,
        ).pack(side="left")

        self.edit_hint_label = ctk.CTkLabel(
            action_frame, text="",
            font=ctk.CTkFont(size=12),
            text_color=C.get("warning", "#E4A36A"),
        )
        self.edit_hint_label.pack(side="left", padx=(16, 0))

        ctk.CTkButton(
            action_frame, text="✕ 清空表单", width=100, height=36,
            fg_color="#9CA3AF", hover_color="#6B7280",
            font=ctk.CTkFont(size=13),
            command=self._clear_form,
        ).pack(side="right", padx=(8, 0))

        self.save_btn = ctk.CTkButton(
            action_frame, text="💾 保存记录", width=100, height=36,
            fg_color=C["primary"], hover_color=C["primary_hover"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._confirm_record,
        )
        self.save_btn.pack(side="right")

        # ═══════════════════════════════════════════
        #  预览板块
        # ═══════════════════════════════════════════

        preview_header = ctk.CTkFrame(self.outer_scroll, fg_color="transparent")
        preview_header.pack(fill="x", padx=24, pady=(8, 4))
        ctk.CTkLabel(
            preview_header, text="📋 比价记录",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            text_color=C["text"],
        ).pack(side="left")

        self._supplier_labels = [ctk.CTkLabel(preview_header, text="", font=ctk.CTkFont(size=11),
                                               text_color=C.get("text_secondary", "#8B7355"))
                                 for _ in range(3)]
        for lbl in self._supplier_labels:
            lbl.pack(side="right", padx=(12, 0))

        z1 = ctk.CTkFrame(self.outer_scroll, fg_color="#E8E2D9", corner_radius=16)
        z1.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        z2 = ctk.CTkFrame(z1, fg_color="#F2F0EB", corner_radius=12)
        z2.pack(fill="both", expand=True, padx=4, pady=4)

        table_frame = ctk.CTkFrame(z2, fg_color=C["card"], corner_radius=C["radius_card"])
        table_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # Treeview
        self._table_columns = ("比价时间", "品名", "项目号", "材质结构", "规格尺寸",
                               "数量", "供应商1", "供应商2", "供应商3", "最终做货供应商", "操作")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TP.Treeview",
                        font=("Microsoft YaHei", 10),
                        rowheight=34,
                        background="#FFFFFF",
                        fieldbackground="#FFFFFF",
                        foreground="#1E293B",
                        borderwidth=0,
                        relief="flat")
        style.configure("TP.Treeview.Heading",
                        font=("Microsoft YaHei", 10, "bold"),
                        background="#F8FAFC",
                        foreground="#475569",
                        relief="flat",
                        borderwidth=0)
        style.map("TP.Treeview",
                  background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])

        tree_wrap = tk.Frame(table_frame, bg="#FFFFFF")
        tree_wrap.pack(fill="both", expand=True, padx=4, pady=4)

        self.tree = ttk.Treeview(
            tree_wrap, style="TP.Treeview",
            columns=self._table_columns, show="headings", height=10, selectmode="browse"
        )

        display_cols = {"比价时间": 95, "品名": 90, "项目号": 65, "材质结构": 85,
                        "规格尺寸": 85, "数量": 80,
                        "供应商1": 95, "供应商2": 95, "供应商3": 95,
                        "最终做货供应商": 120, "操作": 80}
        for col in self._table_columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=display_cols.get(col, 80), minwidth=30, stretch=True, anchor="center")
        self.tree.column("品名", anchor="w")
        self.tree.column("材质结构", anchor="w")
        self.tree.column("规格尺寸", anchor="w")

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.bind("<Double-1>", self._on_row_double_click)
        self.tree.bind("<Button-1>", self._on_single_click)

    # ═══════════════════════════════════════════════
    #  数据加载与表格渲染
    # ═══════════════════════════════════════════════

    def _load_data(self):
        self.records = self.db.get_third_party_records()
        self._render_table()
        total = len(self.records)
        self.stats_label.configure(text=f"共 {total} 条比价记录")

    def _get_supplier_names(self):
        """获取当前表单中的供应商名称（用于表头显示）"""
        s1 = self.supplier1_entry.get().strip() or "供应商1"
        s2 = self.supplier2_entry.get().strip() or "供应商2"
        s3 = self.supplier3_entry.get().strip() or "供应商3"
        return s1, s2, s3

    def _on_supplier_changed(self):
        """供应商名称输入框内容改变时，实时更新预览表头"""
        if not self.tree:
            return
        s1, s2, s3 = self._get_supplier_names()
        cols = list(self._table_columns)
        cols[7], cols[8], cols[9] = s1, s2, s3
        for i, col_name in enumerate(cols):
            self.tree.heading(self._table_columns[i], text=col_name)
        for lbl, name, color in zip(self._supplier_labels,
                                     [s1, s2, s3],
                                     [self.C.get("primary", "#D4917A"), "#6B8FA3", "#9B7EBD"]):
            if name not in ("供应商1", "供应商2", "供应商3"):
                lbl.configure(text=name, text_color=color)
            else:
                lbl.configure(text="")

    def _render_table(self):
        """渲染预览表格，表头中的厂家列名为当前表单中的厂家名称"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 更新列标题和预览区域标签
        s1, s2, s3 = self._get_supplier_names()
        cols = list(self._table_columns)
        cols[7], cols[8], cols[9] = s1, s2, s3
        for i, col in enumerate(cols):
            self.tree.heading(self._table_columns[i], text=col)

        single_supplier = ""
        for lbl, name, color in zip(self._supplier_labels,
                                     [s1, s2, s3],
                                     [self.C.get("primary", "#D4917A"), "#6B8FA3", "#9B7EBD"]):
            if name not in ("供应商1", "供应商2", "供应商3"):
                lbl.configure(text=name, text_color=color)
                if not single_supplier:
                    single_supplier = name
            else:
                lbl.configure(text="")

        for idx, r in enumerate(self.records, 1):
            s1n = r.get("supplier1", "") or "供应商1"
            s2n = r.get("supplier2", "") or "供应商2"
            s3n = r.get("supplier3", "") or "供应商3"
            qty_str = r.get("quantity_tier", "-")
            p1_str = r.get("price1_tier", "") or "-"
            p2_str = r.get("price2_tier", "") or "-"
            p3_str = r.get("price3_tier", "") or "-"

            # 比价时间：取 created_at 前10位 (YYYY-MM-DD)
            created = r.get("created_at", "") or ""
            if created:
                created = created[:10]  # YYYY-MM-DD

            self.tree.insert("", "end", iid=str(r["id"]),
                values=(
                    created,
                    r.get("product_name", "-"),
                    r.get("item_no", "-"),
                    r.get("material_structure", "-"),
                    r.get("spec_size", "-"),
                    qty_str,
                    f"{s1n}\n¥{p1_str}" if p1_str != "-" else s1n,
                    f"{s2n}\n¥{p2_str}" if p2_str != "-" else s2n,
                    f"{s3n}\n¥{p3_str}" if p3_str != "-" else s3n,
                    r.get("final_supplier", "-"),
                    "编辑  删除",
                ))

    # ═══════════════════════════════════════════════
    #  数量+价格多行动态管理
    # ═══════════════════════════════════════════════

    def _add_qp_row(self):
        """添加一行：数量 + 厂家1价格 + 厂家2价格 + 厂家3价格"""
        row_frame = ctk.CTkFrame(self.qp_list_frame, fg_color="transparent")
        row_frame.pack(fill="x", pady=2)

        qty_entry = ctk.CTkEntry(row_frame, width=80, height=30,
                                 font=ctk.CTkFont(size=12),
                                 placeholder_text="数量")
        qty_entry.pack(side="left")

        p1_entry = ctk.CTkEntry(row_frame, width=100, height=30,
                                font=ctk.CTkFont(size=12),
                                placeholder_text="供应商1价格",
                                fg_color=self.C["card"],
                                border_color=self.C.get("primary", "#D4917A"))
        p1_entry.pack(side="left", padx=(4, 0))

        p2_entry = ctk.CTkEntry(row_frame, width=100, height=30,
                                font=ctk.CTkFont(size=12),
                                placeholder_text="供应商2价格",
                                border_color="#6B8FA3")
        p2_entry.pack(side="left", padx=(4, 0))

        p3_entry = ctk.CTkEntry(row_frame, width=100, height=30,
                                font=ctk.CTkFont(size=12),
                                placeholder_text="供应商3价格",
                                border_color="#9B7EBD")
        p3_entry.pack(side="left", padx=(4, 0))

        del_btn = ctk.CTkButton(
            row_frame, text="✕", width=30, height=30,
            fg_color="#B56A6A", hover_color="#A85A5A",
            font=ctk.CTkFont(size=11, weight="bold"),
            command=lambda rf=row_frame: self._remove_qp_row(rf),
        )
        del_btn.pack(side="left", padx=(6, 0))
        self.qp_rows.append((qty_entry, p1_entry, p2_entry, p3_entry, row_frame))

    def _remove_qp_row(self, row_frame):
        if len(self.qp_rows) <= 1:
            return
        self.qp_rows = [r for r in self.qp_rows if r[4] != row_frame]
        row_frame.destroy()

    # ═══════════════════════════════════════════════
    #  确认 / 清空 / 编辑
    # ═══════════════════════════════════════════════

    def _confirm_record(self):
        """确认按钮：保存当前表单并刷新预览（不清空表单）"""
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("错误", "请输入品名")
            return

        quantities = []
        p1s, p2s, p3s = [], [], []
        for qe, pe1, pe2, pe3, _ in self.qp_rows:
            q = qe.get().strip()
            if q:
                quantities.append(q)
                p1s.append(pe1.get().strip())
                p2s.append(pe2.get().strip())
                p3s.append(pe3.get().strip())

        record_data = {
            "product_name": name,
            "item_no": self.item_no_entry.get().strip(),
            "material_structure": self.material_entry.get().strip(),
            "spec_size": self.spec_entry.get().strip(),
            "quantity_tier": ", ".join(quantities),
            "supplier1": self.supplier1_entry.get().strip(),
            "supplier2": self.supplier2_entry.get().strip(),
            "supplier3": self.supplier3_entry.get().strip(),
            "final_supplier": self.final_supplier_entry.get().strip(),
            "price1_tier": ", ".join(p1s),
            "price2_tier": ", ".join(p2s),
            "price3_tier": ", ".join(p3s),
        }

        try:
            if self.editing_oid:
                self.db.update_third_party_record(self.editing_oid, record_data)
                self.edit_hint_label.configure(text=f"✓ 已更新记录 #{self.editing_oid}")
            else:
                oid = self.db.save_third_party_record(record_data)
                self.editing_oid = oid
                self.edit_hint_label.configure(text=f"✓ 已保存记录 #{oid}")

            self._load_data()
        except Exception as e:
            import traceback
            messagebox.showerror("保存失败", f"{e}\n\n{traceback.format_exc()}")

    def _clear_form(self):
        self.editing_oid = None
        self.edit_hint_label.configure(text="")
        self.save_btn.configure(text="💾 保存记录")

        self.name_entry.delete(0, "end")
        self.item_no_entry.delete(0, "end")
        self.material_entry.delete(0, "end")
        self.spec_entry.delete(0, "end")
        self.supplier1_entry.delete(0, "end")
        self.supplier2_entry.delete(0, "end")
        self.supplier3_entry.delete(0, "end")
        self.final_supplier_entry.delete(0, "end")

        for child in list(self.qp_list_frame.winfo_children()):
            child.destroy()
        self.qp_rows.clear()
        self._add_qp_row()

        self._load_data()

    def _fill_form_for_edit(self, oid):
        data = self.db.get_third_party_record(oid)
        if not data:
            return

        self._clear_form()
        self.editing_oid = oid
        self.save_btn.configure(text="💾 保存修改")
        self.edit_hint_label.configure(text=f"🔧 正在编辑记录 #{oid}")

        self.name_entry.insert(0, data.get("product_name", ""))
        self.item_no_entry.insert(0, data.get("item_no", ""))
        self.material_entry.insert(0, data.get("material_structure", ""))
        self.spec_entry.insert(0, data.get("spec_size", ""))
        self.supplier1_entry.insert(0, data.get("supplier1", ""))
        self.supplier2_entry.insert(0, data.get("supplier2", ""))
        self.supplier3_entry.insert(0, data.get("supplier3", ""))
        self.final_supplier_entry.insert(0, data.get("final_supplier", ""))

        # 填充数量+价格行
        for child in list(self.qp_list_frame.winfo_children()):
            child.destroy()
        self.qp_rows.clear()

        qt = data.get("quantity_tier", "")
        p1t = data.get("price1_tier", "")
        p2t = data.get("price2_tier", "")
        p3t = data.get("price3_tier", "")

        q_parts = [p.strip() for p in qt.split(",") if p.strip()] if qt else []
        p1_parts = [p.strip() for p in p1t.split(",") if p.strip()] if p1t else []
        p2_parts = [p.strip() for p in p2t.split(",") if p.strip()] if p2t else []
        p3_parts = [p.strip() for p in p3t.split(",") if p.strip()] if p3t else []

        max_len = max(len(q_parts), len(p1_parts), len(p2_parts), len(p3_parts), 1)
        for i in range(max_len):
            self._add_qp_row()
            if i < len(q_parts):
                self.qp_rows[-1][0].insert(0, q_parts[i])
            if i < len(p1_parts):
                self.qp_rows[-1][1].insert(0, p1_parts[i])
            if i < len(p2_parts):
                self.qp_rows[-1][2].insert(0, p2_parts[i])
            if i < len(p3_parts):
                self.qp_rows[-1][3].insert(0, p3_parts[i])

    # ═══════════════════════════════════════════════
    #  表格交互
    # ═══════════════════════════════════════════════

    def _on_row_double_click(self, event):
        sel = self.tree.selection()
        if sel:
            self._fill_form_for_edit(int(sel[0]))

    def _on_single_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        col_idx = int(col.replace("#", "")) - 1
        oid = int(item)
        if col_idx == 10:  # 操作列（第11列，0-indexed=10）
            bbox = self.tree.bbox(item, col)
            if not bbox:
                return
            x_rel = event.x - bbox[0]
            col_w = bbox[2]
            if x_rel < col_w * 0.5:
                self._fill_form_for_edit(oid)
            else:
                self._delete_record(oid)

    def _delete_record(self, oid):
        rec = next((r for r in self.records if r["id"] == oid), None)
        name = rec.get("product_name", "") if rec else ""
        if messagebox.askyesno("删除确认", f"确定删除比价记录「{name}」？", icon="warning"):
            self.db.delete_third_party_record(oid)
            if self.editing_oid == oid:
                self.editing_oid = None
                self.edit_hint_label.configure(text="")
                self.save_btn.configure(text="💾 保存记录")
            self._load_data()

    # ═══════════════════════════════════════════════
    #  导出Excel（品名_项目号_供应商名_YYYY-MM-DD_比价表）
    # ═══════════════════════════════════════════════

    def _export_excel(self):
        # ── 选中行检查：只导出选中的记录 ──
        sel = self.tree.selection()
        if sel:
            oid = int(sel[0])
            export_records = [r for r in self.records if r["id"] == oid]
            if not export_records:
                messagebox.showinfo("提示", "未找到选中记录")
                return
        else:
            messagebox.showinfo("提示", "请先在表格中点选一条比价记录")
            return

        # 使用选中记录的品名+项目号+供应商名作为默认文件名
        first = export_records[0]
        pname = first.get("product_name", "比价记录")
        pitem = first.get("item_no", "")
        sup_name = first.get("supplier1", "") or first.get("supplier2", "") or first.get("supplier3", "")
        date_str = datetime.now().strftime("%Y-%m-%d")

        parts = [pname]
        if pitem:
            parts.append(pitem)
        if sup_name:
            parts.append(sup_name)
        parts.append(date_str)
        parts.append("比价表")
        default_name = "_".join(parts) + ".xlsx"

        filepath = filedialog.asksaveasfilename(
            title="保存比价表",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            template_path = _get_resource_path("assets/比价表模板.xlsx")
            wb = load_workbook(template_path)

            # ── 0. 清除除模板sheet外的其他表格 ──
            for sn in list(wb.sheetnames):
                if sn != "比价":
                    del wb[sn]
            ws = wb["比价"]

            # 模板字体：数据行 24号宋体，表头 22号加粗宋体
            data_font = Font(name="宋体", size=24)
            header_font = Font(name="宋体", size=22, bold=True)
            date_font = Font(name="宋体", size=22)
            thin = Side(style="thin")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # ── 辅助函数：安全设置单元格值（绕过 MergedCell 只读限制）──
            # openpyxl 的 ws.cell() 可能返回缓存的 MergedCell 对象（即使已 unmerge）。
            # 解决方案：使用坐标字符串赋值法 ws['A1'] = value，
            # 该方法会自动创建/覆盖普通 Cell，不受 MergedCell 缓存影响。
            def _set_cell(ws, row, col, value=None, **props):
                from openpyxl.utils import get_column_letter
                coord = f"{get_column_letter(col)}{row}"
                # 方法1：坐标字符串赋值（核心——绕过 MergedCell 缓存）
                try:
                    ws[coord] = value
                except Exception:
                    # 极端 fallback：强制从 _cells 字典取真实 Cell
                    cell = ws._cells.get((row, col))
                    if cell is not None and not isinstance(cell, type(None).__class__):
                        try:
                            cell.value = value
                        except Exception:
                            pass
                    return
                # 设置样式属性
                if props:
                    cell = ws.cell(row=row, column=col)
                    for k, v in props.items():
                        try:
                            setattr(cell, k, v)
                        except (AttributeError, TypeError, ValueError):
                            pass

            # ════════════════════════════════════════════
            # 核心策略：零行操作（不用 insert_rows/delete_rows）
            #
            # 原因：openpyxl 的 insert_rows/delete_rows 会推移合并单元格，
            #       导致数据区出现 MergedCell（只读），且 ws.cell() 返回缓存对象。
            #       即使 unmerge 后重新获取，缓存仍返回旧 MergedCell。
            #
            # 解决：
            #   1. 只解除数据区合并 A6:A9~E6:E9（5个）
            #   2. 数据行 ≤ 4 时：直接覆写行 6~N，多余行清空（不删除）
            #   3. 数据行 > 4 时：在 row 10 之后 insert（不影响上方任何合并）
            #   4. 所有赋值用 ws['A1'] 坐标法（绕过 MergedCell 缓存）
            # ════════════════════════════════════════════

            # ── 1. 只解除数据区域的合并（A6:A9 ~ E6:E9）──
            DATA_MERGE_RANGES = ["A6:A9", "B6:B9", "C6:C9", "D6:D9", "E6:E9"]
            for mr in DATA_MERGE_RANGES:
                try:
                    ws.unmerge_cells(mr)
                except Exception:
                    pass

            # ── 2. 写入表头信息（固定区域，无合并冲突）──
            s1 = export_records[0].get("supplier1", "") or "厂家1"
            s2 = export_records[0].get("supplier2", "") or "厂家2"
            s3 = export_records[0].get("supplier3", "") or "厂家3"

            for col, val in [(7, s1), (8, s2), (9, s3)]:
                _set_cell(ws, 5, col, val, font=header_font, alignment=center)

            apply_date = self.date_entry.get().strip()
            if apply_date:
                _set_cell(ws, 2, 8, apply_date, font=date_font, alignment=center)

            # ── 3. 准备数据：展开每个记录的每个阶梯 ──
            DATA_START = 6
            TEMPLATE_DATA_ROWS = 4  # 模板默认4条数据行 (rows 6-9)

            expanded = []
            for r in export_records:
                qt = r.get("quantity_tier", "")
                p1 = r.get("price1_tier", "")
                p2 = r.get("price2_tier", "")
                p3 = r.get("price3_tier", "")
                q_parts = [x.strip() for x in qt.split(",") if x.strip()] if qt else [""]
                p1_parts = [x.strip() for x in p1.split(",") if x.strip()] if p1 else []
                p2_parts = [x.strip() for x in p2.split(",") if x.strip()] if p2 else []
                p3_parts = [x.strip() for x in p3.split(",") if x.strip()] if p3 else []
                max_len = max(len(q_parts), len(p1_parts), len(p2_parts), len(p3_parts), 1)
                for i in range(max_len):
                    expanded.append((
                        r,
                        q_parts[i] if i < len(q_parts) else "",
                        p1_parts[i] if i < len(p1_parts) else "",
                        p2_parts[i] if i < len(p2_parts) else "",
                        p3_parts[i] if i < len(p3_parts) else "",
                    ))

            total_data_rows = len(expanded)

            # ── 4. 行数调整（零风险策略）──
            if total_data_rows > TEMPLATE_DATA_ROWS:
                # 需要更多行：在模板数据区域末尾之后插入（row=10）
                # row 10 在所有数据区合并之下，insert 不会影响已有合并
                ws.insert_rows(DATA_START + TEMPLATE_DATA_ROWS, total_data_rows - TEMPLATE_DATA_ROWS)
            # 注意：total_data_rows < TEMPLATE_DATA_ROWS 时不调用 delete_rows！
            # 多余的模板数据行将在步骤 5 中被直接覆写为空。

            # ── 5. 清除并写入数据区域 ──
            # 清除范围：row 6 到 max(实际数据行, 模板行数)
            CLEAR_END = DATA_START + max(total_data_rows, TEMPLATE_DATA_ROWS)
            for r in range(DATA_START, CLEAR_END):
                for c in range(1, 10):
                    _set_cell(ws, r, c, "")

            # 跟踪每个记录的行范围，用于合并产品信息
            record_row_ranges = {}

            exp_idx = 0
            for r in export_records:
                oid = r["id"]
                qt = r.get("quantity_tier", "")
                q_parts = [x.strip() for x in qt.split(",") if x.strip()] if qt else [""]
                tier_count = max(len(q_parts), 1)
                first_row = DATA_START + exp_idx
                last_row = DATA_START + exp_idx + tier_count - 1
                record_row_ranges[oid] = (first_row, last_row)
                exp_idx += tier_count

            # 写入数据
            for idx, (r, qty, price1, price2, price3) in enumerate(expanded):
                row = DATA_START + idx
                ws.row_dimensions[row].height = 36

                _set_cell(ws, row, 1, idx + 1,          font=data_font, alignment=center,   border=border_all)
                _set_cell(ws, row, 2, r.get("product_name", ""), font=data_font, alignment=left_wrap, border=border_all)
                _set_cell(ws, row, 3, r.get("item_no", ""),     font=data_font, alignment=center,   border=border_all)
                _set_cell(ws, row, 4, r.get("material_structure", ""), font=data_font, alignment=left_wrap, border=border_all)
                _set_cell(ws, row, 5, r.get("spec_size", ""),    font=data_font, alignment=left_wrap, border=border_all)
                _set_cell(ws, row, 6, qty,                      font=data_font, alignment=center,   border=border_all)
                _set_cell(ws, row, 7, price1,                   font=data_font, alignment=center,   border=border_all)
                _set_cell(ws, row, 8, price2,                   font=data_font, alignment=center,   border=border_all)
                _set_cell(ws, row, 9, price3,                   font=data_font, alignment=center,   border=border_all)

            # ── 6. 重建产品信息合并单元格（仅多阶梯记录）──
            for oid, (first_row, last_row) in record_row_ranges.items():
                if first_row != last_row:
                    for col_range in ["A", "B", "C", "D", "E"]:
                        ws.merge_cells(f"{col_range}{first_row}:{col_range}{last_row}")
                    # 只修正对齐方式，不覆盖值（merge已保留左上角单元格的值）
                    for c in [1, 2, 3, 4, 5]:
                        cell = ws.cell(row=first_row, column=c)
                        try:
                            cell.alignment = left_wrap if c in [2, 4, 5] else center
                        except Exception:
                            pass

            # ── 7. 最终做货供应商（固定写入模板签名区第12行）──
            # 模板中 D12:F12="最终做货供应商"标签, G12:I12=值域（合并单元格）
            # 注意：无论数据多少行，此区域固定在 row 12
            FS_ROW = 12  # 固定行号，与模板一致
            fs = export_records[0].get("final_supplier", "").strip()
            if not fs:
                fs = self.final_supplier_entry.get().strip()
            if fs:
                _set_cell(ws, FS_ROW, 7, fs, font=data_font, alignment=center)

            # ── 9. 保存 ──
            try:
                wb.save(filepath)
            except PermissionError:
                messagebox.showerror(
                    "文件被占用",
                    f"无法写入文件，可能已被其他程序打开（如 Excel）：\n{filepath}\n\n请先关闭该文件后再试。"
                )
                return
            except Exception as save_err:
                messagebox.showerror("保存失败", f"无法保存文件：\n{save_err}")
                return

            messagebox.showinfo("导出成功", f"比价表已导出到：\n{filepath}")
        except Exception as e:
            import traceback
            messagebox.showerror("导出失败", f"{e}\n\n{traceback.format_exc()}")
