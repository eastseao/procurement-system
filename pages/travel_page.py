#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""差旅报销页面 v1.2 - 新增导入xlsx"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from datetime import datetime, date, timedelta


class TravelPage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.db = db
        self.C = colors
        self.show_archived = False
        self._build()
        self._load_records()

    def _build(self):
        # 顶部工具栏（去掉内部标题，按钮靠左+去图标）
        header = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0, height=44)
        header.pack(fill="x", padx=20, pady=(12, 8))
        header.pack_propagate(False)

        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="left", padx=0, pady=5)

        self.archive_btn = ctk.CTkButton(
            btn_frame, text="查看归档", width=100, height=34,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=14), command=self._toggle_archive, corner_radius=8, border_width=1, border_color="#8B7D6B")
        self.archive_btn.pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="新增差旅", width=100, height=34,
            fg_color="transparent", hover_color="#A85A5A",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._open_form, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="导出Excel", width=100, height=34,
            fg_color="transparent", hover_color="#7A9A6E",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._export_xlsx, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)

        # ── v1.2 新增：导入xlsx 按钮 ──
        ctk.CTkButton(
            btn_frame, text="导入xlsx", width=100, height=34,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=14),
            command=self._import_xlsx, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)

        # 统计栏
        self.stats_frame = ctk.CTkFrame(self, fg_color=self.C["card"],
                                         corner_radius=self.C["radius_card"], height=72)
        self.stats_frame.pack(fill="x", padx=16, pady=(12, 0))
        self.stats_frame.pack_propagate(False)
        self.stat_labels = {}
        for key, label, color in [
            ("total_count", "出差次数", self.C["primary"]),
            ("total_days", "累计出差天数", self.C["warning"]),
            ("total_amount", "报销总金额 (¥)", self.C["warning"]),
            ("unreimbursed", "未报销金额 (¥)", self.C["danger"]),
        ]:
            f = ctk.CTkFrame(self.stats_frame, fg_color="transparent")
            f.pack(side="left", padx=24, pady=8)
            ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=12),
                         text_color=self.C["text_secondary"]).pack()
            lbl = ctk.CTkLabel(f, text="0", font=ctk.CTkFont(size=20, weight="bold"),
                               text_color=color)
            lbl.pack()
            self.stat_labels[key] = lbl

        # 筛选栏
        filter_frame = ctk.CTkFrame(self, fg_color=self.C["card"],
                                     corner_radius=self.C["radius_card"], height=52)
        filter_frame.pack(fill="x", padx=16, pady=8)
        filter_frame.pack_propagate(False)

        ctk.CTkLabel(filter_frame, text="报销状态:",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left", padx=(16, 4), pady=12)
        self.filter_reimburse = ctk.CTkComboBox(
            filter_frame, values=["全部", "未报销", "已报销"], width=100, height=28,
            command=lambda _: self._load_records())
        self.filter_reimburse.set("全部")
        self.filter_reimburse.pack(side="left", padx=4)

        ctk.CTkButton(filter_frame, text="↺ 刷新", width=60, height=28,
                      fg_color=self.C["border"],
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=13),
                      command=self._load_records, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=8)

        # 列表
        # 表格区域 - 阴影分层：z1 → z2 → z3
        # z1（底层）：fg_color=#E8E2D9, corner_radius=16, 内边距 4
        z1_frame = ctk.CTkFrame(self, fg_color="transparent", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        # z2（中层）：fg_color=#F2F0EB, corner_radius=8, 内边距 4
        z2_frame = ctk.CTkFrame(z1_frame, fg_color="transparent", corner_radius=8, border_width=1, border_color="#8B7D6B")
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # z3（上层卡片）：fg_color=C["card"], corner_radius=C["radius_card"]
        list_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        list_frame.pack(fill="both", expand=True, padx=4, pady=4)

        cols = [
            ("start_date", "出发日期", 95),
            ("reason", "出差事由", 160),
            ("destination", "目的地", 100),
            ("duration", "天数", 55),
            ("handler", "出差人", 70),
            ("transport_summary", "交通摘要", 180),
            ("hotel_summary", "住宿摘要", 120),
            ("total", "合计(¥)", 90),
            ("reimburse", "报销", 70),
            ("action", "操作", 140),
        ]

        style = ttk.Style()
        style.configure("Travel.Treeview", font=("Microsoft YaHei", 9),
                         rowheight=38, background="#FFFFFF", fieldbackground="#FFFFFF",
                         foreground="#1E293B")
        style.configure("Travel.Treeview.Heading", font=("Microsoft YaHei", 9, "bold"),
                         background="#F8FAFC", foreground="#475569", relief="flat")
        style.map("Travel.Treeview", background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])

        tree_frame = tk.Frame(list_frame, bg="#FFFFFF")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(tree_frame, style="Travel.Treeview",
                                  columns=[c[0] for c in cols], show="headings",
                                  selectmode="browse")
        for cid, label, width in cols:
            self.tree.heading(cid, text=label)
            self.tree.column(cid, width=width, minwidth=40, stretch=True,
                              anchor="w" if cid in ("reason", "transport_summary", "hotel_summary") else "center")

        vsb = ctk.CTkScrollbar(tree_frame, orientation="vertical", command=self.tree.yview, button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<Button-1>", self._on_click)

        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.tag_configure("unreimbursed", foreground="#DC2626")
        self.tree.tag_configure("archived", foreground="#94A3B8")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)

        # ── 右键菜单 ──
        self._context_menu = tk.Menu(self, tearoff=0,
            bg="#FFFFFF", fg="#4A3728",
            activebackground="#E8D5C4", activeforeground="#4A3728",
            font=("Microsoft YaHei", 10),
        )
        self._context_menu.add_command(label="📝 编辑", command=self._on_edit_selected)
        self._context_menu.add_command(label="📋 复制", command=self._on_copy_selected)
        self._context_menu.add_separator()
        self._context_menu.add_command(label="🗑️ 删除", command=self._on_delete_selected)
        self.tree.bind("<Button-3>", self._on_tree_right_click)

    def _load_records(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        records = self.db.get_travels(archived=1 if self.show_archived else 0)
        ri_filter = self.filter_reimburse.get()

        total_amount = 0
        unreimbursed = 0
        total_days = 0

        for r in records:
            if ri_filter != "全部" and r["reimbursement_status"] != ri_filter:
                continue

            total_amount += r["total"]
            if r["reimbursement_status"] == "未报销":
                unreimbursed += r["total"]
            total_days += r.get("duration") or 0

            t_summary = "；".join(
                f"{t.get('transport_type','')}({t.get('departure','')}→{t.get('destination','')})"
                for t in r["transports"]
            ) or "—"
            if len(t_summary) > 25:
                t_summary = t_summary[:25] + "…"

            h_summary = f"{len(r['hotels'])}晚" if r["hotels"] else "—"

            tag = "unreimbursed" if r["reimbursement_status"] == "未报销" else ""
            if self.show_archived:
                tag = "archived"

            self.tree.insert("", "end", iid=str(r["id"]),
                             values=(
                                 r["start_date"], r["reason"], r["destination"],
                                 f"{r.get('duration', 0)}天",
                                 r.get("handler", "—"),
                                 t_summary, h_summary,
                                 f"{r['total']:.2f}",
                                 r["reimbursement_status"],
                                 "编辑  归档  删除" if not self.show_archived else "查看  删除"
                             ), tags=(tag,))

        count = len(self.tree.get_children())
        self.stat_labels["total_count"].configure(text=str(count))
        self.stat_labels["total_days"].configure(text=str(total_days))
        self.stat_labels["total_amount"].configure(text=f"{total_amount:.2f}")
        self.stat_labels["unreimbursed"].configure(text=f"{unreimbursed:.2f}")

    def _toggle_archive(self):
        self.show_archived = not self.show_archived
        if self.show_archived:
            self.archive_btn.configure(text="返回列表", fg_color=self.C["success"])
        else:
            self.archive_btn.configure(text="查看归档", fg_color="transparent")
        self._load_records()

    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        tid = int(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == 9:
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                if not self.show_archived:
                    if x_rel < col_w * 0.33:
                        self._open_form(tid)
                    elif x_rel < col_w * 0.66:
                        self._archive_record(tid)
                    else:
                        self._delete_record(tid)
                else:
                    if x_rel < col_w * 0.5:
                        self._open_form(tid, view_only=True)
                    else:
                        self._delete_record(tid)
            else:
                # bbox 为 None（行不可见），直接打开表单
                self._open_form(tid)
        else:
            self._open_form(tid)

    def _archive_record(self, tid):
        if messagebox.askyesno("归档确认", "确定将此差旅记录归档？"):
            self.db.archive_travel(tid)
            self._load_records()

    def _delete_record(self, tid):
        if messagebox.askyesno("删除确认", "确定删除此差旅记录？", icon="warning"):
            self.db.delete_travel(tid)
            self._load_records()

    def _open_form(self, tid=None, view_only=False):
        try:
            form = TravelForm(self, self.db, self.C, tid=tid,
                              view_only=view_only,
                              on_save=self._load_records)
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    # ── 导出 Excel ─────────────────────────────────
    def _export_xlsx(self):
        records = self.db.get_travels(archived=1 if self.show_archived else 0)
        if not records:
            messagebox.showinfo("提示", "没有可导出的数据")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出差旅报销", defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"差旅报销_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not filepath:
            return

        headers = ["出差事由", "目的地", "出差人", "开始日期", "结束日期", "出差天数",
                    "类型", "明细日期/入住日期", "出发地/退房日期", "目的地/房间数",
                    "金额", "开票状态", "报销状态", "备注"]
        rows = []
        for r in records:
            base = {
                "出差事由": r.get("reason", ""),
                "目的地": r.get("destination", ""),
                "出差人": r.get("handler", ""),
                "开始日期": r.get("start_date", ""),
                "结束日期": r.get("end_date", ""),
                "出差天数": r.get("duration", ""),
                "报销状态": r.get("reimbursement_status", ""),
                "备注": r.get("remark", ""),
            }
            transports = r.get("transports") or []
            hotels = r.get("hotels") or []

            if transports or hotels:
                for t in transports:
                    rows.append({**base, "类型": f"交通-{t.get('transport_type','')}",
                                 "明细日期/入住日期": t.get("travel_date", ""),
                                 "出发地/退房日期": t.get("departure", ""),
                                 "目的地/房间数": t.get("destination", ""),
                                 "金额": t.get("amount", ""),
                                 "开票状态": ""})
                for h in hotels:
                    rows.append({**base, "类型": "酒店住宿",
                                 "明细日期/入住日期": h.get("checkin_date", ""),
                                 "出发地/退房日期": h.get("checkout_date", ""),
                                 "目的地/房间数": h.get("room_count", ""),
                                 "金额": h.get("amount", ""),
                                 "开票状态": h.get("invoice_status", "")})
            else:
                rows.append({**base, "类型": "", "明细日期/入住日期": "", "出发地/退房日期": "",
                             "目的地/房间数": "", "金额": "", "开票状态": ""})

        try:
            self.db.export_to_xlsx(filepath, "差旅报销", headers, rows,
                                   col_widths=[16, 12, 8, 12, 12, 8, 10, 14, 14, 12, 10, 10, 10, 16])
            messagebox.showinfo("导出成功", f"已导出 {len(rows)} 条明细到：\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── v1.2 新增：导入xlsx ────────────────────────────
    def _import_xlsx(self):
        filepath = filedialog.askopenfilename(
            title="选择差旅报销xlsx文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            col_map = {}
            field_candidates = {
                "reason":               ["出差事由", "事由", "reason"],
                "destination":          ["目的地", "destination"],
                "handler":              ["出差人", "经手人", "handler"],
                "start_date":           ["开始日期", "出发日期", "start_date"],
                "end_date":             ["结束日期", "返回日期", "end_date"],
                "duration":             ["出差天数", "天数", "duration"],
                "reimbursement_status": ["报销状态", "reimbursement_status"],
                "invoice_status":       ["开票状态", "invoice_status"],
                "remark":               ["备注", "remark"],
                "type":                 ["类型", "type"],
                "date1":                ["明细日期/入住日期", "明细日期", "入住日期", "乘坐日期"],
                "departure":            ["出发地/退房日期", "出发地", "departure"],
                "dest2":                ["目的地/房间数", "目的地", "房间数"],
                "amount":               ["金额", "amount"],
                "invoice2":             ["开票状态"],
            }
            for idx, h in enumerate(headers):
                for field, candidates in field_candidates.items():
                    if h in candidates and field not in col_map:
                        col_map[field] = idx
                        break

            if "reason" not in col_map and "start_date" not in col_map:
                messagebox.showerror("导入失败", "未找到必要列（出差事由/开始日期），请检查xlsx格式")
                return

            def _cell(row_vals, key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row_vals):
                    return ""
                v = row_vals[idx]
                return str(v).strip() if v is not None else ""

            # 按出差事由+出发日期分组
            groups = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                reason = _cell(row, "reason")
                start = _cell(row, "start_date")
                if not reason and not start:
                    continue

                gkey = (reason, start)
                if gkey not in groups:
                    groups[gkey] = {
                        "reason": reason,
                        "destination": _cell(row, "destination"),
                        "handler": _cell(row, "handler") or "纪委",
                        "start_date": start,
                        "end_date": _cell(row, "end_date") or start,
                        "duration": int(_cell(row, "duration") or 1),
                        "reimbursement_status": _cell(row, "reimbursement_status") or "未报销",
                        "invoice_status": _cell(row, "invoice_status") or "未开票",
                        "remark": _cell(row, "remark"),
                        "transports": [],
                        "hotels": [],
                    }

                rtype = _cell(row, "type")
                amt_str = _cell(row, "amount")
                try:
                    amt = float(amt_str.replace("¥", "").replace(",", "")) if amt_str else 0
                except (ValueError, AttributeError):
                    amt = 0

                if "酒店" in rtype or "住宿" in rtype:
                    groups[gkey]["hotels"].append({
                        "checkin_date": _cell(row, "date1"),
                        "checkout_date": _cell(row, "departure"),
                        "room_count": 1,
                        "amount": amt,
                        "invoice_status": _cell(row, "invoice2") or "未开票",
                    })
                elif rtype:
                    tp = rtype.replace("交通-", "").strip() or "其他"
                    groups[gkey]["transports"].append({
                        "transport_type": tp,
                        "travel_date": _cell(row, "date1"),
                        "departure": _cell(row, "departure"),
                        "destination": _cell(row, "dest2"),
                        "amount": amt,
                    })

            imported = 0
            for gdata in groups.values():
                data = {k: v for k, v in gdata.items() if k not in ("transports", "hotels")}
                self.db.save_travel(data, gdata["transports"], gdata["hotels"])
                imported += 1

            self._load_records()
            messagebox.showinfo("导入成功", f"成功导入 {imported} 条差旅记录")

        except Exception as e:
            import traceback
            messagebox.showerror("导入失败", f"{e}\n\n{traceback.format_exc()}")



    # ── 单击选中高亮行 ─────────────────────
    def _on_click(self, event):
        """单击选中行，高亮显示"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

    # ── 右键菜单回调 ──────────────────────────────
    def _on_tree_right_click(self, event):
        """右键点击表格行→弹出菜单"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self._context_menu.post(event.x_root, event.y_root)

    def _on_edit_selected(self):
        """编辑选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        tid = int(selection[0])
        self._open_form(tid, view_only=self.show_archived)

    def _on_copy_selected(self):
        """复制选中行信息到剪贴板"""
        selection = self.tree.selection()
        if not selection:
            return
        tid = int(selection[0])
        records = self.db.get_travels(archived=1 if self.show_archived else 0)
        rec = next((r for r in records if r["id"] == tid), None)
        if rec:
            text = f"出差事由:{rec.get('reason','')} 目的地:{rec.get('destination','')} 合计:¥{rec.get('total',0):.2f}"
            self.clipboard_clear()
            self.clipboard_append(text)

    def _on_delete_selected(self):
        """删除选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        tid = int(selection[0])
        self._delete_record(tid)

    def _on_hover(self, event):
        """鼠标悬停高亮行"""
        item = self.tree.identify_row(event.y)
        for i in self.tree.get_children():
            tags = list(self.tree.item(i, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(i, tags=tuple(tags))
        if item:
            tags = list(self.tree.item(item, "tags"))
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tuple(tags))

    def _on_leave(self, event):
        """鼠标离开时清除悬停高亮"""
        for i in self.tree.get_children():
            tags = list(self.tree.item(i, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(i, tags=tuple(tags))

# ============================
# 差旅表单弹窗
# ============================
class TravelForm(ctk.CTkToplevel):
    def __init__(self, parent, db, colors, tid=None, view_only=False, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = colors
        self.tid = tid
        self.view_only = view_only
        self.on_save = on_save
        self.transport_rows = []
        self.hotel_rows = []

        title = "查看差旅记录" if view_only else ("编辑差旅记录" if tid else "新增差旅记录")
        self.title(title)
        self.geometry("960x750")
        self.resizable(True, True)
        self.configure(fg_color=self.C["bg"])

        self.update_idletasks()
        pw, ph = parent.winfo_rootx(), parent.winfo_rooty()
        self.geometry(f"960x750+{max(0,pw+60)}+{max(0,ph+30)}")

        self._build()
        if tid:
            self._load_data(tid)

    def _build(self):
        outer = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"])
        outer.pack(fill="both", expand=True)

        card = ctk.CTkFrame(outer, fg_color=self.C["card"], corner_radius=self.C["radius_modal"])
        card.pack(fill="both", expand=True, padx=16, pady=16)

        self._section(card, "🗺️ 基础信息")

        row0 = ctk.CTkFrame(card, fg_color="transparent")
        row0.pack(fill="x", padx=16, pady=(0, 4))

        ctk.CTkLabel(row0, text="出差事由 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=0, column=0, padx=4, pady=4, sticky="w")
        self.reason_var = tk.StringVar()
        ctk.CTkEntry(row0, textvariable=self.reason_var, width=220, height=32,
                     placeholder_text="如：参加供应商洽谈").grid(row=0, column=1, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(row0, text="目的地 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=0, column=2, padx=(16, 4), pady=4, sticky="w")
        self.destination_var = tk.StringVar()
        ctk.CTkEntry(row0, textvariable=self.destination_var, width=150, height=32,
                     placeholder_text="如：上海").grid(row=0, column=3, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(row0, text="出差人", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=0, column=4, padx=(16, 4), pady=4, sticky="w")
        self.handler_var = tk.StringVar(value="纪委")
        ctk.CTkEntry(row0, textvariable=self.handler_var, width=100, height=32).grid(
            row=0, column=5, padx=4, pady=4, sticky="w")

        row1 = ctk.CTkFrame(card, fg_color="transparent")
        row1.pack(fill="x", padx=16, pady=(0, 4))

        ctk.CTkLabel(row1, text="出发日期 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 4))
        self.start_date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        ctk.CTkEntry(row1, textvariable=self.start_date_var, width=130, height=32,
                     placeholder_text="YYYY-MM-DD").pack(side="left", padx=4)
        self.start_date_var.trace_add("write", self._calc_duration)

        ctk.CTkLabel(row1, text="  返回日期 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(8, 4))
        self.end_date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        ctk.CTkEntry(row1, textvariable=self.end_date_var, width=130, height=32,
                     placeholder_text="YYYY-MM-DD").pack(side="left", padx=4)
        self.end_date_var.trace_add("write", self._calc_duration)

        ctk.CTkLabel(row1, text="  出差时长：", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(12, 0))
        self.duration_label = ctk.CTkLabel(row1, text="0 天",
                                            font=ctk.CTkFont(size=14, weight="bold"),
                                            text_color=self.C["primary"])
        self.duration_label.pack(side="left", padx=4)

        row2 = ctk.CTkFrame(card, fg_color="transparent")
        row2.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(row2, text="报销状态:", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 4))
        self.reimburse_var = tk.StringVar(value="未报销")
        for v in ["未报销", "已报销"]:
            ctk.CTkRadioButton(row2, text=v, variable=self.reimburse_var, value=v,
                               font=ctk.CTkFont(size=13)).pack(side="left", padx=8)

        ctk.CTkLabel(row2, text="  开票状态:", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(16, 4))
        self.invoice_var = tk.StringVar(value="未开票")
        for v in ["未开票", "已开票"]:
            ctk.CTkRadioButton(row2, text=v, variable=self.invoice_var, value=v,
                               font=ctk.CTkFont(size=13)).pack(side="left", padx=8)

        self._section(card, "🚌 交通出行明细")

        t_header = ctk.CTkFrame(card, fg_color="transparent", corner_radius=6)
        t_header.pack(fill="x", padx=16, pady=(0, 2))
        for txt, w in [("交通工具", 110), ("乘坐日期", 110), ("出发地", 110),
                        ("目的地", 110), ("金额(¥)", 90), ("", 36)]:
            ctk.CTkLabel(t_header, text=txt, width=w,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=self.C["text_secondary"]).pack(side="left", padx=2, pady=5)

        self.transport_container = ctk.CTkFrame(card, fg_color="transparent")
        self.transport_container.pack(fill="x", padx=16)
        self._add_transport_row()

        ctk.CTkButton(card, text="＋ 添加交通行程", height=28, width=130,
                      fg_color="transparent",
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=12),
                      command=self._add_transport_row, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(anchor="w", padx=16, pady=4)

        self._section(card, "🏨 酒店住宿明细")

        h_header = ctk.CTkFrame(card, fg_color="transparent", corner_radius=6)
        h_header.pack(fill="x", padx=16, pady=(0, 2))
        for txt, w in [("入住日期", 110), ("退房日期", 110), ("房间数", 70),
                        ("金额(¥)", 90), ("开票状态", 100), ("", 36)]:
            ctk.CTkLabel(h_header, text=txt, width=w,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=self.C["text_secondary"]).pack(side="left", padx=2, pady=5)

        self.hotel_container = ctk.CTkFrame(card, fg_color="transparent")
        self.hotel_container.pack(fill="x", padx=16)
        self._add_hotel_row()

        ctk.CTkButton(card, text="＋ 添加住宿记录", height=28, width=130,
                      fg_color="transparent",
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=12),
                      command=self._add_hotel_row, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(anchor="w", padx=16, pady=4)

        total_row = ctk.CTkFrame(card, fg_color="transparent", corner_radius=8, border_width=1, border_color="#8B7D6B")
        total_row.pack(fill="x", padx=16, pady=8)
        ctk.CTkLabel(total_row, text="差旅总费用：",
                     font=ctk.CTkFont(size=14), text_color=self.C["text_secondary"]).pack(
            side="left", padx=16, pady=8)
        self.total_label = ctk.CTkLabel(total_row, text="¥ 0.00",
                                         font=ctk.CTkFont(size=19, weight="bold"),
                                         text_color=self.C["success"])
        self.total_label.pack(side="left")

        self._section(card, "📝 备注")
        self.remark_text = ctk.CTkTextbox(card, height=70, font=ctk.CTkFont(size=13))
        self.remark_text.pack(fill="x", padx=16, pady=(0, 12))

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=12)
        if not self.view_only:
            ctk.CTkButton(btn_row, text="✓ 保存", width=120, height=38,
                          fg_color="transparent", hover_color=self.C["primary_hover"],
                          font=ctk.CTkFont(size=16, weight="bold"),
                          command=self._save, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="✕ 关闭", width=80, height=38,
                      fg_color="transparent", hover_color="#4B5563",
                      font=ctk.CTkFont(size=14), command=self.destroy, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)

    def _section(self, parent, title):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=self.C["primary"]).pack(side="left")
        ctk.CTkFrame(f, height=1, fg_color=self.C["border"]).pack(
            side="left", fill="x", expand=True, padx=(8, 0))

    def _calc_duration(self, *args):
        try:
            s = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d").date()
            e = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d").date()
            days = max(0, (e - s).days + 1)
            self.duration_label.configure(text=f"{days} 天")
        except ValueError:
            self.duration_label.configure(text="—")

    def _add_transport_row(self, data=None):
        row = ctk.CTkFrame(self.transport_container, fg_color="transparent")
        row.pack(fill="x", pady=2)
        vars_ = {}

        transport_type_var = tk.StringVar(value=data.get("transport_type", "") if data else "")
        vars_["transport_type"] = transport_type_var
        ctk.CTkComboBox(row, values=["飞机", "高铁", "动车", "火车", "汽车", "轮船", "出租车", "地铁", "其他"],
                         variable=transport_type_var, width=110, height=28).pack(side="left", padx=2)

        for fkey, w, ph in [("travel_date", 110, "YYYY-MM-DD"),
                              ("departure", 110, "出发地"),
                              ("destination", 110, "目的地"),
                              ("amount", 90, "金额")]:
            v = tk.StringVar(value=str(data.get(fkey, "")) if data else "")
            vars_[fkey] = v
            ctk.CTkEntry(row, textvariable=v, width=w, height=28,
                         placeholder_text=ph, font=ctk.CTkFont(size=12)).pack(side="left", padx=2)

        ctk.CTkButton(row, text="✕", width=32, height=28,
                      fg_color="transparent",
                      hover_color="#FECACA", font=ctk.CTkFont(size=12),
                      command=lambda r=row: self._remove_transport(r), corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=2)
        self.transport_rows.append({"frame": row, "vars": vars_})

    def _remove_transport(self, frame):
        self.transport_rows = [r for r in self.transport_rows if r["frame"] != frame]
        frame.destroy()
        self._update_total()

    def _add_hotel_row(self, data=None):
        row = ctk.CTkFrame(self.hotel_container, fg_color="transparent")
        row.pack(fill="x", pady=2)
        vars_ = {}

        for fkey, w, ph in [("checkin_date", 110, "入住日期"),
                              ("checkout_date", 110, "退房日期"),
                              ("room_count", 70, "房间数"),
                              ("amount", 90, "金额")]:
            v = tk.StringVar(value=str(data.get(fkey, "")) if data else "")
            vars_[fkey] = v
            ctk.CTkEntry(row, textvariable=v, width=w, height=28,
                         placeholder_text=ph, font=ctk.CTkFont(size=12)).pack(side="left", padx=2)
            if fkey == "amount":
                v.trace_add("write", lambda *a: self._update_total())

        inv_var = tk.StringVar(value=data.get("invoice_status", "未开票") if data else "未开票")
        vars_["invoice_status"] = inv_var
        ctk.CTkComboBox(row, values=["未开票", "已开票"], variable=inv_var,
                         width=100, height=28).pack(side="left", padx=2)

        ctk.CTkButton(row, text="✕", width=32, height=28,
                      fg_color="transparent",
                      hover_color="#FECACA", font=ctk.CTkFont(size=12),
                      command=lambda r=row: self._remove_hotel(r), corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=2)
        self.hotel_rows.append({"frame": row, "vars": vars_})

    def _remove_hotel(self, frame):
        self.hotel_rows = [r for r in self.hotel_rows if r["frame"] != frame]
        frame.destroy()
        self._update_total()

    def _update_total(self):
        total = 0
        for r in self.transport_rows:
            try:
                total += float(r["vars"]["amount"].get() or 0)
            except ValueError:
                pass
        for r in self.hotel_rows:
            try:
                total += float(r["vars"]["amount"].get() or 0)
            except ValueError:
                pass
        self.total_label.configure(text=f"¥ {total:.2f}")

    def _load_data(self, tid):
        all_r = self.db.get_travels() + self.db.get_travels(archived=1)
        r = next((x for x in all_r if x["id"] == tid), None)
        if not r:
            return
        self.reason_var.set(r["reason"])
        self.destination_var.set(r["destination"])
        self.handler_var.set(r.get("handler", "纪委"))
        self.start_date_var.set(r["start_date"])
        self.end_date_var.set(r["end_date"])
        self.reimburse_var.set(r["reimbursement_status"])
        self.invoice_var.set(r["invoice_status"])
        self.remark_text.delete("1.0", "end")
        if r.get("remark"):
            self.remark_text.insert("1.0", r["remark"])
        for row_data in self.transport_rows:
            row_data["frame"].destroy()
        self.transport_rows.clear()
        for t in r["transports"]:
            self._add_transport_row(t)
        if not r["transports"]:
            self._add_transport_row()
        for row_data in self.hotel_rows:
            row_data["frame"].destroy()
        self.hotel_rows.clear()
        for h in r["hotels"]:
            self._add_hotel_row(h)
        if not r["hotels"]:
            self._add_hotel_row()
        self._calc_duration()
        self._update_total()

    def _save(self):
        reason = self.reason_var.get().strip()
        destination = self.destination_var.get().strip()
        start = self.start_date_var.get().strip()
        end = self.end_date_var.get().strip()
        if not reason:
            messagebox.showerror("验证失败", "请填写出差事由", parent=self)
            return
        if not destination:
            messagebox.showerror("验证失败", "请填写目的地", parent=self)
            return
        try:
            s = datetime.strptime(start, "%Y-%m-%d").date()
            e = datetime.strptime(end, "%Y-%m-%d").date()
            duration = max(0, (e - s).days + 1)
        except ValueError:
            messagebox.showerror("验证失败", "日期格式不正确，请使用 YYYY-MM-DD", parent=self)
            return

        transports = []
        for r in self.transport_rows:
            v = r["vars"]
            tp = v["transport_type"].get().strip()
            if not tp:
                continue
            try:
                amt = float(v["amount"].get() or 0)
            except ValueError:
                amt = 0
            transports.append({
                "transport_type": tp,
                "travel_date": v["travel_date"].get().strip(),
                "departure": v["departure"].get().strip(),
                "destination": v["destination"].get().strip(),
                "amount": amt,
            })

        hotels = []
        for r in self.hotel_rows:
            v = r["vars"]
            checkin = v["checkin_date"].get().strip()
            if not checkin:
                continue
            try:
                amt = float(v["amount"].get() or 0)
                rooms = int(v["room_count"].get() or 1)
            except ValueError:
                amt = 0
                rooms = 1
            hotels.append({
                "checkin_date": checkin,
                "checkout_date": v["checkout_date"].get().strip(),
                "room_count": rooms,
                "amount": amt,
                "invoice_status": v["invoice_status"].get(),
            })

        data = {
            "reason": reason,
            "destination": destination,
            "start_date": start,
            "end_date": end,
            "duration": duration,
            "handler": self.handler_var.get().strip() or "纪委",
            "invoice_status": self.invoice_var.get(),
            "reimbursement_status": self.reimburse_var.get(),
            "remark": self.remark_text.get("1.0", "end").strip(),
        }

        if self.tid:
            self.db.update_travel(self.tid, data, transports, hotels)
            messagebox.showinfo("成功", "差旅记录已更新", parent=self)
        else:
            self.db.save_travel(data, transports, hotels)
            messagebox.showinfo("成功", "差旅记录已保存", parent=self)

        if self.on_save:
            self.on_save()
        self.destroy()