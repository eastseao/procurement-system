#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""物料下单页面 v1.5 - 配色统一为莫兰迪暖色"""

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
from datetime import datetime, date, timedelta
import re, os, tempfile, shutil, zipfile
from lxml import etree


# 配色统一使用主窗口传入的 self.C（莫兰迪暖色）
# 不再使用本地的 CLOSORS 字典


class PackagingPage(ctk.CTkFrame):
    """物料下单主页面"""

    def __init__(self, parent, db, C):
        super().__init__(parent, fg_color=C["bg"], corner_radius=0)
        self.db = db
        self.C = C
        self.show_archived = False
        self.records = []
        self._loading = False  # 防止 _load_data 递归
        try:
            self._build_ui()
        except Exception as e:
            import traceback
            # 显示 _build_ui 错误
            err_label = ctk.CTkLabel(self, text=f"UI构建错误:\n{e}\n\n{traceback.format_exc()}",
                                     text_color="red", justify="left",
                                     font=ctk.CTkFont(family="Consolas", size=12))
            err_label.pack(padx=20, pady=20, fill="both", expand=True)
            return
        try:
            self._load_data()
        except Exception as e:
            import traceback
            self.stats_label.configure(text=f"数据加载错误: {e}")
            traceback.print_exc()

    # ── UI 构建 ───────────────────────────────────────
    def _build_ui(self):
        # 顶部工具栏：去掉内部标题，按钮靠左排列并去掉按钮内图标
        toolbar = ctk.CTkFrame(self, fg_color="transparent", height=44)
        toolbar.pack(fill="x", padx=20, pady=(12, 8))
        toolbar.pack_propagate(False)

        # 按钮组（靠左排列，去掉按钮内图标）
        btn_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        btn_frame.pack(side="left", pady=5)

        ctk.CTkButton(
            btn_frame, text="归档", width=100, height=34,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=14), command=self._toggle_archive, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # ── v1.2 新增：导出xlsx 按钮 ──
        ctk.CTkButton(
            btn_frame, text="导出", width=100, height=34,
            fg_color="transparent", hover_color="#7A9A6E",
            font=ctk.CTkFont(size=14),
            command=self._export_xlsx, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="新增", width=100, height=34,
            fg_color="transparent", hover_color="#A85A5A",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._open_form, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # ── 上传合同按钮 (v1.9.2 新增) ──
        ctk.CTkButton(
            btn_frame, text="上传", width=100, height=34,
            fg_color="transparent", hover_color="#3A7BC4",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._upload_contract, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # ── 筛选栏 ──
        filter_bar = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"], height=44)
        filter_bar.pack(fill="x", padx=24, pady=(0, 8))
        filter_bar.pack_propagate(False)

        # 项目筛选
        f1 = ctk.CTkFrame(filter_bar, fg_color="transparent")
        f1.pack(side="left", padx=(16, 12), pady=6)
        ctk.CTkLabel(f1, text="项目：", font=ctk.CTkFont(family="Segoe UI", size=14),
                      text_color=self.C["text"]).pack(side="left")
        self.project_var = ctk.StringVar(value="全部")
        self.project_combo = ctk.CTkComboBox(
            f1, variable=self.project_var,
            values=["全部"], width=130, height=32,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            command=lambda _: self._load_data(),
        )
        self.project_combo.pack(side="left", padx=(4, 0))

        # 供应商筛选（文本输入框，实时模糊匹配）
        f2 = ctk.CTkFrame(filter_bar, fg_color="transparent")
        f2.pack(side="left", padx=(0, 12), pady=6)
        ctk.CTkLabel(f2, text="供应商：", font=ctk.CTkFont(family="Segoe UI", size=14),
                      text_color=self.C["text"]).pack(side="left")
        self.factory_filter_entry = ctk.CTkEntry(
            f2, width=130, height=32,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            placeholder_text="输入供应商",
        )
        self.factory_filter_entry.pack(side="left", padx=(4, 0))
        self.factory_filter_entry.bind("<KeyRelease>", lambda e: self._load_data())

        # 项目号筛选
        f3 = ctk.CTkFrame(filter_bar, fg_color="transparent")
        f3.pack(side="left", padx=(0, 8), pady=6)
        ctk.CTkLabel(f3, text="项目号：", font=ctk.CTkFont(family="Segoe UI", size=14),
                      text_color=self.C["text"]).pack(side="left")
        self.project_no_filter_entry = ctk.CTkEntry(
            f3, width=110, height=32,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            placeholder_text="8位数字",
        )
        self.project_no_filter_entry.pack(side="left", padx=(4, 0))
        self.project_no_filter_entry.bind("<KeyRelease>", self._on_project_no_key)
        self.project_no_filter_entry.bind("<Return>", lambda e: self._load_data())

        # 清除筛选按钮
        ctk.CTkButton(
            filter_bar, text="✕ 清除筛选", width=90, height=30,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=13),
            command=self._clear_filters, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right", padx=(0, 12), pady=6)

        # 统计栏
        stats = ctk.CTkFrame(self, fg_color="transparent", height=28)
        stats.pack(fill="x", padx=24, pady=(0, 4))

        self.stats_label = ctk.CTkLabel(
            stats, text="", font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=self.C.get("text_secondary", "#5D5D5D"),
        )
        self.stats_label.pack(side="left")

        # 表格区域 - 阴影分层：z1 → z2 → z3
        # z1（底层）：fg_color=#E8E2D9, corner_radius=16, 内边距 4
        z1_frame = ctk.CTkFrame(self, fg_color="transparent", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        # z2（中层）：fg_color=#F2F0EB, corner_radius=8, 内边距 4
        z2_frame = ctk.CTkFrame(z1_frame, fg_color="transparent", corner_radius=8, border_width=1, border_color="#8B7D6B")
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # z3（上层卡片）：fg_color=C["card"], corner_radius=C["radius_card"]
        table_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        table_frame.pack(fill="both", expand=True, padx=4, pady=4)

        columns = ("物料名称", "项目号", "下单数量", "下单厂家", "所属项目", "比价单价", "合同状态", "通知日期", "预计发货", "发货日期", "预计到货", "状态", "操作")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Packaging.Treeview",
                         font=("Microsoft YaHei", 9),
                         rowheight=36,
                         background="#FFFFFF",
                         fieldbackground="#FFFFFF",
                         foreground="#1E293B",
                         borderwidth=0,
                         relief="flat")
        style.configure("Packaging.Treeview.Heading",
                         font=("Microsoft YaHei", 9, "bold"),
                         background="#F8FAFC",
                         foreground="#475569",
                         relief="flat",
                         borderwidth=0)
        style.map("Packaging.Treeview",
                  background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])
        style.layout("Packaging.Treeview", [
            ("Treeview.treearea", {"sticky": "nswe"})
        ])

        tree_wrap = tk.Frame(table_frame, bg="#FFFFFF")
        tree_wrap.pack(fill="both", expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(
            tree_wrap, style="Packaging.Treeview",
            columns=columns, show="headings", height=15, selectmode="browse"
        )

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, minwidth=40, stretch=True, anchor="center")
        self.tree.column("物料名称", width=140, minwidth=60, stretch=True, anchor="w")
        self.tree.column("项目号", width=80, minwidth=40, stretch=True)
        self.tree.column("下单数量", width=80, minwidth=40, stretch=True)
        self.tree.column("下单厂家", width=100, minwidth=40, stretch=True)
        self.tree.column("所属项目", width=100, minwidth=40, stretch=True)
        self.tree.column("比价单价", width=80, minwidth=40, stretch=True)
        self.tree.column("合同状态", width=80, minwidth=40, stretch=True)
        self.tree.column("通知日期", width=90, minwidth=40, stretch=True)
        self.tree.column("预计发货", width=90, minwidth=40, stretch=True)
        self.tree.column("发货日期", width=90, minwidth=40, stretch=True)
        self.tree.column("预计到货", width=90, minwidth=40, stretch=True)
        self.tree.column("状态", width=70, minwidth=40, stretch=True)
        self.tree.column("操作", width=90, minwidth=50, stretch=True, anchor="center")

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.tag_configure("odd", background="#F8FAFC")
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)

        self.tree.bind("<Double-1>", self._on_row_double_click)
        self.tree.bind("<Button-1>", self._on_row_click)

        # ── 右键菜单 ──
        self._context_menu = tk.Menu(self, tearoff=0,
            bg="#FFFFFF", fg="#4A3728",
            activebackground="#E8D5C4", activeforeground="#4A3728",
            font=("Microsoft YaHei", 10),
        )
        self._context_menu.add_command(label="📝 编辑", command=self._on_edit_selected)
        self._context_menu.add_command(label="📋 复制", command=self._on_copy_selected)
        self._context_menu.add_separator()
        self._context_menu.add_command(label="🗑️ 删除", command=self._on_delete_selected)
        self.tree.bind("<Button-3>", self._on_tree_right_click)

    # ── 数据加载 ────────────────────────────────────────
    def _load_data(self):
        if self._loading:
            return
        self._loading = True
        try:
            self.records = self.db.get_packagings(archived=1 if self.show_archived else 0)

            # 项目筛选
            project_filter = self.project_var.get()
            if project_filter != "全部":
                self.records = [r for r in self.records if r.get("project") == project_filter]

            # 供应商筛选（文本模糊匹配）
            factory_filter = self.factory_filter_entry.get().strip()
            if factory_filter:
                self.records = [r for r in self.records if factory_filter in (r.get("order_factory") or "")]

            # 项目号筛选（8位数字前缀匹配）
            project_no_filter = self.project_no_filter_entry.get().strip()
            if project_no_filter:
                self.records = [r for r in self.records if (r.get("project_no") or "").startswith(project_no_filter)]

            self._refresh_project_filter()
            self._render_table()

            # 统计信息
            total = len(self.records)
            filter_parts = []
            if project_filter != "全部":
                filter_parts.append(f"项目={project_filter}")
            if factory_filter:
                filter_parts.append(f"供应商={factory_filter}")
            if project_no_filter:
                filter_parts.append(f"项目号={project_no_filter}")

            status_text = f"共 {total} 条记录 | 当前显示：{'已归档' if self.show_archived else '进行中'}"
            if filter_parts:
                status_text += f" | 筛选：{', '.join(filter_parts)}"
            self.stats_label.configure(text=status_text)
        finally:
            self._loading = False

    def _refresh_project_filter(self):
        projects = self.db.get_projects()
        values = ["全部"] + projects
        current = self.project_var.get()
        # 临时移除 command 回调，防止 configure(values=...) 触发 _load_data 导致递归
        self.project_combo.configure(values=values, command=None)
        if current not in values:
            self.project_var.set("全部")
        # 恢复 command 回调
        self.project_combo.configure(command=lambda _: self._load_data())

    def _on_project_no_key(self, event):
        """项目号筛选：只允许输入数字，实时筛选"""
        val = self.project_no_filter_entry.get()
        # 移除非数字字符，限制最多8位
        if val and not val.isdigit():
            cleaned = ''.join(c for c in val if c.isdigit())[:8]
            self.project_no_filter_entry.delete(0, "end")
            self.project_no_filter_entry.insert(0, cleaned)
            return
        if len(val) > 8:
            self.project_no_filter_entry.delete(0, "end")
            self.project_no_filter_entry.insert(0, val[:8])
            return
        self._load_data()

    def _clear_filters(self):
        """清除所有筛选条件"""
        self.project_var.set("全部")
        self.factory_filter_entry.delete(0, "end")
        self.project_no_filter_entry.delete(0, "end")
        self._load_data()

    def _render_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        today_str = date.today().strftime("%Y-%m-%d")
        tomorrow = date.today() + timedelta(days=1)
        tomorrow_str = tomorrow.strftime("%Y-%m-%d")

        for r in self.records:
            compare_price = r.get("compare_price")
            price_str = f"¥{compare_price:.2f}" if compare_price else "-"
            contract_status = r.get("contract_status") or "-"
            notify_date = r.get("notify_date") or "-"
            expected_ship = r.get("expected_ship_date") or "-"
            ship_date = r.get("ship_date") or "-"
            expected_arrival = r.get("expected_arrival") or "-"

            # 状态判断
            if r.get("archived"):
                status = "已归档"
            elif r.get("ship_date"):
                status = "已发货"
            elif r.get("contract_status") == "已签订":
                status = "已签合同"
            elif r.get("contract_status") in ("待签批", "已邮寄"):
                status = "合同处理中"
            elif r.get("compare_price"):
                status = "已比价"
            else:
                status = "待比价"

            tags = []
            if r.get("archived"):
                tags.append("archived")

            arr_date = r.get("expected_arrival", "")
            if arr_date == today_str:
                tags.append("arrival_today")
            elif arr_date == tomorrow_str:
                tags.append("arrival_tomorrow")

            tag_tuple = tuple(tags) if tags else ()

            self.tree.insert(
                "", "end", iid=str(r["id"]),
                values=(
                    r.get("material_name", ""),
                    r.get("project_no", ""),
                    r.get("order_quantity", ""),
                    r.get("order_factory", ""),
                    r.get("project", ""),
                    price_str,
                    contract_status,
                    notify_date,
                    expected_ship,
                    ship_date,
                    expected_arrival,
                    status,
                    "删除  归档" if not r.get("archived") else "删除",
                ),
                tags=tag_tuple,
            )

        self.tree.tag_configure("archived", foreground="#9CA3AF")
        self.tree.tag_configure("arrival_today", background="#FFF3E0", foreground="#C9A96E")
        self.tree.tag_configure("arrival_tomorrow", background="#FFF8F0", foreground="#8B5E3C")

    # ── 事件处理 ────────────────────────────────────────
    def _on_row_double_click(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        oid = int(selection[0])
        self._open_form(oid=oid)

    def _on_tree_right_click(self, event):
        """右键点击表格行→弹出菜单"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self._context_menu.post(event.x_root, event.y_root)

    def _on_edit_selected(self):
        """编辑选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        oid = int(selection[0])
        self._open_form(oid=oid)

    def _on_copy_selected(self):
        """复制选中行信息到剪贴板"""
        import tkinter as tk
        selection = self.tree.selection()
        if not selection:
            return
        oid = int(selection[0])
        rec = None
        for r in self.records:
            if r["id"] == oid:
                rec = r
                break
        if rec:
            text = f"物料:{rec.get('material_name','')} 厂家:{rec.get('order_factory','')} 数量:{rec.get('order_quantity','')}"
            self.clipboard_clear()
            self.clipboard_append(text)
            try:
                self.tree.event_generate("<<TreeviewSelect>>")
            except Exception:
                pass

    def _on_delete_selected(self):
        """删除选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        oid = int(selection[0])
        rec = None
        for r in self.records:
            if r["id"] == oid:
                rec = r
                break
        if rec:
            self._delete_order(oid, rec)

    def _on_row_click(self, event):
        """操作列点击：删除/归档"""
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        # 先选中当前行
        self.tree.selection_set(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx != 12:  # 操作列是第13列（0-indexed: 12）= 13列中的最后一列
            return
        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x_rel = event.x - bbox[0]
        col_w = bbox[2]
        oid = int(item)
        rec = None
        for r in self.records:
            if r["id"] == oid:
                rec = r
                break
        if rec and not rec.get("archived") and x_rel < col_w * 0.5:
            # 左半部分 = 删除
            self._delete_order(oid, rec)
        elif rec and not rec.get("archived") and x_rel >= col_w * 0.5:
            # 右半部分 = 归档
            self._archive_order(oid, rec)
        elif rec and rec.get("archived") and x_rel < col_w * 0.5:
            # 已归档的只有删除
            self._delete_order(oid, rec)

    def _on_hover(self, event):
        """鼠标悬浮时高亮行背景"""
        item = self.tree.identify_row(event.y)
        if item and item != getattr(self, "_last_hover", None):
            # 清除上一个悬浮行的高亮
            if hasattr(self, "_last_hover") and self._last_hover:
                tags = list(self.tree.item(self._last_hover, "tags")[0])
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(self._last_hover, tags=tags)
            # 高亮当前行
            tags = list(self.tree.item(item, "tags")[0])
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tags)
            self._last_hover = item
        elif not item and hasattr(self, "_last_hover") and self._last_hover:
            # 鼠标离开表格区域，清除高亮
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_leave(self, event):
        """鼠标离开表格时清除悬浮高亮"""
        if hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _delete_order(self, oid, rec):
        name = rec.get("material_name", "")
        if messagebox.askyesno("删除确认", f"确定删除「{name}」？删除后不可恢复！", icon="warning"):
            self.db.delete_packaging(oid)
            self._load_data()

    def _archive_order(self, oid, rec):
        name = rec.get("material_name", "")
        if messagebox.askyesno("归档确认", f"确定归档「{name}」？", icon="question"):
            self.db.archive_packaging(oid)
            self._load_data()

    def _open_form(self, oid=None):
        try:
            form = PackagingForm(
                self, self.db, self.C, oid=oid,
                on_save=self._load_data,
            )
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    # ── v1.9.2 新增：上传合同自动生成记录 ──
    def _upload_contract(self):
        """上传合同并自动生成物料下单记录"""
        filepath = filedialog.askopenfilename(
            title="选择合同文件",
            filetypes=[("Word文档", "*.docx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            data = self._parse_contract(filepath)
            if not data or not data.get("material_name"):
                messagebox.showerror(
                    "解析失败",
                    "无法从合同中提取有效信息\n请确保上传的是正确的合同文件"
                )
                return

            # 确保项目存在
            project = data.get("project", "默认项目")
            try:
                self.db.add_project(project)
            except Exception:
                pass

            # 合同状态默认为"待签批"
            data["contract_status"] = "待签批"
            data["project"] = project

            # 补全数据库所需的所有字段默认值
            _defaults = {
                "compare_date": "",
                "compare_remark": "",
                "contract_remark": "",
                "notify_date": "",
                "expected_delivery_date": "",
                "notify_remark": "",
                "production_cycle": "",
                "expected_ship_date": "",
                "production_remark": "",
                "ship_date": "",
                "ship_method": "",
                "tracking_no": "",
                "expected_arrival": "",
                "notify_warehouse": "",
                "archived": 0,
            }
            for _k, _v in _defaults.items():
                data.setdefault(_k, _v)

            # 保存记录
            self.db.save_packaging(data)

            # 刷新表格
            self._load_data()

            material_name = data.get("material_name", "")
            messagebox.showinfo(
                "成功",
                f"已从合同生成物料下单记录：\n{material_name}"
            )
        except Exception as e:
            import traceback
            messagebox.showerror("上传失败", f"错误：\n{str(e)}")
            traceback.print_exc()

    def _parse_contract(self, filepath):
        """解析合同 .docx，提取物料信息"""
        W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'

        tmp_dir = tempfile.mkdtemp(prefix="contract_parse_")
        try:
            # 解压 .docx
            with zipfile.ZipFile(filepath, 'r') as zf:
                zf.extractall(tmp_dir)

            # 解析 document.xml
            doc_path = os.path.join(tmp_dir, 'word', 'document.xml')
            if not os.path.exists(doc_path):
                return None

            tree = etree.parse(doc_path)
            root = tree.getroot()

            data = {}

            # ── 1. 查找产品表格，提取物料信息 ──
            tables = root.findall('.//' + W + 'tbl')

            for tbl in tables:
                rows = tbl.findall('.//' + W + 'tr')
                if len(rows) < 2:
                    continue

                # 检查表头
                header_row = rows[0]
                headers = []
                for cell in header_row.findall('.//' + W + 'tc'):
                    cell_text = ''.join(
                        t.text or '' for t in cell.findall('.//' + W + 't'))
                    headers.append(cell_text.strip())

                header_str = ' '.join(headers)
                # 检查是否是产品表格
                if ('物料名称' in header_str or
                        '物料项目号' in header_str or
                        '产品名称' in header_str):
                    # 提取数据行（第一行数据）
                    for row_idx in range(1, len(rows)):
                        data_row = rows[row_idx]
                        cells = data_row.findall('.//' + W + 'tc')

                        row_data = {}
                        for i, cell in enumerate(cells):
                            cell_text = ''.join(
                                t.text or '' for t in cell.findall('.//' + W + 't'))
                            cell_text = cell_text.strip()

                            if i < len(headers):
                                header = headers[i]
                                if ('物料名称' in header or
                                        '产品名称' in header):
                                    row_data['material_name'] = cell_text
                                elif '项目号' in header:
                                    row_data['project_no'] = cell_text
                                elif '数量' in header:
                                    row_data['order_quantity'] = cell_text
                                elif ('单价' in header and
                                      '金额' not in header):
                                    price_match = re.search(r'[\d.]+',
                                                                  cell_text)
                                    if price_match:
                                        try:
                                            row_data['compare_price'] = float(
                                                price_match.group())
                                        except ValueError:
                                            pass

                        # 只提取第一行有效数据
                        if row_data.get('material_name'):
                            data.update(row_data)
                            break

                    # 找到产品表格后就退出
                    if data.get('material_name'):
                        break

            # ── 2. 提取乙方（厂家）名称 ──
            # 遍历所有段落，查找"乙  方："后面的文本
            all_paras = root.findall('.//' + W + 'p')
            factory_name = ""
            for para in all_paras:
                para_text = ''.join(
                    t.text or '' for t in para.findall('.//' + W + 't'))

                # 查找乙方段落
                if ('乙' in para_text and '方' in para_text and
                        '甲' not in para_text):
                    # 获取段落后面的段落（通常乙方名称在下一个段落）
                    # 或者在本段落后面查找
                    # 简化的方法：取段落中 "：" 后面的文本
                    if '：' in para_text:
                        after_colon = para_text.split('：', 1)[1].strip()
                        if after_colon and len(after_colon) > 1:
                            factory_name = after_colon
                            break

            # 如果上面没找到，尝试另一种方法
            if not factory_name:
                full_text = ''.join(
                    t.text or '' for t in root.iter(W + 't'))
                # 查找 "乙  方：" 模式
                party_b_match = re.search(
                    r'乙\s*方\s*[：:]\s*([^\n\s].{2,50}?)\s*(?:地|联|电|传)',
                    full_text)
                if party_b_match:
                    factory_name = party_b_match.group(1).strip()

            if factory_name:
                data['order_factory'] = factory_name

            # ── 3. 设置默认值 ──
            if 'project' not in data:
                data['project'] = "默认项目"

            return data if data.get('material_name') else None

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def _toggle_archive(self):
        self.show_archived = not getattr(self, "show_archived", False)
        if hasattr(self, 'archive_btn') and self.archive_btn:
            self.archive_btn.configure(
                text="查看进行中" if self.show_archived else "归档"
            )
        self._load_data()

    def set_show_archived(self, val):
        self.show_archived = val

    # ── v1.2 新增：导出xlsx ─────────────────────────────
    def _export_xlsx(self):
        if not self.records:
            messagebox.showinfo("提示", "没有可导出的数据")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出物料下单", defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"物料下单_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not filepath:
            return

        headers = ["物料名称", "项目号", "下单数量", "下单厂家", "所属项目", "比价单价", "比价日期", "比价备注",
                   "合同状态", "合同备注", "通知日期", "沟通货期", "通知备注",
                   "生产周期", "预定发货日期", "生产备注",
                   "发货日期", "发货方式", "物流单号", "预计到货", "通知库房", "状态"]
        rows = []
        for r in self.records:
            compare_price = r.get("compare_price")
            if r.get("archived"):
                status = "已归档"
            elif r.get("ship_date"):
                status = "已发货"
            elif r.get("contract_status") == "已签订":
                status = "已签合同"
            elif r.get("contract_status") in ("待签批", "已邮寄"):
                status = "合同处理中"
            elif compare_price:
                status = "已比价"
            else:
                status = "待比价"

            rows.append({
                "物料名称": r.get("material_name", ""),
                "项目号": r.get("project_no", ""),
                "下单数量": r.get("order_quantity", ""),
                "下单厂家": r.get("order_factory", ""),
                "所属项目": r.get("project", ""),
                "比价单价": f"{compare_price:.2f}" if compare_price else "",
                "比价日期": r.get("compare_date", ""),
                "比价备注": r.get("compare_remark", ""),
                "合同状态": r.get("contract_status", ""),
                "合同备注": r.get("contract_remark", ""),
                "通知日期": r.get("notify_date", ""),
                "沟通货期": r.get("expected_delivery_date", ""),
                "通知备注": r.get("notify_remark", ""),
                "生产周期": r.get("production_cycle", ""),
                "预定发货日期": r.get("expected_ship_date", ""),
                "生产备注": r.get("production_remark", ""),
                "发货日期": r.get("ship_date", ""),
                "发货方式": r.get("ship_method", ""),
                "物流单号": r.get("tracking_no", ""),
                "预计到货": r.get("expected_arrival", ""),
                "通知库房": "是" if r.get("notify_warehouse") else "否",
                "状态": status,
            })

        try:
            self.db.export_to_xlsx(filepath, "物料下单", headers, rows,
                                   col_widths=[16, 8, 10, 12, 10, 12, 16, 10, 16, 12, 12, 16,
                                               10, 12, 16, 12, 10, 14, 12, 10, 10])
            messagebox.showinfo("导出成功", f"已导出 {len(rows)} 条记录到：\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── v1.2 新增：导入xlsx ─────────────────────────────
    def _import_xlsx(self):
        filepath = filedialog.askopenfilename(
            title="选择物料下单xlsx文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            # 列名映射（支持中英文）
            col_map = {}
            for idx, h in enumerate(headers):
                if h in ("物料名称", "material_name"):
                    col_map["material_name"] = idx
                elif h in ("所属项目", "project"):
                    col_map["project"] = idx
                elif h in ("项目号", "project_no"):
                    col_map["project_no"] = idx
                elif h in ("下单厂家", "order_factory"):
                    col_map["order_factory"] = idx
                elif h in ("比价单价", "compare_price"):
                    col_map["compare_price"] = idx
                elif h in ("比价日期", "compare_date"):
                    col_map["compare_date"] = idx
                elif h in ("比价备注", "compare_remark"):
                    col_map["compare_remark"] = idx
                elif h in ("合同状态", "contract_status"):
                    col_map["contract_status"] = idx
                elif h in ("合同备注", "contract_remark"):
                    col_map["contract_remark"] = idx
                elif h in ("通知日期", "notify_date"):
                    col_map["notify_date"] = idx
                elif h in ("沟通货期", "expected_delivery_date"):
                    col_map["expected_delivery_date"] = idx
                elif h in ("通知备注", "notify_remark"):
                    col_map["notify_remark"] = idx
                elif h in ("生产周期", "production_cycle"):
                    col_map["production_cycle"] = idx
                elif h in ("预定发货日期", "expected_ship_date"):
                    col_map["expected_ship_date"] = idx
                elif h in ("生产备注", "production_remark"):
                    col_map["production_remark"] = idx
                elif h in ("发货日期", "ship_date"):
                    col_map["ship_date"] = idx
                elif h in ("发货方式", "ship_method"):
                    col_map["ship_method"] = idx
                elif h in ("物流单号", "tracking_no"):
                    col_map["tracking_no"] = idx
                elif h in ("预计到货", "expected_arrival"):
                    col_map["expected_arrival"] = idx

            if "material_name" not in col_map:
                messagebox.showerror("导入失败", "未找到「物料名称」列，请检查xlsx格式")
                return

            def _cell(row_vals, key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row_vals):
                    return ""
                v = row_vals[idx]
                return str(v).strip() if v is not None else ""

            imported = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue

                name = _cell(row, "material_name")
                if not name:
                    skipped += 1
                    continue

                project = _cell(row, "project")
                if not project:
                    project = "默认项目"
                    # 确保项目存在
                try:
                    self.db.add_project(project)
                except Exception:
                    pass

                try:
                    price_str = _cell(row, "compare_price")
                    compare_price = float(price_str.replace("¥", "").replace(",", "")) if price_str else None
                except (ValueError, AttributeError):
                    compare_price = None

                data = {
                    "material_name": name,
                    "project": project,
                    "project_no": _cell(row, "project_no"),
                    "order_quantity": _cell(row, "order_quantity"),
                    "order_factory": _cell(row, "order_factory"),
                    "compare_price": compare_price or "",
                    "compare_date": _cell(row, "compare_date"),
                    "compare_remark": _cell(row, "compare_remark"),
                    "contract_status": _cell(row, "contract_status"),
                    "contract_remark": _cell(row, "contract_remark"),
                    "notify_date": _cell(row, "notify_date"),
                    "expected_delivery_date": _cell(row, "expected_delivery_date"),
                    "notify_remark": _cell(row, "notify_remark"),
                    "production_cycle": _cell(row, "production_cycle"),
                    "expected_ship_date": _cell(row, "expected_ship_date"),
                    "production_remark": _cell(row, "production_remark"),
                    "ship_date": _cell(row, "ship_date"),
                    "ship_method": _cell(row, "ship_method"),
                    "tracking_no": _cell(row, "tracking_no"),
                    "expected_arrival": _cell(row, "expected_arrival"),
                    "notify_warehouse": 0,
                    "archived": 0,
                }
                self.db.save_packaging(data)
                imported += 1

            self._load_data()
            msg = f"成功导入 {imported} 条物料记录"
            if skipped:
                msg += f"\n（跳过 {skipped} 条空白行）"
            messagebox.showinfo("导入成功", msg)

        except Exception as e:
            import traceback
            messagebox.showerror("导入失败", f"{e}\n\n{traceback.format_exc()}")


# ============================
# 物料下单表单弹窗
# ============================

class PackagingForm(ctk.CTkToplevel):
    """物料下单表单弹窗"""

    CONTRACT_STATUS = ["待签批", "已签订", "已邮寄"]
    SHIP_METHODS = ["快递", "物流", "自提", "其他"]

    def __init__(self, parent, db, C, oid=None, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = C
        self.oid = oid
        self.on_save = on_save
        self.title("编辑物料订单" if oid else "新增物料订单")
        self.geometry("820x750")
        self.configure(fg_color=self.C["bg"])

        self._build_ui()
        if oid:
            self._load_data()

    def _build_ui(self):
        canvas = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"], corner_radius=0)
        canvas.pack(fill="both", expand=True, padx=16, pady=(12, 8))

        # ── 基本信息 ──────────────────────────────────
        info_frame = ctk.CTkFrame(canvas, fg_color=self.C["card"], corner_radius=self.C["radius_modal"])
        info_frame.pack(fill="x", pady=(0, 12), padx=4)

        ctk.CTkLabel(
            info_frame, text="📋 基本信息",
            font=ctk.CTkFont(family="Segoe UI", size=17, weight="bold"),
            text_color=self.C["text"],
        ).pack(anchor="w", padx=16, pady=(12, 8))

        # 物料名称
        row1 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(row1, text="物料名称 *", width=90, anchor="w",
                      font=ctk.CTkFont(family="Segoe UI", size=14)).pack(side="left")
        self.name_entry = ctk.CTkEntry(
            row1, height=34, font=ctk.CTkFont(size=14),
            placeholder_text="请输入物料名称",
        )
        self.name_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # 项目号
        row1b = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1b.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(row1b, text="项目号", width=90, anchor="w",
                      font=ctk.CTkFont(family="Segoe UI", size=14)).pack(side="left")
        self.project_no_entry = ctk.CTkEntry(
            row1b, height=34, font=ctk.CTkFont(size=14),
            placeholder_text="请输入项目号",
        )
        self.project_no_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # 下单数量
        row1b2 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1b2.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(row1b2, text="下单数量", width=90, anchor="w",
                      font=ctk.CTkFont(family="Segoe UI", size=14)).pack(side="left")
        self.order_quantity_entry = ctk.CTkEntry(
            row1b2, height=34, font=ctk.CTkFont(size=14),
            placeholder_text="请输入数量",
        )
        self.order_quantity_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # 下单厂家
        row1c = ctk.CTkFrame(info_frame, fg_color="transparent")
        row1c.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(row1c, text="下单厂家", width=90, anchor="w",
                      font=ctk.CTkFont(family="Segoe UI", size=14)).pack(side="left")
        self.order_factory_entry = ctk.CTkEntry(
            row1c, height=34, font=ctk.CTkFont(size=14),
            placeholder_text="请输入下单厂家",
        )
        self.order_factory_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # 所属项目
        row2 = ctk.CTkFrame(info_frame, fg_color="transparent")
        row2.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(row2, text="所属项目 *", width=90, anchor="w",
                      font=ctk.CTkFont(family="Segoe UI", size=14)).pack(side="left")
        self.project_var = ctk.StringVar()
        projects = self.db.get_projects()
        self.project_combo = ctk.CTkComboBox(
            row2, variable=self.project_var,
            values=projects, width=200, height=34,
            font=ctk.CTkFont(family="Segoe UI", size=14),
        )
        self.project_combo.pack(side="left", padx=(8, 4))
        ctk.CTkButton(
            row2, text="＋", width=36, height=34,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=16, weight="bold"),
            command=self._add_project, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left")

        # ── 五个环节 ─────────────────────────────────
        self._build_section(canvas, "① 物料比价", self._build_compare)
        self._build_section(canvas, "② 合同跟踪", self._build_contract)
        self._build_section(canvas, "③ 通知厂家", self._build_notify)
        self._build_section(canvas, "④ 生产进度", self._build_production)
        self._build_section(canvas, "⑤ 发货跟进", self._build_ship)

        # ── 底部按钮 ─────────────────────────────────
        btn_frame = ctk.CTkFrame(canvas, fg_color="transparent")
        btn_frame.pack(fill="x", pady=16, padx=4)

        self.arrival_btn = ctk.CTkButton(
            btn_frame, text="✅ 确认到货并归档", width=150, height=38,
            fg_color="transparent", hover_color="#7A9A6E",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._confirm_arrival, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",)

        ctk.CTkButton(
            btn_frame, text="💾 保存", width=100, height=38,
            fg_color="transparent", hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._save, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right", padx=8)

        ctk.CTkButton(
            btn_frame, text="取消", width=80, height=38,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=14),
            command=self.destroy, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right", padx=8)

    # ── 环节板块构建辅助 ─────────────────────────────
    def _build_section(self, parent, title, builder):
        frame = ctk.CTkFrame(parent, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        frame.pack(fill="x", pady=(0, 12), padx=4)

        header = ctk.CTkFrame(frame, fg_color="transparent", height=40)
        header.pack(fill="x", padx=16, pady=(8, 4))
        header.pack_propagate(False)

        self._section_visible = getattr(self, "_section_visible", {})
        section_key = title
        if section_key not in self._section_visible:
            self._section_visible[section_key] = True

        toggle_btn = ctk.CTkButton(
            header, text="▼" if self._section_visible[section_key] else "▶",
            width=28, height=28, fg_color="transparent", hover_color=self.C["bg"],
            font=ctk.CTkFont(size=13), corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",)
        toggle_btn.pack(side="left")

        ctk.CTkLabel(
            header, text=title,
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left", padx=(4, 0))

        content = ctk.CTkFrame(frame, fg_color="transparent")
        content.pack(fill="x", padx=16, pady=(0, 12))

        setattr(self, f"_section_toggle_{section_key}", toggle_btn)
        setattr(self, f"_section_content_{section_key}", content)

        def _toggle():
            visible = not self._section_visible[section_key]
            self._section_visible[section_key] = visible
            toggle_btn.configure(text="▼" if visible else "▶")
            if visible:
                content.pack(fill="x", padx=16, pady=(0, 12))
            else:
                content.pack_forget()

        toggle_btn.configure(command=_toggle)
        builder(content)

    # ── ① 比价环节 ───────────────────────────────────
    def _build_compare(self, parent):
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="比价单价", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.compare_price_entry = ctk.CTkEntry(row1, height=30, font=ctk.CTkFont(size=13))
        self.compare_price_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="比价日期", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.compare_date_entry = ctk.CTkEntry(row2, height=30, font=ctk.CTkFont(size=13),
                                               placeholder_text="YYYY-MM-DD")
        self.compare_date_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row3 = ctk.CTkFrame(parent, fg_color="transparent")
        row3.pack(fill="x", pady=3)
        ctk.CTkLabel(row3, text="比价备注", width=90, anchor="nw",
                      font=ctk.CTkFont(size=13)).pack(side="left", pady=4)
        self.compare_remark_text = ctk.CTkTextbox(row3, height=60, font=ctk.CTkFont(size=13))
        self.compare_remark_text.pack(side="left", fill="x", expand=True, padx=(8, 0))

    # ── ② 合同环节 ───────────────────────────────────
    def _build_contract(self, parent):
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="合同状态", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.contract_status_var = ctk.StringVar()
        self.contract_status_combo = ctk.CTkComboBox(
            row1, variable=self.contract_status_var,
            values=self.CONTRACT_STATUS, width=160, height=30,
            font=ctk.CTkFont(size=13),
        )
        self.contract_status_combo.pack(side="left", padx=(8, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="合同备注", width=90, anchor="nw",
                      font=ctk.CTkFont(size=13)).pack(side="left", pady=4)
        self.contract_remark_text = ctk.CTkTextbox(row2, height=60, font=ctk.CTkFont(size=13))
        self.contract_remark_text.pack(side="left", fill="x", expand=True, padx=(8, 0))

    # ── ③ 通知厂家环节 ────────────────────────────────
    def _build_notify(self, parent):
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="通知日期", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.notify_date_entry = ctk.CTkEntry(row1, height=30, font=ctk.CTkFont(size=13),
                                              placeholder_text="YYYY-MM-DD")
        self.notify_date_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="沟通货期", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.expected_delivery_entry = ctk.CTkEntry(row2, height=30, font=ctk.CTkFont(size=13),
                                                     placeholder_text="YYYY-MM-DD")
        self.expected_delivery_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row3 = ctk.CTkFrame(parent, fg_color="transparent")
        row3.pack(fill="x", pady=3)
        ctk.CTkLabel(row3, text="通知备注", width=90, anchor="nw",
                      font=ctk.CTkFont(size=13)).pack(side="left", pady=4)
        self.notify_remark_text = ctk.CTkTextbox(row3, height=60, font=ctk.CTkFont(size=13))
        self.notify_remark_text.pack(side="left", fill="x", expand=True, padx=(8, 0))

    # ── ④ 生产进度环节 ────────────────────────────────
    def _build_production(self, parent):
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="生产周期", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.production_cycle_entry = ctk.CTkEntry(row1, height=30, font=ctk.CTkFont(size=13),
                                                    placeholder_text="如：15天")
        self.production_cycle_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="预定发货", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.expected_ship_entry = ctk.CTkEntry(row2, height=30, font=ctk.CTkFont(size=13),
                                                  placeholder_text="YYYY-MM-DD")
        self.expected_ship_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row3 = ctk.CTkFrame(parent, fg_color="transparent")
        row3.pack(fill="x", pady=3)
        ctk.CTkLabel(row3, text="生产备注", width=90, anchor="nw",
                      font=ctk.CTkFont(size=13)).pack(side="left", pady=4)
        self.production_remark_text = ctk.CTkTextbox(row3, height=60, font=ctk.CTkFont(size=13))
        self.production_remark_text.pack(side="left", fill="x", expand=True, padx=(8, 0))

    # ── ⑤ 发货跟进环节 ────────────────────────────────
    def _build_ship(self, parent):
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="发货日期", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.ship_date_entry = ctk.CTkEntry(row1, height=30, font=ctk.CTkFont(size=13),
                                             placeholder_text="YYYY-MM-DD")
        self.ship_date_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="发货方式", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.ship_method_var = ctk.StringVar()
        self.ship_method_combo = ctk.CTkComboBox(
            row2, variable=self.ship_method_var,
            values=self.SHIP_METHODS, width=140, height=30,
            font=ctk.CTkFont(size=13),
        )
        self.ship_method_combo.pack(side="left", padx=(8, 0))

        row3 = ctk.CTkFrame(parent, fg_color="transparent")
        row3.pack(fill="x", pady=3)
        ctk.CTkLabel(row3, text="物流单号", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.tracking_no_entry = ctk.CTkEntry(row3, height=30, font=ctk.CTkFont(size=13))
        self.tracking_no_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row4 = ctk.CTkFrame(parent, fg_color="transparent")
        row4.pack(fill="x", pady=3)
        ctk.CTkLabel(row4, text="预计到货", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.expected_arrival_entry = ctk.CTkEntry(row4, height=30, font=ctk.CTkFont(size=13),
                                                     placeholder_text="YYYY-MM-DD")
        self.expected_arrival_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        row5 = ctk.CTkFrame(parent, fg_color="transparent")
        row5.pack(fill="x", pady=3)
        ctk.CTkLabel(row5, text="通知库房", width=90, anchor="w",
                      font=ctk.CTkFont(size=13)).pack(side="left")
        self.notify_warehouse_var = ctk.IntVar(value=0)
        ctk.CTkCheckBox(
            row5, text="已通知库房", variable=self.notify_warehouse_var,
            font=ctk.CTkFont(size=13),
            checkbox_width=20, checkbox_height=20,
        ).pack(side="left", padx=(8, 0))

    # ── 数据加载 ────────────────────────────────────────
    def _load_data(self):
        if not self.oid:
            return
        data = self.db.get_packaging(self.oid)
        if not data:
            messagebox.showerror("错误", "记录不存在")
            self.destroy()
            return

        self.name_entry.insert(0, data.get("material_name", ""))
        if data.get("project_no"):
            self.project_no_entry.insert(0, data["project_no"])
        if data.get("order_factory"):
            self.order_factory_entry.insert(0, data["order_factory"])
        if data.get("order_quantity"):
            self.order_quantity_entry.insert(0, str(data["order_quantity"]))
        self.project_var.set(data.get("project", ""))

        if data.get("compare_price"):
            self.compare_price_entry.insert(0, str(data["compare_price"]))
        if data.get("compare_date"):
            self.compare_date_entry.insert(0, data["compare_date"])
        if data.get("compare_remark"):
            self.compare_remark_text.insert("0.0", data["compare_remark"])

        if data.get("contract_status"):
            self.contract_status_var.set(data["contract_status"])
        if data.get("contract_remark"):
            self.contract_remark_text.insert("0.0", data["contract_remark"])

        if data.get("notify_date"):
            self.notify_date_entry.insert(0, data["notify_date"])
        if data.get("expected_delivery_date"):
            self.expected_delivery_entry.insert(0, data["expected_delivery_date"])
        if data.get("notify_remark"):
            self.notify_remark_text.insert("0.0", data["notify_remark"])

        if data.get("production_cycle"):
            self.production_cycle_entry.insert(0, data["production_cycle"])
        if data.get("expected_ship_date"):
            self.expected_ship_entry.insert(0, data["expected_ship_date"])
        if data.get("production_remark"):
            self.production_remark_text.insert("0.0", data["production_remark"])

        if data.get("ship_date"):
            self.ship_date_entry.insert(0, data["ship_date"])
        if data.get("ship_method"):
            self.ship_method_var.set(data["ship_method"])
        if data.get("tracking_no"):
            self.tracking_no_entry.insert(0, data["tracking_no"])
        if data.get("expected_arrival"):
            self.expected_arrival_entry.insert(0, data["expected_arrival"])
        if data.get("notify_warehouse"):
            self.notify_warehouse_var.set(data["notify_warehouse"])

    # ── 保存 ────────────────────────────────────────────
    def _save(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showerror("错误", "请输入物料名称")
            return

        project = self.project_var.get()
        if not project:
            messagebox.showerror("错误", "请选择所属项目")
            return

        def _get_text(widget):
            if isinstance(widget, ctk.CTkTextbox):
                return widget.get("0.0", "end").strip()
            return widget.get().strip()

        def _get_entry(widget):
            if isinstance(widget, ctk.CTkEntry):
                return widget.get().strip()
            return ""

        data = {
            "material_name": name,
            "project": project,
            "project_no": self.project_no_entry.get().strip() or None,
            "order_factory": self.order_factory_entry.get().strip() or None,
            "order_quantity": self.order_quantity_entry.get().strip() or None,
            "compare_price": float(self.compare_price_entry.get().strip() or 0) or None,
            "compare_date": _get_entry(self.compare_date_entry) or None,
            "compare_remark": _get_text(self.compare_remark_text) or None,
            "contract_status": self.contract_status_var.get() or None,
            "contract_remark": _get_text(self.contract_remark_text) or None,
            "notify_date": _get_entry(self.notify_date_entry) or None,
            "expected_delivery_date": _get_entry(self.expected_delivery_entry) or None,
            "notify_remark": _get_text(self.notify_remark_text) or None,
            "production_cycle": _get_entry(self.production_cycle_entry) or None,
            "expected_ship_date": _get_entry(self.expected_ship_entry) or None,
            "production_remark": _get_text(self.production_remark_text) or None,
            "ship_date": _get_entry(self.ship_date_entry) or None,
            "ship_method": self.ship_method_var.get() or None,
            "tracking_no": _get_entry(self.tracking_no_entry) or None,
            "expected_arrival": _get_entry(self.expected_arrival_entry) or None,
            "notify_warehouse": self.notify_warehouse_var.get(),
            "archived": 0,
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""

        try:
            if self.oid:
                self.db.update_packaging(self.oid, data)
                messagebox.showinfo("成功", "包材订单已更新")
            else:
                self.db.save_packaging(data)
                messagebox.showinfo("成功", "包材订单已新增")
            if self.on_save:
                self.on_save()
            self.destroy()
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    # ── 确认到货并归档 ─────────────────────────────────
    def _confirm_arrival(self):
        if not self.oid:
            messagebox.showwarning("提示", "请先保存订单信息")
            return
        if not self.ship_date_entry.get().strip():
            messagebox.showwarning("提示", "请先填写发货日期")
            return
        if not messagebox.askyesno("确认到货", "确认已到货？确认后该记录将自动归档。"):
            return
        try:
            self._save()
            self.db.archive_packaging(self.oid)
            messagebox.showinfo("成功", "已确认到货并归档！")
            if self.on_save:
                self.on_save()
            self.destroy()
        except Exception as e:
            messagebox.showerror("操作失败", str(e))

    # ── 添加项目 ────────────────────────────────────────
    def _add_project(self):
        from tkinter import simpledialog
        name = simpledialog.askstring("添加项目", "请输入新项目名称：")
        if name and name.strip():
            name = name.strip()
            try:
                self.db.add_project(name)
                projects = self.db.get_projects()
                self.project_combo.configure(values=projects)
                self.project_var.set(name)
                messagebox.showinfo("成功", f"项目「{name}」已添加")
            except Exception as e:
                messagebox.showerror("添加失败", str(e))
