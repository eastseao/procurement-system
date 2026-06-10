#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""催款记录页面 - v1.2"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from datetime import datetime


TABLE_COLS = [
    ("supplier_name",   "供应商名称", 140, "w"),
    ("contact_person",  "联系人名称", 100, "center"),
    ("wechat",          "微信名称",   120, "center"),
    ("reminder_date",   "催款时间",   100, "center"),
    ("amount_due",      "应付金额",   100, "center"),
    ("notify_internal", "通知内勤",    80, "center"),
    ("notify_manager",  "通知经理",    80, "center"),
]


class CollectionPage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.db = db
        self.C = colors
        self._build()
        # 默认不搜索、不显示数据

    def _build(self):
        # ── 顶部标题栏 ─────────────────────────────────
        header = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=0, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header, text="💰  催款记录",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left", padx=24, pady=16)

        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="right", padx=16)

        ctk.CTkButton(
            btn_frame, text="＋ 新增记录", width=110, height=34,
            fg_color=self.C["danger"], hover_color="#A85A5A",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._open_form,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="📤 导出Excel", width=100, height=34,
            fg_color=self.C["success"], hover_color="#7A9A6E",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._export_xlsx,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="📥 导入xlsx", width=100, height=34,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=13),
            command=self._import_xlsx,
        ).pack(side="right", padx=4)

        # ── 搜索栏 ─────────────────────────────────────
        search_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"], height=56)
        search_frame.pack(fill="x", padx=16, pady=(12, 0))
        search_frame.pack_propagate(False)

        inner = ctk.CTkFrame(search_frame, fg_color="transparent")
        inner.pack(side="left", fill="y", padx=16)

        ctk.CTkLabel(inner, text="模糊搜索：",
                     font=ctk.CTkFont(size=12), text_color=self.C["text_secondary"]).pack(
            side="left", pady=12)
        self.kw_var = tk.StringVar()
        kw_entry = ctk.CTkEntry(
            inner, textvariable=self.kw_var, width=200, height=32,
            font=ctk.CTkFont(size=12), placeholder_text="供应商/联系人/微信…",
        )
        kw_entry.pack(side="left", padx=(4, 16))
        kw_entry.bind("<Return>", lambda e: self._do_search())

        ctk.CTkButton(
            inner, text="查询", width=70, height=32,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._do_search,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            inner, text="重置", width=60, height=32,
            fg_color=self.C["border"], text_color=self.C["text"],
            hover_color="#CBD5E1", font=ctk.CTkFont(size=12),
            command=self._reset,
        ).pack(side="left", padx=4)

        self.count_lbl = ctk.CTkLabel(
            search_frame, text="",
            font=ctk.CTkFont(size=12), text_color=self.C["text_secondary"],
        )
        self.count_lbl.pack(side="right", padx=20)

        # ── 表格区域 - 阴影分层：z1 → z2 → z3 ─────────────────────────────────────
        # z1（底层）：fg_color=#E8E2D9, corner_radius=16, 内边距 4
        z1_frame = ctk.CTkFrame(self, fg_color="#E8E2D9", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=16, pady=(8, 16))

        # z2（中层）：fg_color=#F2F0EB, corner_radius=12, 内边距 4
        z2_frame = ctk.CTkFrame(z1_frame, fg_color="#F2F0EB", corner_radius=12)
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # z3（上层卡片）：fg_color=C["card"], corner_radius=C["radius_card"]
        list_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        list_frame.pack(fill="both", expand=True, padx=4, pady=4)

        style = ttk.Style()
        style.configure("Collection.Treeview",
                        font=("Microsoft YaHei", 11), rowheight=36,
                        background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#1E293B")
        style.configure("Collection.Treeview.Heading",
                        font=("Microsoft YaHei", 11, "bold"),
                        background="#F8FAFC", foreground="#475569", relief="flat")
        style.map("Collection.Treeview",
                  background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])

        tree_frame = tk.Frame(list_frame, bg="#FFFFFF")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        col_ids = [c[0] for c in TABLE_COLS] + ["action"]
        self.tree = ttk.Treeview(
            tree_frame, style="Collection.Treeview",
            columns=col_ids, show="headings", selectmode="browse",
        )
        for cid, label, width, anchor in TABLE_COLS:
            self.tree.heading(cid, text=label)
            self.tree.column(cid, width=width, minwidth=40, stretch=True, anchor=anchor)
        self.tree.heading("action", text="操作")
        self.tree.column("action", width=100, minwidth=40, stretch=True, anchor="center")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<Button-1>", self._on_click)


        self.tree.tag_configure("odd", background="#F8FAFC")
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)
        self.tree.bind("<Double-1>", self._on_double_click)

    # ── 搜索 / 重置 ────────────────────────────────────
    def _do_search(self):
        kw = self.kw_var.get().strip()
        records = self.db.get_collections(keyword=kw if kw else None)
        self._render(records)
        self.count_lbl.configure(text=f"共 {len(records)} 条")

    def _reset(self):
        self.kw_var.set("")
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.count_lbl.configure(text="")

    def _render(self, records):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for i, r in enumerate(records):
            tag = "odd" if i % 2 == 0 else "even"
            amount = r.get("amount_due")
            amount_str = f"¥{amount:,.2f}" if amount else "—"
            self.tree.insert("", "end", iid=str(r["id"]),
                             values=(
                                 r.get("supplier_name") or "—",
                                 r.get("contact_person") or "—",
                                 r.get("wechat") or "—",
                                 r.get("reminder_date") or "—",
                                 amount_str,
                                 "是" if r.get("notify_internal") else "否",
                                 "是" if r.get("notify_manager") else "否",
                                 "编辑  删除",
                             ), tags=(tag,))

    # ── 双击操作 ───────────────────────────────────────
    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        cid = int(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == len(TABLE_COLS):  # 操作列
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                if x_rel < col_w * 0.5:
                    self._open_form(cid)
                else:
                    self._delete(cid)
        else:
            self._open_form(cid)

    def _on_hover(self, event):
        """鼠标悬浮时高亮行背景"""
        item = self.tree.identify_row(event.y)
        if item and item != getattr(self, "_last_hover", None):
            if hasattr(self, "_last_hover") and self._last_hover:
                tags = list(self.tree.item(self._last_hover, "tags"))[0]
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(self._last_hover, tags=tags)
            tags = list(self.tree.item(item, "tags"))[0]
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tags)
            self._last_hover = item
        elif not item and hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags"))[0]
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_leave(self, event):
        """鼠标离开时清除悬浮高亮"""
        if hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags"))[0]
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _open_form(self, cid=None):
        try:
            form = CollectionForm(self, self.db, self.C, cid=cid, on_save=self._do_search)
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    def _delete(self, cid):
        rec = self.db.get_collection(cid)
        name = rec.get("supplier_name", "") if rec else ""
        if messagebox.askyesno("删除确认", f"确定删除「{name}」的催款记录？", icon="warning"):
            self.db.delete_collection(cid)
            self._do_search()

    # ── 导出 Excel ─────────────────────────────────────
    def _export_xlsx(self):
        kw = self.kw_var.get().strip()
        records = self.db.get_collections(keyword=kw if kw else None)
        if not records:
            messagebox.showinfo("提示", "没有可导出的数据，请先查询")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出催款记录", defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"催款记录_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )
        if not filepath:
            return

        headers = ["供应商名称", "联系人名称", "微信名称", "催款时间",
                   "应付金额", "通知内勤", "通知经理", "备注"]
        rows = []
        for r in records:
            amount = r.get("amount_due")
            rows.append({
                "供应商名称": r.get("supplier_name") or "",
                "联系人名称": r.get("contact_person") or "",
                "微信名称": r.get("wechat") or "",
                "催款时间": r.get("reminder_date") or "",
                "应付金额": f"{amount:,.2f}" if amount else "",
                "通知内勤": "是" if r.get("notify_internal") else "否",
                "通知经理": "是" if r.get("notify_manager") else "否",
                "备注": r.get("remark") or "",
            })
        try:
            self.db.export_to_xlsx(filepath, "催款记录", headers, rows,
                                   col_widths=[16, 10, 14, 12, 12, 8, 8, 20])
            messagebox.showinfo("导出成功", f"已导出 {len(rows)} 条记录到：\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── 导入 Excel ─────────────────────────────────────
    def _import_xlsx(self):
        filepath = filedialog.askopenfilename(
            title="选择催款记录xlsx文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not filepath:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            field_map = {
                "供应商名称": "supplier_name", "联系人名称": "contact_person",
                "微信名称": "wechat", "催款时间": "reminder_date",
                "应付金额": "amount_due", "通知内勤": "notify_internal",
                "通知经理": "notify_manager", "备注": "remark",
            }
            col_index = {}
            for i, h in enumerate(headers):
                if h in field_map:
                    col_index[field_map[h]] = i

            if "supplier_name" not in col_index:
                messagebox.showerror("导入失败", "Excel中未找到「供应商名称」列，请检查表头")
                return

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row):
                    continue
                def get_val(field):
                    idx = col_index.get(field)
                    if idx is None:
                        return ""
                    v = row[idx]
                    return str(v).strip() if v is not None else ""

                name = get_val("supplier_name")
                if not name:
                    continue

                amount_str = get_val("amount_due")
                try:
                    amount = float(amount_str.replace("¥", "").replace(",", "")) if amount_str else 0
                except (ValueError, AttributeError):
                    amount = 0

                ni_str = get_val("notify_internal")
                notify_internal = 1 if ni_str in ("是", "1", "yes", "YES", "Yes") else 0
                nm_str = get_val("notify_manager")
                notify_manager = 1 if nm_str in ("是", "1", "yes", "YES", "Yes") else 0

                data = {
                    "supplier_name": name,
                    "contact_person": get_val("contact_person"),
                    "wechat": get_val("wechat"),
                    "reminder_date": get_val("reminder_date"),
                    "amount_due": amount,
                    "notify_internal": notify_internal,
                    "notify_manager": notify_manager,
                    "remark": get_val("remark"),
                }
                self.db.save_collection(data)
                count += 1

            messagebox.showinfo("导入成功", f"成功导入 {count} 条催款记录")
            self._do_search()
        except ImportError:
            messagebox.showerror("缺少依赖", "请安装 openpyxl：pip install openpyxl")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))



    # ── 单击选中高亮行 ─────────────────────
    def _on_click(self, event):
        """单击选中行，高亮显示"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
    
# ============================
# 催款记录表单弹窗
# ============================
class CollectionForm(ctk.CTkToplevel):
    def __init__(self, parent, db, colors, cid=None, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = colors
        self.cid = cid
        self.on_save = on_save

        self.title("编辑催款记录" if cid else "新增催款记录")
        self.geometry("600x520")
        self.resizable(True, True)
        self.configure(fg_color=self.C["bg"])

        self.update_idletasks()
        pw, ph = parent.winfo_rootx(), parent.winfo_rooty()
        self.geometry(f"600x520+{pw+80}+{ph+40}")

        self._build()
        if cid:
            self._load(cid)

    def _build(self):
        outer = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"])
        outer.pack(fill="both", expand=True)

        card = ctk.CTkFrame(outer, fg_color=self.C["card"], corner_radius=self.C["radius_modal"])
        card.pack(fill="both", expand=True, padx=16, pady=16)

        def row(parent, label, widget_fn, **kw):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=16, pady=4)
            ctk.CTkLabel(f, text=label, width=90, anchor="w",
                         font=ctk.CTkFont(size=12),
                         text_color=self.C["text_secondary"]).pack(side="left")
            widget_fn(f, **kw)

        def entry(parent, var, ph=""):
            e = ctk.CTkEntry(parent, textvariable=var, height=32,
                             font=ctk.CTkFont(size=12), placeholder_text=ph)
            e.pack(side="left", fill="x", expand=True, padx=(8, 0))

        # ── 字段 ──
        self.v_supplier  = tk.StringVar()
        self.v_contact   = tk.StringVar()
        self.v_wechat    = tk.StringVar()
        self.v_date      = tk.StringVar()
        self.v_amount    = tk.StringVar()
        self.v_internal  = tk.IntVar(value=0)
        self.v_manager   = tk.IntVar(value=0)

        self._section(card, "📋 基本信息")
        row(card, "供应商名称 *", lambda p, **k: entry(p, self.v_supplier, "供应商全称"))
        row(card, "联系人名称",   lambda p, **k: entry(p, self.v_contact,  "联系人姓名"))
        row(card, "微信名称",     lambda p, **k: entry(p, self.v_wechat,   "微信号"))

        self._section(card, "💰 催款信息")
        row(card, "催款时间",     lambda p, **k: entry(p, self.v_date,     "YYYY-MM-DD"))
        row(card, "应付金额",     lambda p, **k: entry(p, self.v_amount,   "如：50000"))

        self._section(card, "📢 通知状态")
        notify_frame = ctk.CTkFrame(card, fg_color="transparent")
        notify_frame.pack(fill="x", padx=16, pady=8)
        ctk.CTkCheckBox(notify_frame, text="已通知内勤", variable=self.v_internal,
                        font=ctk.CTkFont(size=13),
                        checkbox_width=20, checkbox_height=20).pack(side="left", padx=8)
        ctk.CTkCheckBox(notify_frame, text="已通知经理", variable=self.v_manager,
                        font=ctk.CTkFont(size=13),
                        checkbox_width=20, checkbox_height=20).pack(side="left", padx=8)

        self._section(card, "📝 备注")
        remark_frame = ctk.CTkFrame(card, fg_color="transparent")
        remark_frame.pack(fill="x", padx=16, pady=4)
        self.remark_text = ctk.CTkTextbox(remark_frame, height=80, font=ctk.CTkFont(size=12))
        self.remark_text.pack(fill="x")

        # 按钮
        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=12)
        ctk.CTkButton(btn_row, text="✓ 保存", width=120, height=38,
                      fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._save).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="✕ 关闭", width=80, height=38,
                      fg_color="#6B7280", hover_color="#4B5563",
                      font=ctk.CTkFont(size=13), command=self.destroy).pack(side="left", padx=4)

    def _section(self, parent, title):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=self.C["primary"]).pack(side="left")
        ctk.CTkFrame(f, height=1, fg_color=self.C["border"]).pack(
            side="left", fill="x", expand=True, padx=(8, 0))

    def _load(self, cid):
        r = self.db.get_collection(cid)
        if not r:
            return
        self.v_supplier.set(r.get("supplier_name") or "")
        self.v_contact.set(r.get("contact_person") or "")
        self.v_wechat.set(r.get("wechat") or "")
        self.v_date.set(r.get("reminder_date") or "")
        amount = r.get("amount_due")
        self.v_amount.set(f"{amount:,.2f}" if amount else "")
        self.v_internal.set(1 if r.get("notify_internal") else 0)
        self.v_manager.set(1 if r.get("notify_manager") else 0)
        if r.get("remark"):
            self.remark_text.insert("1.0", r["remark"])

    def _save(self):
        name = self.v_supplier.get().strip()
        if not name:
            messagebox.showerror("验证失败", "请填写供应商名称", parent=self)
            return

        amount_str = self.v_amount.get().strip()
        try:
            amount = float(amount_str.replace("¥", "").replace(",", "")) if amount_str else 0
        except (ValueError, AttributeError):
            amount = 0

        data = {
            "supplier_name": name,
            "contact_person": self.v_contact.get().strip(),
            "wechat": self.v_wechat.get().strip(),
            "reminder_date": self.v_date.get().strip(),
            "amount_due": amount,
            "notify_internal": self.v_internal.get(),
            "notify_manager": self.v_manager.get(),
            "remark": self.remark_text.get("1.0", "end").strip(),
        }

        if self.cid:
            self.db.update_collection(self.cid, data)
            messagebox.showinfo("成功", "催款记录已更新", parent=self)
        else:
            self.db.save_collection(data)
            messagebox.showinfo("成功", "催款记录已添加", parent=self)
        if self.on_save:
            self.on_save()
        self.destroy()
