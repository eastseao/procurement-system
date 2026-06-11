#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""报价单页面 v2.0 - 供方配置库、勾选导出、新文件名格式"""

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
from datetime import datetime
import os
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import range_boundaries


def _get_resource_path(relative_path):
    """获取资源绝对路径（兼容打包后 _MEIPASS）"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


class QuotationPage(ctk.CTkFrame):
    """产品报价单主页面 V2.0"""

    def __init__(self, parent, db, C):
        super().__init__(parent, fg_color=C["bg"], corner_radius=0)
        self.db = db
        self.C = C
        self.products = []
        self.config_data = {}
        self.checked_ids = set()  # 勾选导出的产品ID集合
        self._build_ui()
        self._load_data()

    # ═══════════════════════════════════════════════
    #  UI 构建（按照合同生成页面排版）
    # ═══════════════════════════════════════════════

    def _build_ui(self):
        # ── 顶部工具栏 ──
        toolbar = ctk.CTkFrame(self, fg_color="transparent", height=56)
        toolbar.pack(fill="x", padx=24, pady=(16, 8))
        toolbar.pack_propagate(False)

        ctk.CTkLabel(
            toolbar, text="📋  产品报价单",
            font=ctk.CTkFont(family="Microsoft YaHei", size=20, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left", pady=12)

        btn_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        btn_frame.pack(side="right", pady=8)

        ctk.CTkButton(
            btn_frame, text="📤 导出报价单", width=120, height=34,
            fg_color=self.C["success"], hover_color="#7A9A6E",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._export_quotation,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="+ 供方配置", width=110, height=34,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=13),
            command=self._open_supplier_config,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="⚙ 需方配置", width=100, height=34,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=13),
            command=self._open_config,
        ).pack(side="right", padx=4)

        # ── 供应商检索栏（参照合同生成页面）──
        sb = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        sb.pack(fill="x", padx=24, pady=(0, 8))

        ctk.CTkLabel(
            sb, text="🔍  供应商检索：",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left", padx=(16, 6), pady=12)

        # 下拉选择框
        self.supplier_combo = ttk.Combobox(
            sb, width=18, height=6, state="readonly",
            font=("Microsoft YaHei", 11),
        )
        self.supplier_combo.pack(side="left", padx=(0, 10), pady=12)
        self.supplier_combo.bind("<<ComboboxSelected>>", self._on_supplier_selected)

        ctk.CTkLabel(sb, text="或输入：", font=ctk.CTkFont(size=12),
                      text_color="#9CA3AF").pack(side="left", padx=(0, 6), pady=12)

        self.search_entry = ctk.CTkEntry(
            sb, width=160, height=32,
            placeholder_text="输入简称/全称...", font=ctk.CTkFont(size=12),
        )
        self.search_entry.pack(side="left", padx=(0, 10), pady=12)
        self.search_entry.bind("<KeyRelease>", self._on_search_key)
        self.search_entry.bind("<Return>", self._on_search)

        ctk.CTkButton(
            sb, text="检索", width=60, height=30,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=12),
            command=self._on_search,
        ).pack(side="left", padx=(0, 10), pady=12)

        # 确认按钮
        self.confirm_btn = ctk.CTkButton(
            sb, text="确认", width=80, height=30,
            fg_color=self.C["success"], hover_color="#7A9A6E",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._on_confirm_supplier,
        )
        self.confirm_btn.pack(side="left", padx=(0, 10), pady=12)

        # 重置按钮
        ctk.CTkButton(
            sb, text="重置", width=80, height=30,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._on_reset_all,
        ).pack(side="left", padx=(0, 10), pady=12)

        # 确认状态标签
        self.confirm_status_label = ctk.CTkLabel(sb, text="", font=ctk.CTkFont(size=11),
                                               text_color="#9CA3AF")
        self.confirm_status_label.pack(side="left", padx=(0, 10), pady=12)

        # ── 滚动区域 ──
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent", corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        self._build_product_info(scroll)
        self._build_tier_price(scroll)
        self._build_product_table(scroll)

    # ═══════════════════════════════════════════════
    #  产品信息板块（内联表单）
    # ═══════════════════════════════════════════════

    def _build_product_info(self, parent):
        """产品信息板块 — 内联表单"""
        card = ctk.CTkFrame(parent, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        card.pack(fill="x", pady=(0, 12))

        # 标题行
        hf = ctk.CTkFrame(card, fg_color="transparent")
        hf.pack(fill="x", padx=20, pady=(14, 8))
        ctk.CTkLabel(
            hf, text="📦  产品信息",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left")

        # 表单区域
        form = ctk.CTkFrame(card, fg_color="transparent")
        form.pack(fill="x", padx=20, pady=(0, 14))

        fields = [
            ("产品名称 *", "product_name", 200),
            ("项目号", "item_no", 200),
            ("材质工艺/描述", "material_process", 280),
            ("产品尺寸", "product_size", 200),
            ("供货周期", "supply_cycle", 200),
            ("发货箱规", "carton_spec", 200),
            ("单位", "unit", 100),
        ]

        self.inline_entries = {}
        for label, key, width in fields:
            row = ctk.CTkFrame(form, fg_color="transparent")
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(
                row, text=label,
                width=120, anchor="e",
                font=ctk.CTkFont(size=12),
                text_color=self.C.get("text_secondary", "#8B7355"),
            ).pack(side="left", padx=(0, 8))
            e = ctk.CTkEntry(row, width=width, height=30, font=ctk.CTkFont(size=12))
            e.pack(side="left")
            if key == "unit":
                e.insert(0, "PCS")
            self.inline_entries[key] = e

        # 按钮栏
        bf = ctk.CTkFrame(card, fg_color="transparent")
        bf.pack(fill="x", padx=20, pady=(0, 14))
        ctk.CTkButton(
            bf, text="💾 保存产品", width=100, height=32,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._save_inline_product,
        ).pack(side="right", padx=4)
        ctk.CTkButton(
            bf, text="🗑 清空", width=80, height=32,
            fg_color="#9CA3AF", hover_color="#6B7280",
            font=ctk.CTkFont(size=12),
            command=self._clear_inline_form,
        ).pack(side="right", padx=4)

    # ═══════════════════════════════════════════════
    #  阶梯价格板块（内联表单）
    # ═══════════════════════════════════════════════

    def _build_tier_price(self, parent):
        """阶梯价格板块 — 内联表单，支持多行阶梯"""
        card = ctk.CTkFrame(parent, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        card.pack(fill="x", pady=(0, 12))

        # 标题行 + 添加阶梯按钮
        hf = ctk.CTkFrame(card, fg_color="transparent")
        hf.pack(fill="x", padx=20, pady=(14, 8))
        ctk.CTkLabel(
            hf, text="💰  阶梯价格",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left")

        ctk.CTkButton(
            hf, text="＋ 添加阶梯", width=100, height=28,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=11),
            command=self._add_inline_tier_row,
        ).pack(side="right")

        # 阶梯行容器
        self.inline_tier_frame = ctk.CTkFrame(card, fg_color="transparent")
        self.inline_tier_frame.pack(fill="x", padx=20, pady=(0, 14))

        # 表头
        hdr = ctk.CTkFrame(self.inline_tier_frame, fg_color=self.C.get("sidebar", "#EEE5DA"), corner_radius=4)
        hdr.pack(fill="x", pady=(0, 4))
        for h_text, w in [("最低数量", 140), ("最高数量(留空=无限)", 180), ("操作", 60)]:
            ctk.CTkLabel(
                hdr, text=h_text, width=w, anchor="center",
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=self.C["text"],
            ).pack(side="left", padx=4, pady=4)

        # 提示文字
        hint = ctk.CTkLabel(
            self.inline_tier_frame,
            text="💡 填入阶梯数量后保存产品，如：100 | 500-999 | ≥1000",
            font=ctk.CTkFont(size=11), text_color=self.C.get("text_secondary", "#8B7355"),
        )
        hint.pack(anchor="w", pady=(2, 6))

        # 默认添加一行
        self.inline_tier_rows = []
        self._add_inline_tier_row()

    def _add_inline_tier_row(self):
        """在阶梯价格区域新增一行输入框"""
        row = ctk.CTkFrame(self.inline_tier_frame, fg_color="transparent")
        row.pack(fill="x", pady=2)

        min_entry = ctk.CTkEntry(row, width=140, height=30, font=ctk.CTkFont(size=12),
                                  placeholder_text="最低数量")
        min_entry.pack(side="left", padx=4)
        max_entry = ctk.CTkEntry(row, width=180, height=30, font=ctk.CTkFont(size=12),
                                  placeholder_text="留空=无上限")
        max_entry.pack(side="left", padx=4)

        def _remove():
            row.destroy()
            self.inline_tier_rows = [tr for tr in self.inline_tier_rows if tr["row"] is not row]

        del_btn = ctk.CTkButton(row, text="✕", width=40, height=30,
                                fg_color=self.C["danger"], hover_color="#A85A5A",
                                font=ctk.CTkFont(size=12), command=_remove)
        del_btn.pack(side="left", padx=4)

        tier_data = {"row": row, "min": min_entry, "max": max_entry}
        self.inline_tier_rows.append(tier_data)

    # ═══════════════════════════════════════════════
    #  产品列表表格板块
    # ═══════════════════════════════════════════════

    def _build_product_table(self, parent):
        """产品列表表格板块 — Treeview + 勾选列 + 统计信息"""
        card = ctk.CTkFrame(parent, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        card.pack(fill="both", expand=True, pady=(0, 16))

        # 标题行 + 操作按钮
        hf = ctk.CTkFrame(card, fg_color="transparent")
        hf.pack(fill="x", padx=20, pady=(14, 8))
        ctk.CTkLabel(
            hf, text="📋  产品列表",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left")
        ctk.CTkButton(
            hf, text="＋ 添加产品", width=100, height=28,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=11),
            command=lambda: self._open_form(),
        ).pack(side="right")
        ctk.CTkButton(
            hf, text="全选", width=60, height=28,
            fg_color="#6B8FA3", hover_color="#5A7A93",
            font=ctk.CTkFont(size=11),
            command=self._select_all,
        ).pack(side="right", padx=(0, 6))
        ctk.CTkButton(
            hf, text="清选", width=60, height=28,
            fg_color="#9CA3AF", hover_color="#6B7280",
            font=ctk.CTkFont(size=11),
            command=self._deselect_all,
        ).pack(side="right", padx=(0, 6))

        # 表格 - 阴影分层：z1 → z2 → z3
        z1_frame = ctk.CTkFrame(card, fg_color="#E8E2D9", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=4, pady=4)

        z2_frame = ctk.CTkFrame(z1_frame, fg_color="#F5F0EB", corner_radius=12)
        z2_frame.pack(fill="both", expand=True, padx=2, pady=2)

        z3_frame = ctk.CTkFrame(z2_frame, fg_color="white", corner_radius=8)
        z3_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Treeview 列定义
        columns = ("idx", "item_no", "name", "size", "process", "cycle", "carton", "tiers", "action")
        col_widths = [40, 80, 160, 100, 140, 80, 100, 130, 90]

        self.tree = ttk.Treeview(z3_frame, columns=columns, show="headings", height=8,
                                  selectmode="browse")

        # 配置样式
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("QuotationTreeview",
                        background="white",
                        foreground="#333333",
                        fieldbackground="white",
                        font=("Microsoft YaHei", 10),
                        rowheight=34)
        style.configure("QuotationTreeview.Heading",
                        background="#F0EAE0",
                        foreground="#4A4A4A",
                        font=("Microsoft YaHei", 10, "bold"))
        style.map("QuotationTreeview",
                  background=[("selected", "#D4C4B0")],
                  foreground=[("selected", "#000000")])
        style.layout("QuotationTreeview", [("QuotationTreeview.treearea", {"sticky": "nswe"})])

        self.tree.configure(style="QuotationTreeview")

        headings = ["序号", "项目号", "产品名称", "尺寸", "材质/工艺描述", "供货周期", "发货箱规", "阶梯价格", "操作"]
        for col, heading, w in zip(columns, headings, col_widths):
            self.tree.heading(col, text=heading)
            self.tree.column(col, width=w, minwidth=w, anchor="center" if col != "action" else "w")

        # 绑定事件
        self.tree.bind("<Double-1>", self._on_row_double_click)
        self.tree.bind("<ButtonRelease-1>", self._on_row_click)
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)

        # 滚动条
        vsb = ttk.Scrollbar(z3_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(z3_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        z3_frame.grid_rowconfigure(0, weight=1)
        z3_frame.grid_columnconfigure(0, weight=1)

        # Tag 配置颜色
        self.tree.tag_configure("checked", background="#E8F0E8")
        self.tree.tag_configure("odd", background="#FAFAF8")
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("hover", background="#F0E0D0")

        # 底部统计标签
        stats_frame = ctk.CTkFrame(card, fg_color="transparent")
        stats_frame.pack(fill="x", padx=20, pady=(8, 14))
        self.stats_label = ctk.CTkLabel(
            stats_frame, text="加载中...",
            font=ctk.CTkFont(size=11),
            text_color=self.C.get("text_secondary", "#8B7355"),
        )
        self.stats_label.pack(side="left")

    # ═══════════════════════════════════════════════
    #  内联表单操作（保存/清空/确认/重置）
    # ═══════════════════════════════════════════════

    def _save_inline_product(self):
        """从内联表单保存产品（含阶梯价格）到数据库"""
        name = (self.inline_entries.get("product_name").get().strip()
                if "product_name" in self.inline_entries else "")
        if not name:
            messagebox.showerror("错误", "请输入产品名称")
            return

        # 收集产品基本信息
        product_data = {
            "item_no": self.inline_entries.get("item_no", ctk.CTkEntry()).get().strip(),
            "product_name": name,
            "product_size": self.inline_entries.get("product_size", ctk.CTkEntry()).get().strip(),
            "material_process": self.inline_entries.get("material_process", ctk.CTkEntry()).get().strip(),
            "supply_cycle": self.inline_entries.get("supply_cycle", ctk.CTkEntry()).get().strip(),
            "carton_spec": self.inline_entries.get("carton_spec", ctk.CTkEntry()).get().strip(),
            "unit": self.inline_entries.get("unit", ctk.CTkEntry()).get().strip() or "PCS",
        }

        # 收集阶梯价格
        valid_tiers = []
        for tr in getattr(self, "inline_tier_rows", []):
            try:
                min_qty = int(tr["min"].get().strip())
            except (ValueError, TypeError):
                min_qty = 0
            max_val = tr["max"].get().strip()
            max_qty = int(max_val) if max_val else None
            valid_tiers.append({"min_qty": min_qty, "max_qty": max_qty})

        try:
            pid = self.db.save_quotation_product(product_data)
            for t in valid_tiers:
                self.db.save_quotation_tier({**t, "product_id": pid})
            messagebox.showinfo("成功", f"产品「{name}」已保存")
            self._clear_inline_form()
            self._load_data()
        except Exception as e:
            import traceback
            messagebox.showerror("保存失败", f"{e}\n\n{traceback.format_exc()}")

    def _clear_inline_form(self):
        """清空内联所有表单字段"""
        if hasattr(self, "inline_entries"):
            for key, entry in self.inline_entries.items():
                entry.delete(0, "end")
                if key == "unit":
                    entry.insert(0, "PCS")
        # 清空阶梯价格行（保留一行空白）
        if hasattr(self, "inline_tier_frame") and hasattr(self, "inline_tier_rows"):
            for tr in self.inline_tier_rows:
                if tr["row"].winfo_exists():
                    tr["row"].destroy()
            self.inline_tier_rows = []
            self._add_inline_tier_row()

    def _on_confirm_supplier(self):
        """确认按钮：将选中的供应商信息注入当前上下文"""
        selected = self.supplier_combo.get()
        if not selected or selected.startswith("—"):
            # 尝试用文本框内容
            keyword = self.search_entry.get().strip()
            if not keyword:
                messagebox.showwarning("提示", "请先选择或检索供应商，再点击确认")
                return
            selected = keyword

        # 存储到实例变量供导出时使用
        self._pending_supplier = {"supplier_name": selected}
        self.confirm_status_label.configure(
            text=f"✅ 已确认：{selected}",
            text_color=self.C.get("success", "#6B9E6B"),
            font=ctk.CTkFont(size=11, weight="bold"),
        )

    def _on_reset_all(self):
        """重置所有状态：清空表单、重置下拉框和检索"""
        self._clear_inline_form()
        # 重置供应商下拉框
        if hasattr(self, "supplier_combo"):
            vals = self.supplier_combo.cget("values")
            if vals:
                self.supplier_combo.current(0)
        # 清空检索框
        if hasattr(self, "search_entry"):
            self.search_entry.delete(0, "end")
        # 重置确认状态
        if hasattr(self, "confirm_status_label"):
            self.confirm_status_label.configure(text="", text_color="#9CA3AF",
                                                font=ctk.CTkFont(size=11))
        # 清除待定供应商
        self._pending_supplier = {}
        # 刷新全部数据
        self._load_data()

    # ═══════════════════════════════════════════════
    #  数据加载与表格渲染（带勾选状态）
    # ═══════════════════════════════════════════════

    def _load_data(self):
        self.products = self.db.get_quotation_products()
        self.config_data = self.db.get_quotation_config()
        self._render_table()
        total = len(self.products)
        total_tiers = sum(len(p.get("tiers", [])) for p in self.products)
        buyer = self.config_data.get("buyer_name", "")
        checked_count = len(self.checked_ids)
        self.stats_label.configure(
            text=f"共 {total} 款产品 | {total_tiers} 个价格阶梯 | 已选 {checked_count} 条导出 | 需方：{buyer}"
        )
        # 刷新供应商下拉列表
        self._refresh_supplier_combo()

    def _render_table(self):
        self._render_items(self.products)

    def _render_items(self, items):
        """通用表格渲染（支持全量或过滤后的产品列表）"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        for idx, p in enumerate(items, 1):
            pid = p["id"]
            is_checked = pid in self.checked_ids
            check_mark = "☑" if is_checked else "☐"

            tiers = p.get("tiers", [])
            if tiers:
                tier_texts = []
                for t in tiers:
                    mn = int(t.get("min_qty", 0))
                    mx = t.get("max_qty")
                    up = t.get("unit_price", 0)
                    if mx:
                        tier_texts.append(f"{mn}-{int(mx)-1}")
                    else:
                        tier_texts.append(f">={mn}")
                tier_str = " | ".join(tier_texts)
            else:
                tier_str = "未设置"

            tags = ("checked",) if is_checked else (("odd",) if idx % 2 == 1 else ("even",))

            self.tree.insert("", "end", iid=str(pid),
                values=(
                    check_mark,
                    idx,
                    p.get("item_no", "-"),
                    p.get("product_name", ""),
                    p.get("product_size", "-"),
                    p.get("material_process", "-"),
                    p.get("supply_cycle", "-"),
                    p.get("carton_spec", "-"),
                    tier_str,
                    "编辑  删除",
                ),
                tags=tags,
            )

    def _render_filtered(self, filtered):
        """渲染检索过滤后的列表"""
        self._render_items(filtered)
        # 更新统计显示
        total_tiers = sum(len(p.get("tiers", [])) for p in filtered)
        checked_count = sum(1 for p in filtered if p["id"] in self.checked_ids)
        buyer = self.config_data.get("buyer_name", "")
        keyword = self.search_entry.get().strip()
        self.stats_label.configure(
            text=f"检索「{keyword}」: {len(filtered)} 款匹配 | {total_tiers} 阶梯 | 已选 {checked_count} 条 | 需方：{buyer}"
        )

    def _select_all(self):
        """全选所有产品"""
        self.checked_ids = set(p["id"] for p in self.products)
        self._render_table()
        self._update_stats()

    def _deselect_all(self):
        """清空所有选择"""
        self.checked_ids.clear()
        self._render_table()
        self._update_stats()

    def _update_stats(self):
        total_tiers = sum(len(p.get("tiers", [])) for p in self.products)
        buyer = self.config_data.get("buyer_name", "")
        checked_count = len(self.checked_ids)
        self.stats_label.configure(
            text=f"共 {len(self.products)} 款产品 | {total_tiers} 个价格阶梯 | 已选 {checked_count} 条导出 | 需方：{buyer}"
        )

    # ═══════════════════════════════════════════════
    #  供应商检索
    # ═══════════════════════════════════════════════

    def _on_search_key(self, event):
        """实时检索（每次按键触发，300ms 防抖）"""
        if hasattr(self, "_search_after"):
            self.after_cancel(self._search_after)
        self._search_after = self.after(300, self._on_search)

    def _on_search(self, event=None):
        """按供应商名称检索产品并刷新表格"""
        keyword = self.search_entry.get().strip()
        if not keyword:
            # 检索词为空 → 显示全部
            self._render_table()
            return
        keyword_lower = keyword.lower()
        # 过滤 products 中匹配的产品（按名称模糊匹配）
        filtered = [
            p for p in self.products
            if keyword_lower in (p.get("product_name") or "").lower()
               or keyword_lower in (p.get("item_no") or "").lower()
        ]
        # 临时替换渲染
        self._render_filtered(filtered)

    # ═══════════════════════════════════════════════
    #  供应商下拉选择
    # ═══════════════════════════════════════════════

    def _refresh_supplier_combo(self):
        """从数据库加载供应商列表，刷新下拉选项"""
        try:
            suppliers = self.db.get_all_quotation_suppliers()
            names = ["— 选择供应商 —"] + [s["supplier_name"] for s in suppliers if s.get("supplier_name")]
            self.supplier_combo["values"] = names
            if names:
                self.supplier_combo.current(0)
        except Exception:
            # 表可能尚未创建（首次运行）
            self.supplier_combo["values"] = []
            self.supplier_combo.set("")

    def _on_supplier_selected(self, event=None):
        """下拉选择了某个供应商 → 填入文本框并触发检索"""
        selected = self.supplier_combo.get()
        if not selected or selected.startswith("—"):
            return
        self.search_entry.delete(0, "end")
        self.search_entry.insert(0, selected)
        # 自动触发检索
        self._on_search()

    # ═══════════════════════════════════════════════
    #  表格交互（含勾选切换）
    # ═══════════════════════════════════════════════

    def _on_row_double_click(self, event):
        sel = self.tree.selection()
        if sel:
            self._open_form(oid=int(sel[0]))

    def _on_row_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return

        self.tree.selection_set(item)
        col_idx = int(col.replace("#", "")) - 1
        oid = int(item)

        # 第0列：点击勾选列 → 切换选中状态
        if col_idx == 0:
            if oid in self.checked_ids:
                self.checked_ids.discard(oid)
            else:
                self.checked_ids.add(oid)
            self._render_table()
            self._update_stats()
            return

        # 最后一列（索引9）：操作列 → 编辑 / 删除
        if col_idx == 9:
            bbox = self.tree.bbox(item, col)
            if not bbox:
                return
            x_rel = event.x - bbox[0]
            col_w = bbox[2]
            if x_rel < col_w * 0.5:
                self._edit_product(oid)
            else:
                self._delete_product(oid)

    def _on_hover(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        if item != getattr(self, "_last_hover", None):
            last = getattr(self, "_last_hover", None)
            if last:
                tags = list(self.tree.item(last, "tags"))
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(last, tags=tags)
            tags = list(self.tree.item(item, "tags"))
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tags)
            self._last_hover = item

    def _on_leave(self, event):
        last = getattr(self, "_last_hover", None)
        if last:
            tags = list(self.tree.item(last, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(last, tags=tags)
            self._last_hover = None

    def _edit_product(self, oid):
        self._open_form(oid=oid)

    def _delete_product(self, oid):
        rec = next((p for p in self.products if p["id"] == oid), None)
        name = rec.get("product_name", "") if rec else ""
        if messagebox.askyesno("删除确认", f"确定删除产品「{name}」及其所有阶梯价格？", icon="warning"):
            self.db.delete_quotation_product(oid)
            self.checked_ids.discard(oid)  # 同时清除勾选
            self._load_data()

    # ═══════════════════════════════════════════════
    #  弹窗打开
    # ═══════════════════════════════════════════════

    def _open_form(self, oid=None):
        try:
            form = QuotationForm(self, self.db, self.C, oid=oid, on_save=self._load_data)
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    def _open_config(self):
        try:
            dlg = QuotationConfigDialog(self, self.db, self.C, on_save=self._load_data)
            dlg.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开配置失败", f"{e}\n\n{traceback.format_exc()}")

    def _open_supplier_config(self):
        try:
            dlg = QuotationSupplierLibraryDialog(self, self.db, self.C, on_save=self._load_data)
            dlg.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开供方配置失败", f"{e}\n\n{traceback.format_exc()}")

    # ═══════════════════════════════════════════════
    #  导出Excel（只导出勾选的产品）
    # ═══════════════════════════════════════════════

    TEMPLATE_REL = "assets/产品包装报价单_模板.xlsx"

    def _get_template_path(self):
        """获取模板文件路径（支持打包后 _MEIPASS）"""
        if hasattr(sys, '_MEIPASS'):
            p = os.path.join(sys._MEIPASS, self.TEMPLATE_REL)
            if os.path.exists(p):
                return p
        project = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        p = os.path.join(project, self.TEMPLATE_REL)
        if os.path.exists(p):
            return p
        alt = r"E:/采购合同/包材供应商合同/产品包装报价单.xlsx"
        if os.path.exists(alt):
            return alt
        raise FileNotFoundError(
            f"找不到报价单模板文件，请确认存在以下任一位置：\n  {p}\n  {alt}"
        )

    def _get_next_export_seq(self, directory, prefix):
        """获取指定前缀下当天最大的流水号+1"""
        max_idx = 0
        if os.path.isdir(directory):
            for fn in os.listdir(directory):
                if fn.startswith(prefix) and fn.endswith(".xlsx"):
                    # 从文件名末尾提取流水号
                    base = fn[:-5]  # 去掉 .xlsx
                    last_part = base.rsplit("-", 1)[-1]
                    try:
                        if last_part.isdigit():
                            idx = int(last_part)
                            if idx > max_idx:
                                max_idx = idx
                    except ValueError:
                        pass
        return max_idx + 1

    def _export_quotation(self):
        """基于模板导出报价单Excel —— 只导出勾选的产品"""
        if not self.checked_ids:
            messagebox.showinfo("提示", "请先在表格中勾选要导出的产品（点击☐列）\n可使用「全选」按钮一键全选")
            return

        # 只取勾选的产品，保持原有顺序
        export_products = [p for p in self.products if p["id"] in self.checked_ids]
        if not export_products:
            return

        # 获取当前供方信息（从供方库取最新一条或默认）
        suppliers = self.db.get_all_quotation_suppliers()
        current_supplier = suppliers[0] if suppliers else self.db.get_quotation_supplier() or {}
        supplier_name = current_supplier.get("supplier_name", "供应商")

        # 文件名格式：供应商名称_品名_产品报价单_YYYY-MM-DD-流水号
        first_product = export_products[0]
        pname = first_product.get("product_name", "产品")
        date_str = datetime.now().strftime("%Y-%m-%d")
        prefix = f"{supplier_name}_{pname}_产品报价单_{date_str}"

        filepath = filedialog.asksaveasfilename(
            title="保存报价单",
            defaultextension=".xlsx",
            initialfile=f"{prefix}-001.xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            self._generate_excel(filepath, export_products, current_supplier)
            messagebox.showinfo("导出成功", f"报价单已导出到：\n{filepath}\n\n共导出 {len(export_products)} 款产品")
        except Exception as e:
            import traceback
            messagebox.showerror("导出失败", f"{e}\n\n{traceback.format_exc()}")

    def _generate_excel(self, filepath, products=None, supplier=None):

        if products is None:
            products = self.products

        template = self._get_template_path()
        wb = load_workbook(template)
        ws = wb.active
        ws.title = "产品包装报价单"

        cfg = self.config_data or {}
        if supplier is None:
            supplier = self.db.get_quotation_supplier() or {}

        normal_font = Font(name="宋体", size=14)
        thin = Side(style="thin")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # ============================================================
        # 1. 需方信息区
        # ============================================================
        ws["F5"] = cfg.get("buyer_name", "")
        ws["F5"].font = normal_font
        ws["F7"] = cfg.get("buyer_contact", "")
        ws["F7"].font = normal_font
        ws["F9"] = str(cfg.get("buyer_phone", ""))
        ws["F9"].font = normal_font
        ws["F11"] = cfg.get("buyer_address", "")
        ws["F11"].font = normal_font

        # ============================================================
        # 2. 供方信息区
        # ============================================================
        ws["O5"] = supplier.get("supplier_name", "")
        ws["O5"].font = normal_font
        ws["O6"] = supplier.get("contact_person", "")
        ws["O6"].font = normal_font
        ws["O7"] = supplier.get("phone", "")
        ws["O7"].font = normal_font
        ws["O8"] = supplier.get("address", "")
        ws["O8"].font = normal_font
        ws["O9"] = supplier.get("quote_date", "")
        ws["O9"].font = normal_font
        ws["O10"] = supplier.get("quote_validity", "")
        ws["O10"].font = normal_font

        # ============================================================
        # 3. 底部条款区
        # ============================================================
        ws["E21"] = cfg.get("payment_terms", "按协议条件付款；")
        ws["E21"].font = normal_font
        ws["E22"] = cfg.get("transport_method", "物料或者专车请提前说明")
        ws["E22"].font = normal_font
        ws["E23"] = cfg.get("delivery_docs", "请随货放【发货单】【厂检报告】")
        ws["E23"].font = Font(name="宋体", size=14, bold=True, color="FFFF0000")
        ws["E24"] = cfg.get("quote_requirement", "需含税含运")
        ws["E24"].font = normal_font

        # ============================================================
        # 4. 数据行 —— 动态处理（只写入勾选的产品！）
        # ============================================================

        total_rows = 0
        for product in products:
            tiers = product.get("tiers", [])
            total_rows += max(1, len(tiers))

        ORIGINAL_DATA_START = 15
        ORIGINAL_DATA_END = 18
        ORIGINAL_DATA_COUNT = ORIGINAL_DATA_END - ORIGINAL_DATA_START + 1  # =4

        # 解除数据区域原有的所有合并单元格
        merges_to_remove = []
        for mc in list(ws.merged_cells.ranges):
            mc_str = str(mc)
            min_col, min_row, max_col, max_row = range_boundaries(mc_str)
            if min_row >= ORIGINAL_DATA_START and max_row <= ORIGINAL_DATA_END:
                merges_to_remove.append(mc_str)

        for mr in merges_to_remove:
            try:
                ws.unmerge_cells(mr)
            except Exception:
                pass

        extra_rows = total_rows - ORIGINAL_DATA_COUNT
        if extra_rows > 0:
            ws.insert_rows(ORIGINAL_DATA_END + 1, extra_rows)
        elif extra_rows < 0:
            ws.delete_rows(ORIGINAL_DATA_END + extra_rows + 1, -extra_rows)

        # 清理原数据行
        for r in range(ORIGINAL_DATA_START, ORIGINAL_DATA_START + total_rows):
            for c in range(2, 21):
                cell = ws.cell(row=r, column=c)
                cell.value = None

        # 写入产品数据（注意：用 products 参数而非 self.products）
        current_row = ORIGINAL_DATA_START

        for idx, product in enumerate(products, 1):
            tiers = product.get("tiers", [])
            tier_count = max(1, len(tiers))
            first_r = current_row
            last_r = current_row + tier_count - 1

            for r in range(first_r, last_r + 1):
                ws.row_dimensions[r].height = 36

            if tier_count > 1:
                ws.merge_cells(f"B{first_r}:B{last_r}")
                ws.merge_cells(f"C{first_r}:D{last_r}")
                ws.merge_cells(f"E{first_r}:F{last_r}")
                ws.merge_cells(f"G{first_r}:I{last_r}")
                ws.merge_cells(f"J{first_r}:M{last_r}")
                ws.merge_cells(f"N{first_r}:N{last_r}")
                ws.merge_cells(f"O{first_r}:O{last_r}")

            for ti in range(tier_count):
                r = current_row

                if ti == 0:
                    ws.cell(row=r, column=2, value=idx).font = normal_font
                    ws.cell(row=r, column=2).alignment = center
                    ws.cell(row=r, column=2).border = border_all

                    ws.cell(row=r, column=3, value=product.get("item_no", "")).font = normal_font
                    ws.cell(row=r, column=3).alignment = center
                    ws.cell(row=r, column=3).border = border_all

                    ws.cell(row=r, column=5, value=product.get("product_name", "")).font = normal_font
                    ws.cell(row=r, column=5).alignment = left_wrap
                    ws.cell(row=r, column=5).border = border_all

                    ws.cell(row=r, column=7, value=product.get("product_size", "")).font = normal_font
                    ws.cell(row=r, column=7).alignment = center
                    ws.cell(row=r, column=7).border = border_all

                    ws.cell(row=r, column=10, value=product.get("material_process", "")).font = normal_font
                    ws.cell(row=r, column=10).alignment = left_wrap
                    ws.cell(row=r, column=10).border = border_all

                    ws.cell(row=r, column=14, value=product.get("supply_cycle", "")).font = normal_font
                    ws.cell(row=r, column=14).alignment = center
                    ws.cell(row=r, column=14).border = border_all

                    ws.cell(row=r, column=15, value=product.get("carton_spec", "")).font = normal_font
                    ws.cell(row=r, column=15).alignment = center
                    ws.cell(row=r, column=15).border = border_all

                # 阶梯数量 (P:Q)
                if tiers and ti < len(tiers):
                    tier = tiers[ti]
                    mn = int(tier.get("min_qty", 0))
                    mx = tier.get("max_qty")
                    if mx:
                        qty_text = f"{mn}-{int(mx)-1}"
                    else:
                        qty_text = f">={mn}"
                    ws.merge_cells(f"P{r}:Q{r}")
                    ws.cell(row=r, column=16, value=qty_text).font = normal_font
                    ws.cell(row=r, column=16).alignment = center
                    ws.cell(row=r, column=16).border = border_all
                else:
                    ws.merge_cells(f"P{r}:Q{r}")
                    ws.cell(row=r, column=16).alignment = center
                    ws.cell(row=r, column=16).border = border_all

                # 单价 (R:T) — 留空给供方填写
                ws.merge_cells(f"R{r}:T{r}")
                ws.cell(row=r, column=18).alignment = center
                ws.cell(row=r, column=18).border = border_all

                for c_idx in range(2, 21):
                    ws.cell(row=r, column=c_idx).border = border_all

                current_row += 1

        ws.page_setup.orientation = 'landscape'
        wb.save(filepath)


# ============================
# 报价单产品表单弹窗
# ============================

class QuotationForm(ctk.CTkToplevel):
    """产品报价单新增/编辑表单"""

    def __init__(self, parent, db, C, oid=None, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = C
        self.oid = oid
        self.on_save = on_save
        self.tiers = []
        self.title("编辑产品" if oid else "添加产品")
        self.geometry("800x680")
        self.configure(fg_color=self.C["bg"])
        self._build_ui()
        if oid:
            self._load_data()

    def _build_ui(self):
        canvas = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"], corner_radius=0)
        canvas.pack(fill="both", expand=True, padx=16, pady=(12, 8))

        info_frame = ctk.CTkFrame(canvas, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        info_frame.pack(fill="x", pady=(0, 12), padx=4)

        ctk.CTkLabel(
            info_frame, text="📦 产品信息",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(anchor="w", padx=16, pady=(12, 8))

        fields = [
            ("项目号", "item_no_entry"), ("产品名称 *", "name_entry"),
            ("产品尺寸", "size_entry"), ("材质/工艺描述", "process_entry"),
            ("供货周期", "cycle_entry"), ("发货箱规", "carton_entry"), ("单位", "unit_entry"),
        ]
        self.entries = {}
        for label, attr in fields:
            r = ctk.CTkFrame(info_frame, fg_color="transparent")
            r.pack(fill="x", padx=16, pady=3)
            ctk.CTkLabel(r, text=label, width=100, anchor="w",
                         font=ctk.CTkFont(size=13)).pack(side="left")
            ent = ctk.CTkEntry(r, height=32, font=ctk.CTkFont(size=13))
            ent.pack(side="left", fill="x", expand=True, padx=(8, 0))
            self.entries[attr] = ent

        self.entries["unit_entry"].insert(0, "PCS")

        # ── 阶梯价格 ──
        tier_frame = ctk.CTkFrame(canvas, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        tier_frame.pack(fill="x", pady=(0, 12), padx=4)

        tier_header = ctk.CTkFrame(tier_frame, fg_color="transparent")
        tier_header.pack(fill="x", padx=16, pady=(12, 4))

        ctk.CTkLabel(
            tier_header, text="💰 阶梯价格",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left")

        ctk.CTkButton(
            tier_header, text="＋ 添加阶梯", width=100, height=30,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=12), command=self._add_tier_row,
        ).pack(side="right")

        self.tier_container = ctk.CTkFrame(tier_frame, fg_color="transparent")
        self.tier_container.pack(fill="x", padx=16, pady=(4, 12))

        tier_hdr = ctk.CTkFrame(self.tier_container, fg_color=self.C["sidebar"], corner_radius=4)
        tier_hdr.pack(fill="x", pady=(0, 4))
        for text, w in [("最低数量", 120), ("最高数量(留空=无限)", 150), ("", 60)]:
            ctk.CTkLabel(tier_hdr, text=text, width=w, anchor="center",
                         font=ctk.CTkFont(size=11)).pack(side="left", padx=2, pady=2)

        hint = ctk.CTkLabel(
            self.tier_container,
            text="💡 提示：填入阶梯数量后，导出时每个阶梯会独立显示一行。如：100-499 | 500-999 | ≥1000",
            font=ctk.CTkFont(size=11), text_color=self.C.get("text_secondary", "#8B7355"),
        )
        hint.pack(anchor="w", pady=(2, 6))

        # 底部按钮
        btn_frame = ctk.CTkFrame(canvas, fg_color="transparent")
        btn_frame.pack(fill="x", pady=16, padx=4)

        ctk.CTkButton(
            btn_frame, text="💾 保存", width=100, height=38,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._save,
        ).pack(side="right", padx=8)

        ctk.CTkButton(
            btn_frame, text="取消", width=80, height=38,
            fg_color="#9CA3AF", hover_color="#6B7280",
            font=ctk.CTkFont(size=13), command=self.destroy,
        ).pack(side="right", padx=8)

    def _add_tier_row(self, data=None):
        row = ctk.CTkFrame(self.tier_container, fg_color="transparent")
        row.pack(fill="x", pady=2)

        min_entry = ctk.CTkEntry(row, width=120, height=30, font=ctk.CTkFont(size=12),
                                  placeholder_text="最低数量，如：100")
        min_entry.pack(side="left", padx=4)

        max_entry = ctk.CTkEntry(row, width=150, height=30, font=ctk.CTkFont(size=12),
                                  placeholder_text="最高数量，留空=无上限")
        max_entry.pack(side="left", padx=4)

        def _remove():
            row.destroy()
            self.tiers = [t for t in self.tiers if t["row"] is not row]

        del_btn = ctk.CTkButton(row, text="✕", width=40, height=30,
                                fg_color=self.C["danger"], hover_color="#9A5555",
                                font=ctk.CTkFont(size=12), command=_remove)
        del_btn.pack(side="left", padx=4)

        tier_data = {"row": row, "min": min_entry, "max": max_entry}
        if data:
            min_entry.insert(0, str(data.get("min_qty", "")))
            if data.get("max_qty"):
                max_entry.insert(0, str(data["max_qty"]))
        self.tiers.append(tier_data)

    def _load_data(self):
        if not self.oid:
            return
        data = self.db.get_quotation_product(self.oid)
        if not data:
            return
        self.entries["item_no_entry"].insert(0, data.get("item_no", ""))
        self.entries["name_entry"].insert(0, data.get("product_name", ""))
        self.entries["size_entry"].insert(0, data.get("product_size", ""))
        self.entries["process_entry"].insert(0, data.get("material_process", ""))
        self.entries["cycle_entry"].insert(0, data.get("supply_cycle", ""))
        self.entries["carton_entry"].insert(0, data.get("carton_spec", ""))
        self.entries["unit_entry"].delete(0, "end")
        self.entries["unit_entry"].insert(0, data.get("unit", "PCS"))

        for t in data.get("tiers", []):
            self._add_tier_row(t)

    def _save(self):
        name = self.entries["name_entry"].get().strip()
        if not name:
            messagebox.showerror("错误", "请输入产品名称")
            return

        valid_tiers = []
        for t in self.tiers:
            try:
                min_qty = int(t["min"].get().strip())
            except (ValueError, TypeError):
                messagebox.showerror("错误", "最低数量必须为整数")
                return
            max_val = t["max"].get().strip()
            max_qty = int(max_val) if max_val else None
            valid_tiers.append({"min_qty": min_qty, "max_qty": max_qty})

        product_data = {
            "item_no": self.entries["item_no_entry"].get().strip(),
            "product_name": name,
            "product_size": self.entries["size_entry"].get().strip(),
            "material_process": self.entries["process_entry"].get().strip(),
            "supply_cycle": self.entries["cycle_entry"].get().strip(),
            "carton_spec": self.entries["carton_entry"].get().strip(),
            "unit": self.entries["unit_entry"].get().strip() or "PCS",
        }

        try:
            if self.oid:
                self.db.update_quotation_product(self.oid, product_data)
                self.db.delete_quotation_tiers(self.oid)
                pid = self.oid
            else:
                pid = self.db.save_quotation_product(product_data)

            for t in valid_tiers:
                self.db.save_quotation_tier({**t, "product_id": pid})

            messagebox.showinfo("成功", "产品已保存")
            if self.on_save:
                self.on_save()
            self.destroy()
        except Exception as e:
            messagebox.showerror("保存失败", str(e))


# ============================
# 报价单配置弹窗（需方信息）
# ============================

class QuotationConfigDialog(ctk.CTkToplevel):
    """需方信息和报价条款配置"""

    def __init__(self, parent, db, C, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = C
        self.on_save = on_save
        self.title("报价单配置 - 需方信息")
        self.geometry("700x600")
        self.configure(fg_color=self.C["bg"])
        self._build_ui()
        self._load_config()

    def _build_ui(self):
        canvas = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"], corner_radius=0)
        canvas.pack(fill="both", expand=True, padx=16, pady=(12, 8))

        frame = ctk.CTkFrame(canvas, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        frame.pack(fill="x", padx=4)

        ctk.CTkLabel(
            frame, text="⚙ 需方信息与条款配置",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        ).pack(anchor="w", padx=16, pady=(12, 8))

        self.fields = {}
        config_fields = [
            ("需方名称", "buyer_name"), ("联系人", "buyer_contact"),
            ("联系方式", "buyer_phone"), ("送货地址", "buyer_address"),
            ("付款方式", "payment_terms"), ("运输方式", "transport_method"),
            ("发货文件要求", "delivery_docs"), ("报价要求", "quote_requirement"),
            ("模板说明", "quote_template_note"), ("底部注释", "footer_note"),
        ]

        for label, key in config_fields:
            r = ctk.CTkFrame(frame, fg_color="transparent")
            r.pack(fill="x", padx=16, pady=3)
            ctk.CTkLabel(r, text=label, width=110, anchor="w",
                         font=ctk.CTkFont(size=13)).pack(side="left")
            ent = ctk.CTkEntry(r, height=32, font=ctk.CTkFont(size=13))
            ent.pack(side="left", fill="x", expand=True, padx=(8, 0))
            self.fields[key] = ent

        btn_frame = ctk.CTkFrame(canvas, fg_color="transparent")
        btn_frame.pack(fill="x", pady=16, padx=4)

        ctk.CTkButton(btn_frame, text="💾 保存配置", width=110, height=38,
                       fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
                       font=ctk.CTkFont(size=13, weight="bold"),
                       command=self._save).pack(side="right", padx=8)
        ctk.CTkButton(btn_frame, text="取消", width=80, height=38,
                       fg_color="#9CA3AF", hover_color="#6B7280",
                       font=ctk.CTkFont(size=13), command=self.destroy).pack(side="right", padx=8)

    def _load_config(self):
        cfg = self.db.get_quotation_config()
        for key, ent in self.fields.items():
            val = cfg.get(key, "")
            if val:
                ent.insert(0, str(val))

    def _save(self):
        data = {}
        for key, ent in self.fields.items():
            data[key] = ent.get().strip()
        try:
            self.db.update_quotation_config(data)
            messagebox.showinfo("成功", "报价单配置已保存")
            if self.on_save:
                self.on_save()
            self.destroy()
        except Exception as e:
            messagebox.showerror("保存失败", str(e))


# ============================
# 报价单供方配置库弹窗（V2 — 多供应商管理）
# ============================

class QuotationSupplierLibraryDialog(ctk.CTkToplevel):
    """供方配置库 —— 多供应商增删改查，每次保存自动入库"""

    def __init__(self, parent, db, C, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = C
        self.on_save = on_save
        self.suppliers = []
        self.editing_sid = None
        self.title("⚙ 供方配置库")
        self.geometry("800x620")
        self.configure(fg_color=self.C["bg"])
        self._build_ui()
        self._load_list()

    def _build_ui(self):
        # 上半部分：表单区
        form_frame = ctk.CTkScrollableFrame(self, fg_color="transparent", height=260)
        form_frame.pack(fill="x", padx=16, pady=(12, 4))
        form_frame.pack_propagate(False)

        card = ctk.CTkFrame(form_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        card.pack(fill="x", padx=4, pady=(0, 4))

        title_text = "编辑供方" if self.editing_sid else "新增供方"
        self.form_title = ctk.CTkLabel(
            card, text=f"📝 {title_text}",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            text_color=self.C["text"],
        )
        self.form_title.pack(anchor="w", padx=16, pady=(12, 6))

        self.fields = {}
        field_defs = [
            ("供应商名称 *", "supplier_name"), ("联系人", "contact_person"),
            ("联系方式", "phone"), ("地址", "address"),
            ("报价日期", "quote_date"), ("报价有效期", "quote_validity"),
        ]

        # 两列布局
        grid_inner = ctk.CTkFrame(card, fg_color="transparent")
        grid_inner.pack(fill="x", padx=16, pady=(0, 10))

        for i, (label, key) in enumerate(field_defs):
            r = i // 2
            c_mod = i % 2
            row_frame = ctk.CTkFrame(grid_inner, fg_color="transparent")
            if c_mod == 0:
                row_frame.pack(fill="x", pady=2)
            ctk.CTkLabel(row_frame, text=label, width=90, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left")
            ent = ctk.CTkEntry(row_frame, height=30, font=ctk.CTkFont(size=12))
            ent.pack(side="left", fill="x", expand=True, padx=(6, 8 if c_mod == 0 else 0))
            self.fields[key] = ent

        # 表单按钮
        form_btns = ctk.CTkFrame(card, fg_color="transparent")
        form_btns.pack(fill="x", padx=16, pady=(0, 12))

        self.save_btn = ctk.CTkButton(
            form_btns, text="💾 保存供方", width=100, height=32,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=12, weight="bold"), command=self._save_supplier,
        )
        self.save_btn.pack(side="right", padx=4)

        ctk.CTkButton(
            form_btns, text="+ 新增", width=80, height=32,
            fg_color=self.C["danger"], hover_color="#A85A5A",
            font=ctk.CTkFont(size=12), command=self._new_supplier,
        ).pack(side="right", padx=4)

        # 下半部分：已保存的供应商列表
        list_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        list_frame.pack(fill="both", expand=True, padx=16, pady=(4, 16))

        ctk.CTkLabel(
            list_frame, text="📂 已保存的供方记录",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13, weight="bold"),
            text_color=self.C["text"],
        ).pack(anchor="w", padx=16, pady=(10, 6))

        # 列表标题
        hdr = ctk.CTkFrame(list_frame, fg_color=self.C.get("sidebar", "#EEE5DA"), corner_radius=4)
        hdr.pack(fill="x", padx=12, pady=(0, 4))
        for txt, w in [("供应商名称", 160), ("联系人", 100), ("电话", 120), ("地址", 200), ("操作", 120)]:
            ctk.CTkLabel(hdr, text=txt, width=w, anchor="center",
                         font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=4, pady=4)

        # 供应商列表容器
        self.list_container = ctk.CTkScrollableFrame(list_frame, fg_color="transparent")
        self.list_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    def _new_supplier(self):
        """清空表单准备新增"""
        self.editing_sid = None
        self.form_title.configure(text="📝 新增供方")
        self.save_btn.configure(text="💾 保存供方")
        for ent in self.fields.values():
            ent.delete(0, "end")

    def _load_list(self):
        """重新加载供应商列表"""
        self.suppliers = self.db.get_all_quotation_suppliers()
        # 清空列表控件
        for widget in self.list_container.winfo_children():
            widget.destroy()
        if not self.suppliers:
            ctk.CTkLabel(self.list_container, text="暂无供方记录，请使用上方表单新增",
                         font=ctk.CTkFont(size=12), text_color="#999").pack(pady=20)
            return
        for s in self.suppliers:
            self._make_list_row(s)

    def _make_list_row(self, s):
        """创建一行供应商列表项"""
        sid = s["id"]
        row = ctk.CTkFrame(self.list_container, fg_color="transparent", height=36)
        row.pack(fill="x", pady=1)

        # 高亮当前编辑中的
        if self.editing_sid == sid:
            row.configure(fg_color=self.C.get("primary_light", "#F0E0D6"))

        name_lbl = ctk.CTkLabel(row, text=s.get("supplier_name", ""), width=150, anchor="w",
                                  font=ctk.CTkFont(size=11))
        name_lbl.pack(side="left", padx=4)
        ctk.CTkLabel(row, text=s.get("contact_person", ""), width=95, anchor="center",
                     font=ctk.CTkFont(size=11)).pack(side="left", padx=4)
        ctk.CTkLabel(row, text=s.get("phone", ""), width=115, anchor="center",
                     font=ctk.CTkFont(size=11)).pack(side="left", padx=4)
        addr = s.get("address", "") or "-"
        if len(addr) > 18:
            addr = addr[:17] + ".."
        ctk.CTkLabel(row, text=addr, width=195, anchor="w",
                     font=ctk.CTkFont(size=11)).pack(side="left", padx=4)

        btn_area = ctk.CTkFrame(row, fg_color="transparent")
        btn_area.pack(side="left", padx=4)

        ctk.CTkButton(btn_area, text="编辑", width=50, height=26,
                      fg_color="#6B8FA3", hover_color="#5A7A93",
                      font=ctk.CTkFont(size=10),
                      command=lambda sid=sid: self._edit_supplier(sid)).pack(side="left", padx=2)
        ctk.CTkButton(btn_area, text="删除", width=50, height=26,
                      fg_color="#B56A6A", hover_color="#A85A5A",
                      font=ctk.CTkFont(size=10),
                      command=lambda sid=sid: self._delete_supplier(sid)).pack(side="left", padx=2)

    def _edit_supplier(self, sid):
        """加载某条供应商到表单进行编辑"""
        s = next((x for x in self.suppliers if x["id"] == sid), None)
        if not s:
            return
        self.editing_sid = sid
        self.form_title.configure(text=f"📝 编辑供方 (#{sid})")
        self.save_btn.configure(text="💾 更新供方")
        for key, ent in self.fields.items():
            ent.delete(0, "end")
            val = s.get(key, "")
            if val:
                ent.insert(0, str(val))
        self._load_list()  # 刷新高亮状态

    def _delete_supplier(self, sid):
        """删除一条供方记录"""
        s = next((x for x in self.suppliers if x["id"] == sid), None)
        name = s.get("supplier_name", "") if s else ""
        if messagebox.askyesno("确认删除", f"确定删除供方「{name}」？", icon="warning"):
            self.db.delete_quotation_supplier_record(sid)
            if self.editing_sid == sid:
                self._new_supplier()
            self._load_list()

    def _save_supplier(self):
        """保存当前表单（新增或更新）"""
        name = self.fields["supplier_name"].get().strip()
        if not name:
            messagebox.showerror("错误", "请输入供应商名称")
            return

        data = {key: ent.get().strip() for key, ent in self.fields.items()}

        try:
            if self.editing_sid:
                self.db.update_quotation_supplier_record(self.editing_sid, data)
                messagebox.showinfo("成功", f"供方「{name}」已更新")
            else:
                new_id = self.db.save_quotation_supplier_record(data)
                messagebox.showinfo("成功", f"供方「{name}」已入库 (ID:{new_id})")

            self._new_supplier()
            self._load_list()
            if self.on_save:
                self.on_save()
        except Exception as e:
            messagebox.showerror("保存失败", str(e))
