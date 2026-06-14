#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""供应商管理页面 - v1.2"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import customtkinter as ctk
from datetime import datetime


CATEGORIES = ["礼盒", "卡盒", "标签", "玻璃瓶", "复合膜", "铝制品", "塑料罐", "物流箱", "其他"]
COOP_STATUS = ["合作中", "接洽中", "打样中"]

# 列定义：(字段key, 显示名, 宽度, 对齐)
TABLE_COLS = [
    ("name",               "名称",     140, "w"),
    ("cooperation_status", "合作状态",  80, "center"),
    ("main_product",       "主营",     130, "w"),
    ("contact_person",     "联系人",    80, "center"),
    ("phone",              "电话",     110, "center"),
    ("wechat",             "微信",     110, "center"),
    ("quote_status",       "询比价",    80, "center"),
    ("sample_status",      "打样",      70, "center"),
    ("payment_method",     "付款方式",  90, "center"),
    ("invoice_type",       "开票类型",  90, "center"),
    ("tax_rate",           "税率",      60, "center"),
]


class SupplierPage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.db = db
        self.C = colors
        self._build()
        # 初始不搜索、不显示数据

    def _build(self):
        # ── 顶部工具栏（去掉内部标题，按钮靠左+去图标）──
        header = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0, height=44)
        header.pack(fill="x", padx=20, pady=(12, 8))
        header.pack_propagate(False)

        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="left", padx=0, pady=5)

        ctk.CTkButton(
            btn_frame, text="新增", width=80, height=34,
            fg_color="transparent", hover_color="#A85A5A",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._open_form, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="导出", width=80, height=34,
            fg_color="transparent", hover_color="#7A9A6E",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._export_xlsx, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="导入", width=80, height=34,
            fg_color="transparent", hover_color="#4B5563",
            font=ctk.CTkFont(size=14),
            command=self._import_xlsx, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # ── 搜索栏 ─────────────────────────────────────
        search_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"], height=56)
        search_frame.pack(fill="x", padx=16, pady=(12, 0))
        search_frame.pack_propagate(False)

        inner = ctk.CTkFrame(search_frame, fg_color="transparent")
        inner.pack(side="left", fill="y", padx=16)

        # 供应商类别下拉
        ctk.CTkLabel(inner, text="供应商类别：",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left", pady=12)
        self.cat_var = tk.StringVar(value="全部")
        self.cat_combo = ctk.CTkComboBox(
            inner, values=["全部"] + CATEGORIES, variable=self.cat_var,
            width=110, height=32, font=ctk.CTkFont(size=13),
        )
        self.cat_combo.pack(side="left", padx=(4, 16), pady=12)

        # 供应商名称搜索框
        ctk.CTkLabel(inner, text="供应商名称：",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left")
        self.kw_var = tk.StringVar()
        kw_entry = ctk.CTkEntry(
            inner, textvariable=self.kw_var, width=180, height=32,
            font=ctk.CTkFont(size=13), placeholder_text="模糊搜索…",
        )
        kw_entry.pack(side="left", padx=(4, 16))
        kw_entry.bind("<Return>", lambda e: self._do_search())

        # 合作状态下拉 (v1.2 新增)
        ctk.CTkLabel(inner, text="合作状态：",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left")
        self.coop_var = tk.StringVar(value="全部")
        self.coop_combo = ctk.CTkComboBox(
            inner, values=["全部"] + COOP_STATUS, variable=self.coop_var,
            width=100, height=32, font=ctk.CTkFont(size=13),
        )
        self.coop_combo.pack(side="left", padx=(4, 16))

        # 查询 & 重置
        ctk.CTkButton(
            inner, text="查询", width=70, height=32,
            fg_color="transparent", hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._do_search, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            inner, text="重置", width=60, height=32,
            fg_color=self.C["border"],
            hover_color="#CBD5E1", font=ctk.CTkFont(size=13),
            command=self._reset, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # 统计
        self.count_lbl = ctk.CTkLabel(
            search_frame, text="",
            font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"],
        )
        self.count_lbl.pack(side="right", padx=20)

        # ── 表格区域 - 阴影分层：z1 → z2 → z3 ────────────────────────────────────
        # z1（底层）：fg_color=#E8E2D9, corner_radius=16, 内边距 4
        z1_frame = ctk.CTkFrame(self, fg_color="transparent", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=16, pady=(8, 16))

        # z2（中层）：fg_color=#F2F0EB, corner_radius=8, 内边距 4
        z2_frame = ctk.CTkFrame(z1_frame, fg_color="transparent", corner_radius=8, border_width=1, border_color="#8B7D6B")
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # z3（上层卡片）：fg_color=C["card"], corner_radius=C["radius_card"]
        list_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        list_frame.pack(fill="both", expand=True, padx=4, pady=4)

        style = ttk.Style()
        style.configure("Supplier.Treeview",
                         font=("Microsoft YaHei", 9), rowheight=36,
                         background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#1E293B")
        style.configure("Supplier.Treeview.Heading",
                         font=("Microsoft YaHei", 9, "bold"),
                         background="#F8FAFC", foreground="#475569", relief="flat")
        style.map("Supplier.Treeview", background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])

        tree_frame = tk.Frame(list_frame, bg="#FFFFFF")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        col_ids = [c[0] for c in TABLE_COLS] + ["action"]
        self.tree = ttk.Treeview(
            tree_frame, style="Supplier.Treeview",
            columns=col_ids, show="headings", selectmode="browse",
        )
        for cid, label, width, anchor in TABLE_COLS:
            self.tree.heading(cid, text=label)
            self.tree.column(cid, width=width, minwidth=40, stretch=True, anchor=anchor)
        self.tree.heading("action", text="操作")
        self.tree.column("action", width=120, minwidth=40, stretch=True, anchor="center")

        vsb = ctk.CTkScrollbar(tree_frame, orientation="vertical", command=self.tree.yview, button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8)
        hsb = ctk.CTkScrollbar(tree_frame, orientation="horizontal", command=self.tree.xview, button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8, height=8)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.tag_configure("odd", background="#F8FAFC")
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)
        self.tree.bind("<Button-1>", self._on_click)
        self.tree.bind("<Double-1>", self._on_double_click)

    # ── 搜索 / 重置 ────────────────────────────────────
    # ── 单击选中高亮行 ────────────────────────────
    def _on_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        self.tree.selection_set(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == len(TABLE_COLS):
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                sid = int(item)
                if x_rel < col_w * 0.5:
                    self._open_form(sid)
                else:
                    self._delete(sid)
            return
        sid = int(item)
        self._open_form(sid)

    def _on_hover(self, event):
        """鼠标悬浮时高亮行背景"""
        item = self.tree.identify_row(event.y)
        if item and item != getattr(self, "_last_hover", None):
            if hasattr(self, "_last_hover") and self._last_hover:
                tags = list(self.tree.item(self._last_hover, "tags")[0])
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(self._last_hover, tags=tags)
            tags = list(self.tree.item(item, "tags")[0])
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tags)
            self._last_hover = item
        elif not item and hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_leave(self, event):
        """鼠标离开时清除悬浮高亮"""
        if hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _do_search(self):
        cat = self.cat_var.get()
        kw = self.kw_var.get().strip()
        coop = self.coop_var.get()
        records = self.db.get_suppliers(
            category=cat if cat != "全部" else None,
            keyword=kw if kw else None,
            cooperation_status=coop if coop != "全部" else None,
        )
        self._render(records)
        self.count_lbl.configure(text=f"共 {len(records)} 条")

    def _reset(self):
        self.cat_var.set("全部")
        self.kw_var.set("")
        self.coop_var.set("全部")
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.count_lbl.configure(text="")

    def _render(self, records):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for i, r in enumerate(records):
            tag = "odd" if i % 2 == 0 else "even"
            self.tree.insert("", "end", iid=str(r["id"]),
                             values=(
                                 r.get("name") or "—",
                                 r.get("cooperation_status") or "—",
                                 r.get("main_product") or "—",
                                 r.get("contact_person") or "—",
                                 r.get("phone") or "—",
                                 r.get("wechat") or "—",
                                 r.get("quote_status") or "—",
                                 r.get("sample_status") or "—",
                                 r.get("payment_method") or "—",
                                 r.get("invoice_type") or "—",
                                 r.get("tax_rate") or "—",
                                 "编辑  删除",
                             ), tags=(tag,))

    # ── 双击操作 ───────────────────────────────────────
    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        sid = int(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == len(TABLE_COLS):  # 操作列
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                if x_rel < col_w * 0.5:
                    self._open_form(sid)
                else:
                    self._delete(sid)
        else:
            self._open_form(sid)

    # ── 新增 / 编辑 ────────────────────────────────────
    def _open_form(self, sid=None):
        try:
            form = SupplierForm(self, self.db, self.C, sid=sid, on_save=self._do_search)
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    def _delete(self, sid):
        rec = self.db.get_supplier(sid)
        name = rec.get("name", "") if rec else ""
        if messagebox.askyesno("删除确认", f"确定删除供应商「{name}」？", icon="warning"):
            self.db.delete_supplier(sid)
            self._do_search()

    # ── 导出 Excel ─────────────────────────────────────
    def _export_xlsx(self):
        cat = self.cat_var.get()
        kw = self.kw_var.get().strip()
        records = self.db.get_suppliers(
            category=cat if cat != "全部" else None,
            keyword=kw if kw else None,
        )
        if not records:
            messagebox.showinfo("提示", "没有可导出的数据，请先查询")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出供应商列表", defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"供应商_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )
        if not filepath:
            return

        headers = ["名称", "合作状态", "供应商类别", "主营", "联系人", "电话", "微信",
                   "询比价", "打样", "付款方式", "开票类型", "税率", "备注"]
        rows = []
        for r in records:
            rows.append({
                "名称": r.get("name") or "",
                "合作状态": r.get("cooperation_status") or "",
                "供应商类别": r.get("category") or "",
                "主营": r.get("main_product") or "",
                "联系人": r.get("contact_person") or "",
                "电话": r.get("phone") or "",
                "微信": r.get("wechat") or "",
                "询比价": r.get("quote_status") or "",
                "打样": r.get("sample_status") or "",
                "付款方式": r.get("payment_method") or "",
                "开票类型": r.get("invoice_type") or "",
                "税率": r.get("tax_rate") or "",
                "备注": r.get("remark") or "",
            })
        try:
            self.db.export_to_xlsx(filepath, "供应商", headers, rows,
                                   col_widths=[16, 10, 16, 8, 13, 13, 8, 8, 10, 10, 6, 20])
            messagebox.showinfo("导出成功", f"已导出 {len(rows)} 条记录到：\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── 导入 Excel ─────────────────────────────────────
    def _import_xlsx(self):
        filepath = filedialog.askopenfilename(
            title="选择供应商xlsx文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not filepath:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            field_map = {
                "名称": "name", "合作状态": "cooperation_status", "供应商类别": "category", "主营": "main_product",
                "联系人": "contact_person", "电话": "phone", "微信": "wechat",
                "询比价": "quote_status", "打样": "sample_status",
                "付款方式": "payment_method", "开票类型": "invoice_type",
                "税率": "tax_rate", "备注": "remark",
            }
            col_index = {}
            for i, h in enumerate(headers):
                if h in field_map:
                    col_index[field_map[h]] = i

            if "name" not in col_index:
                messagebox.showerror("导入失败", "Excel中未找到「名称」列，请检查表头")
                return

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row):
                    continue
                def get_val(field):
                    idx = col_index.get(field)
                    if idx is None:
                        return ""
                    v = row[idx]
                    return str(v).strip() if v is not None else ""

                name = get_val("name")
                if not name:
                    continue
                data = {
                    "name": name,
                    "category": get_val("category"),
                    "main_product": get_val("main_product"),
                    "contact_person": get_val("contact_person"),
                    "phone": get_val("phone"),
                    "wechat": get_val("wechat"),
                    "cooperation_status": get_val("cooperation_status") or "接洽中",
                    "quote_status": get_val("quote_status"),
                    "sample_status": get_val("sample_status"),
                    "payment_method": get_val("payment_method"),
                    "invoice_type": get_val("invoice_type"),
                    "tax_rate": get_val("tax_rate"),
                    "remark": get_val("remark"),
                }
                self.db.save_supplier(data)
                count += 1

            messagebox.showinfo("导入成功", f"成功导入 {count} 条供应商数据")
            self._do_search()
        except ImportError:
            messagebox.showerror("缺少依赖", "请安装 openpyxl：pip install openpyxl")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))


# ============================
# 供应商表单弹窗
# ============================
class SupplierForm(ctk.CTkToplevel):
    PAYMENT_OPTS = ["款到发货", "货到付款", "30天账期", "60天账期", "月结", "其他"]
    INVOICE_OPTS = ["增值税专票", "增值税普票", "无需发票", "其他"]
    QUOTE_OPTS   = ["已询价", "比价中", "已确认", "—"]
    SAMPLE_OPTS  = ["未打样", "打样中", "已确认", "—"]

    def __init__(self, parent, db, colors, sid=None, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = colors
        self.sid = sid
        self.on_save = on_save

        self.title("编辑供应商" if sid else "新增供应商")
        self.geometry("720x680")
        self.resizable(True, True)
        self.configure(fg_color=self.C["bg"])

        self.update_idletasks()
        pw, ph = parent.winfo_rootx(), parent.winfo_rooty()
        self.geometry(f"720x680+{pw+80}+{ph+40}")

        self._build()
        if sid:
            self._load(sid)

    def _build(self):
        outer = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"])
        outer.pack(fill="both", expand=True)

        card = ctk.CTkFrame(outer, fg_color=self.C["card"], corner_radius=self.C["radius_modal"])
        card.pack(fill="both", expand=True, padx=16, pady=16)

        def row(parent, label, widget_fn, **kw):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=16, pady=4)
            ctk.CTkLabel(f, text=label, width=90, anchor="w",
                         font=ctk.CTkFont(size=13),
                         text_color=self.C["text_secondary"]).pack(side="left")
            widget_fn(f, **kw)

        def entry(parent, var, ph=""):
            e = ctk.CTkEntry(parent, textvariable=var, height=32,
                             font=ctk.CTkFont(size=13), placeholder_text=ph)
            e.pack(side="left", fill="x", expand=True, padx=(8, 0))

        def combo(parent, var, values, width=180):
            c = ctk.CTkComboBox(parent, variable=var, values=values,
                                width=width, height=32, font=ctk.CTkFont(size=13))
            c.pack(side="left", padx=(8, 0))

        # ── 字段 ──
        self.v_name    = tk.StringVar()
        self.v_cat     = tk.StringVar()
        self.v_coop    = tk.StringVar(value="接洽中")
        self.v_main    = tk.StringVar()
        self.v_contact = tk.StringVar()
        self.v_phone   = tk.StringVar()
        self.v_wechat  = tk.StringVar()
        self.v_quote   = tk.StringVar(value="—")
        self.v_sample  = tk.StringVar(value="—")
        self.v_payment = tk.StringVar()
        self.v_invoice = tk.StringVar()
        self.v_tax     = tk.StringVar()

        self._section(card, "📋 基本信息")
        row(card, "名称 *",     lambda p, **k: entry(p, self.v_name,    "供应商全称"), )
        row(card, "供应商类别", lambda p, **k: combo(p, self.v_cat,     [""] + CATEGORIES, 160))
        row(card, "合作状态",   lambda p, **k: combo(p, self.v_coop,    COOP_STATUS, 160))
        row(card, "主营产品",   lambda p, **k: entry(p, self.v_main,    "主营产品描述"))

        self._section(card, "👤 联系方式")
        row(card, "联系人",     lambda p, **k: entry(p, self.v_contact, "姓名"))
        row(card, "电话",       lambda p, **k: entry(p, self.v_phone,   "手机/座机"))
        row(card, "微信",       lambda p, **k: entry(p, self.v_wechat,  "微信号"))

        self._section(card, "💼 合作信息")
        row(card, "询比价",     lambda p, **k: combo(p, self.v_quote,   self.QUOTE_OPTS, 160))
        row(card, "打样",       lambda p, **k: combo(p, self.v_sample,  self.SAMPLE_OPTS, 160))
        row(card, "付款方式",   lambda p, **k: combo(p, self.v_payment, self.PAYMENT_OPTS, 180))
        row(card, "开票类型",   lambda p, **k: combo(p, self.v_invoice, self.INVOICE_OPTS, 200))
        row(card, "税率",       lambda p, **k: entry(p, self.v_tax,     "如：13%"))

        self._section(card, "📝 备注")
        remark_frame = ctk.CTkFrame(card, fg_color="transparent")
        remark_frame.pack(fill="x", padx=16, pady=4)
        self.remark_text = ctk.CTkTextbox(remark_frame, height=80, font=ctk.CTkFont(size=13))
        self.remark_text.pack(fill="x")

        # 按钮
        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=12)
        ctk.CTkButton(btn_row, text="✓ 保存", width=120, height=38,
                      fg_color="transparent", hover_color=self.C["primary_hover"],
                      font=ctk.CTkFont(size=16, weight="bold"),
                      command=self._save, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="✕ 关闭", width=80, height=38,
                      fg_color="transparent", hover_color="#4B5563",
                      font=ctk.CTkFont(size=14), command=self.destroy, corner_radius=8, border_width=1, border_color="#8B7D6B", text_color="#000000",).pack(side="left", padx=4)

    def _section(self, parent, title):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=self.C["primary"]).pack(side="left")
        ctk.CTkFrame(f, height=1, fg_color=self.C["border"]).pack(
            side="left", fill="x", expand=True, padx=(8, 0))

    def _load(self, sid):
        r = self.db.get_supplier(sid)
        if not r:
            return
        self.v_name.set(r.get("name") or "")
        self.v_cat.set(r.get("category") or "")
        self.v_coop.set(r.get("cooperation_status") or "接洽中")
        self.v_main.set(r.get("main_product") or "")
        self.v_contact.set(r.get("contact_person") or "")
        self.v_phone.set(r.get("phone") or "")
        self.v_wechat.set(r.get("wechat") or "")
        self.v_quote.set(r.get("quote_status") or "—")
        self.v_sample.set(r.get("sample_status") or "—")
        self.v_payment.set(r.get("payment_method") or "")
        self.v_invoice.set(r.get("invoice_type") or "")
        self.v_tax.set(r.get("tax_rate") or "")
        if r.get("remark"):
            self.remark_text.insert("1.0", r["remark"])

    def _save(self):
        name = self.v_name.get().strip()
        if not name:
            messagebox.showerror("验证失败", "请填写供应商名称", parent=self)
            return
        data = {
            "name": name,
            "category": self.v_cat.get().strip(),
            "cooperation_status": self.v_coop.get().strip(),
            "main_product": self.v_main.get().strip(),
            "contact_person": self.v_contact.get().strip(),
            "phone": self.v_phone.get().strip(),
            "wechat": self.v_wechat.get().strip(),
            "quote_status": self.v_quote.get().strip(),
            "sample_status": self.v_sample.get().strip(),
            "payment_method": self.v_payment.get().strip(),
            "invoice_type": self.v_invoice.get().strip(),
            "tax_rate": self.v_tax.get().strip(),
            "remark": self.remark_text.get("1.0", "end").strip(),
        }
        if self.sid:
            self.db.update_supplier(self.sid, data)
            messagebox.showinfo("成功", "供应商信息已更新", parent=self)
        else:
            self.db.save_supplier(data)
            messagebox.showinfo("成功", "供应商已添加", parent=self)
        if self.on_save:
            self.on_save()
        self.destroy()
