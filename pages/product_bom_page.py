#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""成品BOM页面 V2.2.7 - 重构：顶栏三按钮 + 筛选顺序 + 多物料新增 + 台账风格表格 + 附件模板导入"""

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk


# ── 资源路径（兼容 PyInstaller 打包）──
def _get_resource_path(rel_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, rel_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel_path)


# BOM 模板路径（assets/bom__template.xlsx）
BOM_TEMPLATE_PATH = _get_resource_path("assets/bom__template.xlsx")


class ProductBomPage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.C = colors
        self.db = db
        self._all_data = []
        self._sort_col = None
        self._sort_asc = True
        self._build()
        # 默认不显示数据，需点查询
        self._load_from_db(show=False)

    # ───────────────────────────────────────────────
    #  构建 UI
    # ───────────────────────────────────────────────
    def _build(self):
        # ── 顶部工具栏（去掉内部标题，按钮靠左+去图标）──
        header = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0, height=44)
        header.pack(fill="x", padx=20, pady=(12, 8))
        header.pack_propagate(False)

        # 三个按钮靠左排列，去掉按钮内图标
        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="left", padx=0, pady=5)

        ctk.CTkButton(
            btn_frame, text="导出表格", width=100, height=34,
            fg_color="transparent", hover_color="#B89A5D",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            command=self._on_export, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="导入表格", width=100, height=34,
            fg_color="transparent", hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            command=self._on_import, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="新增BOM", width=100, height=34,
            fg_color="transparent", hover_color="#7A9472",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            command=self._on_add, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=4)

        # ── 筛选区 ─────────────────────────────
        filter_card = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        filter_card.pack(fill="x", padx=16, pady=(8, 8))

        filter_inner = ctk.CTkFrame(filter_card, fg_color="transparent")
        filter_inner.pack(fill="x", padx=16, pady=12)

        # 第一行：物料名称（模糊） + 物料项目号 + 品名（模糊） + 成品项目号
        row1 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row1.pack(fill="x", pady=4)

        ctk.CTkLabel(row1, text="物料名称", font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                     text_color=self.C["text_secondary"]).pack(side="left")
        self.material_name_entry = ctk.CTkEntry(
            row1, width=160, height=34,
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            placeholder_text="支持模糊搜索",
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.material_name_entry.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(row1, text="物料项目号", font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                     text_color=self.C["text_secondary"]).pack(side="left")
        self.material_project_no_entry = ctk.CTkEntry(
            row1, width=150, height=34,
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.material_project_no_entry.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(row1, text="品名", font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                     text_color=self.C["text_secondary"]).pack(side="left")
        self.product_name_entry = ctk.CTkEntry(
            row1, width=160, height=34,
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            placeholder_text="支持模糊搜索",
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.product_name_entry.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(row1, text="成品项目号", font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                     text_color=self.C["text_secondary"]).pack(side="left")
        self.finished_project_no_entry = ctk.CTkEntry(
            row1, width=150, height=34,
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.finished_project_no_entry.pack(side="left", padx=(4, 16))
        self.finished_project_no_entry.bind("<Return>", lambda e: self._on_query())

        # 第二行：查询 + 重置
        row2 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row2.pack(fill="x", pady=(8, 0))

        ctk.CTkButton(
            row2, text="🔍  查询", width=80, height=34,
            fg_color="transparent", hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            command=self._on_query, corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row2, text="🔄  重置", width=80, height=34,
            fg_color="transparent",
            hover_color="#E2E8F0", border_width=1, border_color=self.C["border"],
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            command=self._on_reset, corner_radius=8,
 text_color="#000000",).pack(side="left")

        # 统计栏
        stats_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=self.C["radius_card"], height=48)
        stats_frame.pack(fill="x", padx=16, pady=(0, 8))
        stats_frame.pack_propagate(False)
        self.result_count_lbl = ctk.CTkLabel(
            stats_frame, text="请设置筛选条件后点击「查询」",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            text_color=self.C["text_secondary"],
        )
        self.result_count_lbl.pack(side="left", padx=20, pady=10)

        # ── 表格区（台账查询风格：z1 → z2 → z3 分层） ──
        z1_frame = ctk.CTkFrame(self, fg_color="transparent", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        z2_frame = ctk.CTkFrame(z1_frame, fg_color="transparent", corner_radius=8, border_width=1, border_color="#8B7D6B")
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        table_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        table_frame.pack(fill="both", expand=True, padx=4, pady=4)
        self._build_table(table_frame)

    def _build_table(self, parent):
        """台账查询风格表格：列排序、斑马纹、悬浮高亮、单击选中"""
        self.columns = [
            ("finished_project_no", "成品项目号", 130, "center"),
            ("product_name",        "品名",        180, "w"),
            ("spec",                "规格",        180, "w"),
            ("brand",               "品牌",        110, "center"),
            ("material_project_no", "物料项目号",  130, "center"),
            ("material_name",       "物料名称",    200, "w"),
            ("quantity",            "数量",        80,  "center"),
            ("unit",                "计量单位",    80,  "center"),
        ]

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Bom.Treeview",
                        font=("Microsoft YaHei", 9),
                        rowheight=36,
                        background="#FFFFFF",
                        fieldbackground="#FFFFFF",
                        foreground="#1E293B",
                        borderwidth=0,
                        relief="flat")
        style.configure("Bom.Treeview.Heading",
                        font=("Microsoft YaHei", 9, "bold"),
                        background="#F8FAFC",
                        foreground="#475569",
                        relief="flat",
                        borderwidth=0)
        style.map("Bom.Treeview",
                  background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])
        style.layout("Bom.Treeview", [("Treeview.treearea", {"sticky": "nswe"})])

        tree_wrap = tk.Frame(parent, bg="#FFFFFF")
        tree_wrap.pack(fill="both", expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(
            tree_wrap, style="Bom.Treeview",
            columns=[c[0] for c in self.columns],
            show="headings", selectmode="browse",
        )
        for cid, label, width, anchor in self.columns:
            self.tree.heading(cid, text=label,
                              command=lambda c=cid: self._sort_by_col(c))
            self.tree.column(cid, width=width, minwidth=50, stretch=True, anchor=anchor)

        vsb = ctk.CTkScrollbar(tree_wrap, orientation="vertical", command=self.tree.yview,
                               button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8)
        hsb = ctk.CTkScrollbar(tree_wrap, orientation="horizontal", command=self.tree.xview,
                               button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8, height=8)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.tag_configure("odd",   background="#F8FAFC")
        self.tree.tag_configure("even",  background="#FFFFFF")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)
        self.tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self.tree.bind("<Double-1>", self._on_double_click)

    # ───────────────────────────────────────────────
    #  表格交互
    # ───────────────────────────────────────────────
    def _on_tree_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

    def _on_hover(self, event):
        item = self.tree.identify_row(event.y)
        if item and item != getattr(self, "_last_hover", None):
            if hasattr(self, "_last_hover") and self._last_hover:
                tags = list(self.tree.item(self._last_hover, "tags"))
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(self._last_hover, tags=tags)
            tags = list(self.tree.item(item, "tags"))
            if "hover" not in tags:
                tags.append("hover")
            self.tree.item(item, tags=tags)
            self._last_hover = item
        elif not item and hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_leave(self, event):
        if hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_double_click(self, event):
        item = self.tree.selection()
        if not item:
            return
        idx = self.tree.index(item[0])
        if idx >= len(self._all_data):
            return
        record = self._all_data[idx]
        self._show_edit_dialog(record)

    # ───────────────────────────────────────────────
    #  数据加载 / 查询 / 重置
    # ───────────────────────────────────────────────
    def _load_from_db(self, show=True):
        rows = self.db.get_product_bom()
        if rows:
            self._all_data = rows
        if show:
            self._refresh_table()

    def _on_query(self):
        material_name = self.material_name_entry.get().strip()
        material_project_no = self.material_project_no_entry.get().strip()
        product_name = self.product_name_entry.get().strip()
        finished_project_no = self.finished_project_no_entry.get().strip()

        data = self.db.get_product_bom(
            product_name=product_name if product_name else None,
            finished_project_no=finished_project_no if finished_project_no else None,
            material_project_no=material_project_no if material_project_no else None,
            material_name=material_name if material_name else None,
        )
        self._all_data = data
        self._refresh_table()

    def _on_reset(self):
        self.material_name_entry.delete(0, "end")
        self.material_project_no_entry.delete(0, "end")
        self.product_name_entry.delete(0, "end")
        self.finished_project_no_entry.delete(0, "end")
        self._all_data = []
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.result_count_lbl.configure(
            text="请设置筛选条件后点击「查询」",
            text_color=self.C.get("text_secondary", "#5D5D5D"),
        )

    # ───────────────────────────────────────────────
    #  排序 / 渲染
    # ───────────────────────────────────────────────
    def _sort_by_col(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True

        def sort_key(r):
            v = r.get(col, "")
            if col == "quantity":
                try:
                    return float(v or 0)
                except (ValueError, TypeError):
                    return 0
            return str(v or "").lower()

        self._all_data.sort(key=sort_key, reverse=not self._sort_asc)
        self._refresh_table()

    def _refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        data = self._all_data
        if not data:
            self.result_count_lbl.configure(
                text="请设置筛选条件后点击「查询」",
                text_color=self.C.get("text_secondary", "#5D5D5D"),
            )
            return

        for i, d in enumerate(data):
            tag = "odd" if i % 2 == 0 else "even"
            qty = d.get("quantity") or 0
            qty_str = f"{qty:g}" if qty else "—"
            self.tree.insert("", "end", values=(
                d.get("finished_project_no") or "—",
                d.get("product_name") or "—",
                d.get("spec") or "—",
                d.get("brand") or "—",
                d.get("material_project_no") or "—",
                d.get("material_name") or "—",
                qty_str,
                d.get("unit") or "—",
            ), tags=(tag,))

        self.result_count_lbl.configure(
            text=f"共 {len(data)} 条记录", text_color=self.C["primary"]
        )

    # ───────────────────────────────────────────────
    #  导入（附件模板）
    # ───────────────────────────────────────────────
    def _on_import(self):
        """导入表格：弹出文件选择框，默认指向附件 bom__template.xlsx"""
        if not os.path.exists(BOM_TEMPLATE_PATH):
            messagebox.showwarning("模板缺失", f"未找到BOM模板文件：\n{BOM_TEMPLATE_PATH}")
            filepath = filedialog.askopenfilename(
                title="选择BOM表格（Excel）",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            )
            if not filepath:
                return
        else:
            initial_dir = os.path.dirname(BOM_TEMPLATE_PATH)
            initial_file = os.path.basename(BOM_TEMPLATE_PATH)
            filepath = filedialog.askopenfilename(
                title="选择BOM表格（附件模板）",
                initialdir=initial_dir,
                initialfile=initial_file,
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            )
            if not filepath:
                return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active

            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                rows.append({
                    "finished_project_no": str(row[0]).strip() if row[0] else "",
                    "product_name":        str(row[1]).strip() if row[1] else "",
                    "spec":                str(row[2]).strip() if row[2] else "",
                    "retail_price":        self._to_float(row[3]),
                    "brand":               str(row[4]).strip() if row[4] else "",
                    "material_project_no": str(row[5]).strip() if row[5] else "",
                    "material_name":       str(row[6]).strip() if row[6] else "",
                    "quantity":            self._to_float(row[7]),
                    "unit":                str(row[8]).strip() if row[8] else "",
                })

            if not rows:
                messagebox.showwarning("提示", "表格中未找到有效数据")
                return

            self.db.import_product_bom(rows)
            messagebox.showinfo("导入成功", f"成功导入 {len(rows)} 条BOM数据")
            self._on_query()
        except Exception as e:
            messagebox.showerror("导入失败", f"导入表格时出错：\n{e}")

    @staticmethod
    def _to_float(v):
        try:
            if v is None or str(v).strip() == "":
                return 0.0
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    # ───────────────────────────────────────────────
    #  导出
    # ───────────────────────────────────────────────
    def _on_export(self):
        if not self._all_data:
            messagebox.showwarning("提示", "当前没有数据可导出，请先查询")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出BOM表格",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="成品BOM导出.xlsx",
        )
        if not filepath:
            return

        try:
            headers = [c[1] for c in self.columns]
            rows = []
            for d in self._all_data:
                rows.append({
                    "成品项目号": d.get("finished_project_no", ""),
                    "品名":       d.get("product_name", ""),
                    "规格":       d.get("spec", ""),
                    "品牌":       d.get("brand", ""),
                    "物料项目号": d.get("material_project_no", ""),
                    "物料名称":   d.get("material_name", ""),
                    "数量":       d.get("quantity", 0),
                    "计量单位":   d.get("unit", ""),
                })
            self.db.export_to_xlsx(filepath, "成品BOM", headers, rows,
                                   col_widths=[14, 20, 20, 12, 14, 20, 8, 10])
            messagebox.showinfo("导出成功", f"成功导出 {len(rows)} 条记录")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出表格时出错：\n{e}")

    # ───────────────────────────────────────────────
    #  新增 / 编辑（多物料）
    # ───────────────────────────────────────────────
    def _on_add(self):
        self._show_edit_dialog({})

    def _show_edit_dialog(self, record):
        """新增/编辑对话框：成品信息4个固定字段 + 物料列表（可多行）"""
        is_new = not record.get("id")
        dialog = tk.Toplevel(self)
        dialog.title("新增BOM" if is_new else "编辑BOM")
        dialog.geometry("780x640")
        dialog.resizable(True, True)
        dialog.minsize(720, 560)
        dialog.configure(bg=self.C["card"])
        dialog.transient(self)
        dialog.grab_set()

        # 居中
        dialog.update_idletasks()
        dw, dh = 780, 640
        sw = dialog.winfo_screenwidth()
        sh = dialog.winfo_screenheight()
        dialog.geometry(f"{dw}x{dh}+{(sw-dw)//2}+{(sh-dh)//2}")

        # ── 成品信息（固定字段）──
        head = ctk.CTkLabel(
            dialog, text="成品信息",
            font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
            text_color=self.C["text"],
        )
        head.pack(anchor="w", padx=20, pady=(16, 6))

        head_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        head_frame.pack(fill="x", padx=20)

        # 第一行：成品项目号 + 品名
        head_row1 = ctk.CTkFrame(head_frame, fg_color="transparent")
        head_row1.pack(fill="x", pady=4)

        ctk.CTkLabel(head_row1, text="成品项目号", width=80, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     text_color=self.C["text_secondary"]).pack(side="left")
        finished_var = tk.StringVar(value=str(record.get("finished_project_no", "")) if record.get("finished_project_no") else "")
        ctk.CTkEntry(head_row1, textvariable=finished_var, width=240, height=32,
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     fg_color=self.C["bg"], border_color=self.C["border"]).pack(side="left", padx=(6, 24))

        ctk.CTkLabel(head_row1, text="品名", width=60, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     text_color=self.C["text_secondary"]).pack(side="left")
        product_var = tk.StringVar(value=str(record.get("product_name", "")) if record.get("product_name") else "")
        ctk.CTkEntry(head_row1, textvariable=product_var, width=240, height=32,
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     fg_color=self.C["bg"], border_color=self.C["border"]).pack(side="left", padx=(6, 0))

        # 第二行：规格 + 品牌
        head_row2 = ctk.CTkFrame(head_frame, fg_color="transparent")
        head_row2.pack(fill="x", pady=4)

        ctk.CTkLabel(head_row2, text="规格", width=80, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     text_color=self.C["text_secondary"]).pack(side="left")
        spec_var = tk.StringVar(value=str(record.get("spec", "")) if record.get("spec") else "")
        ctk.CTkEntry(head_row2, textvariable=spec_var, width=240, height=32,
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     fg_color=self.C["bg"], border_color=self.C["border"]).pack(side="left", padx=(6, 24))

        ctk.CTkLabel(head_row2, text="品牌", width=60, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     text_color=self.C["text_secondary"]).pack(side="left")
        brand_var = tk.StringVar(value=str(record.get("brand", "")) if record.get("brand") else "")
        ctk.CTkEntry(head_row2, textvariable=brand_var, width=240, height=32,
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     fg_color=self.C["bg"], border_color=self.C["border"]).pack(side="left", padx=(6, 0))

        # ── 物料列表（可多行）──
        mat_head_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        mat_head_frame.pack(fill="x", padx=20, pady=(16, 6))

        ctk.CTkLabel(mat_head_frame, text="物料清单",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=15, weight="bold"),
                     text_color=self.C["text"]).pack(side="left")

        ctk.CTkButton(
            mat_head_frame, text="➕  添加物料", width=110, height=30,
            fg_color="transparent", hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=lambda: self._add_material_row(material_list_frame, material_rows, {}, focus_first=True),
            corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right")

        # 物料滚动容器
        mat_outer = ctk.CTkFrame(dialog, fg_color=self.C["bg"], corner_radius=8,
                                 border_width=1, border_color=self.C["border"])
        mat_outer.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        # 列头
        header_row = ctk.CTkFrame(mat_outer, fg_color="transparent")
        header_row.pack(fill="x", padx=8, pady=(8, 4))
        ctk.CTkLabel(header_row, text="物料项目号", width=110, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=12, weight="bold"),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(header_row, text="物料名称", width=200, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=12, weight="bold"),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(header_row, text="数量", width=80, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=12, weight="bold"),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(header_row, text="计量单位", width=80, anchor="w",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=12, weight="bold"),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(header_row, text="", width=70, anchor="w").pack(side="right")

        # 滚动区
        scroll = ctk.CTkScrollableFrame(mat_outer, fg_color="transparent", height=260)
        scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        material_list_frame = scroll  # 物料行容器

        material_rows = []  # 存放每行的 row_frame + vars

        # 初始化：编辑模式回填1行；新增模式给1行空行
        if not is_new:
            self._add_material_row(material_list_frame, material_rows, {
                "material_project_no": record.get("material_project_no", ""),
                "material_name":       record.get("material_name", ""),
                "quantity":            record.get("quantity", ""),
                "unit":                record.get("unit", ""),
            })
        else:
            self._add_material_row(material_list_frame, material_rows, {})

        # ── 底部按钮 ──
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(8, 16))

        if not is_new:
            ctk.CTkButton(
                btn_frame, text="🗑  删除", width=90, height=36,
                fg_color="transparent", hover_color="#A05A5A",
                font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                command=lambda: self._on_delete(record["id"], dialog), corner_radius=8,
                border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="left")

        ctk.CTkButton(
            btn_frame, text="✓  保存", width=100, height=36,
            fg_color="transparent", hover_color="#7A9472",
            font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
            command=lambda: _save(), corner_radius=8,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right", padx=(8, 0))

        ctk.CTkButton(
            btn_frame, text="取消", width=80, height=36,
            fg_color="transparent",
            border_color=self.C["border"], border_width=1,
            font=ctk.CTkFont(family="Microsoft YaHei", size=14),
            command=dialog.destroy, corner_radius=8,
 text_color="#000000",).pack(side="right")

        def _save():
            fpn = finished_var.get().strip()
            if not fpn:
                messagebox.showwarning("提示", "成品项目号不能为空")
                return

            # 收集所有物料行
            valid_rows = []
            for r in material_rows:
                mpn = r["mat_no_var"].get().strip()
                mname = r["mat_name_var"].get().strip()
                if not mpn and not mname:
                    continue  # 空行跳过
                if not mpn:
                    messagebox.showwarning("提示", "物料项目号不能为空（已自动跳过全空行）")
                    return
                try:
                    qty = float(r["qty_var"].get().strip()) if r["qty_var"].get().strip() else 0
                except ValueError:
                    messagebox.showwarning("提示", f"物料「{mname or mpn}」的数量格式错误")
                    return
                valid_rows.append({
                    "finished_project_no": fpn,
                    "product_name":        product_var.get().strip(),
                    "spec":                spec_var.get().strip(),
                    "brand":               brand_var.get().strip(),
                    "material_project_no": mpn,
                    "material_name":       mname,
                    "quantity":            qty,
                    "unit":                r["unit_var"].get().strip(),
                })

            if not valid_rows:
                messagebox.showwarning("提示", "请至少填写一条物料")
                return

            try:
                if is_new:
                    # 新增：插入所有行
                    self.db.save_product_bom_batch(valid_rows)
                else:
                    # 编辑：删旧插新（简单可靠）
                    self.db.delete_product_bom(record["id"])
                    self.db.save_product_bom_batch(valid_rows)
                dialog.destroy()
                self._on_query()
                messagebox.showinfo("成功", f"保存成功，共 {len(valid_rows)} 条物料")
            except Exception as e:
                messagebox.showerror("保存失败", f"保存时出错：\n{e}")

    def _add_material_row(self, parent, material_rows, init_data, focus_first=False):
        """添加一行物料输入"""
        row_frame = ctk.CTkFrame(parent, fg_color=self.C["card"], corner_radius=6,
                                 border_width=1, border_color=self.C["border"])
        row_frame.pack(fill="x", pady=3)

        mat_no_var = tk.StringVar(value=str(init_data.get("material_project_no", "")) if init_data.get("material_project_no") else "")
        mat_name_var = tk.StringVar(value=str(init_data.get("material_name", "")) if init_data.get("material_name") else "")
        qty_var = tk.StringVar(value=str(init_data.get("quantity", "")) if init_data.get("quantity") not in (None, "", 0) else "")
        unit_var = tk.StringVar(value=str(init_data.get("unit", "")) if init_data.get("unit") else "")

        e1 = ctk.CTkEntry(row_frame, textvariable=mat_no_var, width=110, height=30,
                          font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                          fg_color=self.C["bg"], border_color=self.C["border"],
                          placeholder_text="物料项目号")
        e1.pack(side="left", padx=(4, 6), pady=4)
        if focus_first and not mat_no_var.get():
            e1.focus_set()

        e2 = ctk.CTkEntry(row_frame, textvariable=mat_name_var, width=220, height=30,
                          font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                          fg_color=self.C["bg"], border_color=self.C["border"],
                          placeholder_text="物料名称")
        e2.pack(side="left", padx=(0, 6), pady=4)

        e3 = ctk.CTkEntry(row_frame, textvariable=qty_var, width=80, height=30,
                          font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                          fg_color=self.C["bg"], border_color=self.C["border"],
                          placeholder_text="数量")
        e3.pack(side="left", padx=(0, 6), pady=4)

        e4 = ctk.CTkEntry(row_frame, textvariable=unit_var, width=80, height=30,
                          font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                          fg_color=self.C["bg"], border_color=self.C["border"],
                          placeholder_text="单位")
        e4.pack(side="left", padx=(0, 6), pady=4)

        def _remove():
            material_rows.remove(row_info)
            row_frame.destroy()

        ctk.CTkButton(
            row_frame, text="🗑", width=40, height=30,
            fg_color="transparent",
            hover_color="#FEE2E2",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=_remove, corner_radius=6,
            border_width=1, border_color="#8B7D6B",
 text_color="#000000",).pack(side="right", padx=4, pady=4)

        row_info = {
            "frame": row_frame,
            "mat_no_var": mat_no_var,
            "mat_name_var": mat_name_var,
            "qty_var": qty_var,
            "unit_var": unit_var,
        }
        material_rows.append(row_info)
        return row_info

    def _on_delete(self, item_id, dialog):
        if messagebox.askyesno("确认删除", "确定要删除这条BOM记录吗？"):
            self.db.delete_product_bom(item_id)
            dialog.destroy()
            self._on_query()
            messagebox.showinfo("成功", "删除成功")
