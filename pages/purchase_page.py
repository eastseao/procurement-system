#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""采购垫付页面 v1.2 - 新增导入xlsx"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import customtkinter as ctk
from datetime import datetime, date


class PurchasePage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.db = db
        self.C = colors
        self.show_archived = False
        self._build()
        self._load_records()

    def _build(self):
        # 顶部标题栏
        header = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0, height=52)
        header.pack(fill="x", padx=20, pady=(16, 8))
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="采购垫付",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=16, weight="bold"),
                     text_color=self.C["text"]).pack(side="left", pady=14)

        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="right", padx=16)

        self.archive_btn = ctk.CTkButton(
            btn_frame, text="📁 查看归档", width=100, height=34,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=14), command=self._toggle_archive, corner_radius=20)
        self.archive_btn.pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="＋ 新增垫付", width=110, height=34,
            fg_color=self.C["danger"], hover_color="#A85A5A",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._open_form, corner_radius=20).pack(side="left", padx=4)

        ctk.CTkButton(
            btn_frame, text="📤 导出Excel", width=110, height=34,
            fg_color=self.C["success"], hover_color="#7A9A6E",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._export_xlsx, corner_radius=20).pack(side="right", padx=4)

        # ── v1.2 新增：导入xlsx 按钮 ──
        ctk.CTkButton(
            btn_frame, text="📥 导入xlsx", width=100, height=34,
            fg_color="#6B7280", hover_color="#4B5563",
            font=ctk.CTkFont(size=14),
            command=self._import_xlsx, corner_radius=20).pack(side="right", padx=4)

        # 统计栏
        self.stats_frame = ctk.CTkFrame(self, fg_color=self.C["card"],
                                         corner_radius=self.C["radius_card"], height=72)
        self.stats_frame.pack(fill="x", padx=16, pady=(12, 0))
        self.stats_frame.pack_propagate(False)
        self.stat_labels = {}
        for key, label, color in [
            ("total_count", "总记录数", self.C["primary"]),
            ("total_amount", "垫付总金额 (¥)", self.C["warning"]),
            ("unreimbursed", "未报销金额 (¥)", self.C["danger"]),
            ("uninvoiced", "未开票数", self.C["warning"]),
        ]:
            f = ctk.CTkFrame(self.stats_frame, fg_color="transparent")
            f.pack(side="left", padx=24, pady=8)
            ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=12),
                         text_color=self.C["text_secondary"]).pack()
            lbl = ctk.CTkLabel(f, text="0", font=ctk.CTkFont(size=20, weight="bold"),
                               text_color=color)
            lbl.pack()
            self.stat_labels[key] = lbl

        # 筛选栏
        filter_frame = ctk.CTkFrame(self, fg_color=self.C["card"],
                                     corner_radius=self.C["radius_card"], height=52)
        filter_frame.pack(fill="x", padx=16, pady=8)
        filter_frame.pack_propagate(False)

        ctk.CTkLabel(filter_frame, text="报销状态:",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left", padx=(16, 4), pady=12)
        self.filter_reimburse = ctk.CTkComboBox(
            filter_frame, values=["全部", "未报销", "已报销"], width=100, height=28,
            command=lambda _: self._load_records())
        self.filter_reimburse.set("全部")
        self.filter_reimburse.pack(side="left", padx=4)

        ctk.CTkLabel(filter_frame, text="开票状态:",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left", padx=(12, 4))
        self.filter_invoice = ctk.CTkComboBox(
            filter_frame, values=["全部", "未开票", "已开票"], width=100, height=28,
            command=lambda _: self._load_records())
        self.filter_invoice.set("全部")
        self.filter_invoice.pack(side="left", padx=4)

        ctk.CTkLabel(filter_frame, text="项目:",
                     font=ctk.CTkFont(size=13), text_color=self.C["text_secondary"]).pack(
            side="left", padx=(12, 4))
        self.filter_project = ctk.CTkComboBox(
            filter_frame, values=["全部"], width=120, height=28,
            command=lambda _: self._load_records())
        self.filter_project.set("全部")
        self.filter_project.pack(side="left", padx=4)
        self._refresh_project_filter()

        ctk.CTkButton(filter_frame, text="↺ 刷新", width=60, height=28,
                      fg_color=self.C["border"], text_color=self.C["text"],
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=13),
                      command=self._load_records, corner_radius=20).pack(side="left", padx=8)

        # 列表区（带滚动）
        # 表格区域 - 阴影分层：z1 → z2 → z3
        # z1（底层）：fg_color=#E8E2D9, corner_radius=16, 内边距 4
        z1_frame = ctk.CTkFrame(self, fg_color="#E8E2D9", corner_radius=16)
        z1_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        # z2（中层）：fg_color=#F2F0EB, corner_radius=12, 内边距 4
        z2_frame = ctk.CTkFrame(z1_frame, fg_color="#F2F0EB", corner_radius=12)
        z2_frame.pack(fill="both", expand=True, padx=4, pady=4)

        # z3（上层卡片）：fg_color=C["card"], corner_radius=C["radius_card"]
        list_frame = ctk.CTkFrame(z2_frame, fg_color=self.C["card"], corner_radius=self.C["radius_card"])
        list_frame.pack(fill="both", expand=True, padx=4, pady=4)

        cols = [("date", "日期", 90), ("project", "项目", 100), ("handler", "经手人", 70),
                ("payment", "支付方式", 90), ("items_summary", "物料摘要", 260),
                ("total", "合计(¥)", 90), ("invoice", "开票", 70),
                ("reimburse", "报销", 70), ("action", "操作", 140)]

        style = ttk.Style()
        style.configure("Purchase.Treeview", font=("Microsoft YaHei", 9),
                         rowheight=38, background="#FFFFFF", fieldbackground="#FFFFFF",
                         foreground="#1E293B")
        style.configure("Purchase.Treeview.Heading", font=("Microsoft YaHei", 9, "bold"),
                         background="#F8FAFC", foreground="#475569", relief="flat")
        style.map("Purchase.Treeview",
                  background=[("selected", "#E8D5C4")],
                  foreground=[("selected", "#4A3728")])

        tree_frame = tk.Frame(list_frame, bg="#FFFFFF")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(tree_frame, style="Purchase.Treeview",
                                  columns=[c[0] for c in cols], show="headings",
                                  selectmode="browse")
        for cid, label, width in cols:
            self.tree.heading(cid, text=label)
            self.tree.column(cid, width=width, minwidth=40, stretch=True, anchor="center" if cid != "items_summary" else "w")

        vsb = ctk.CTkScrollbar(tree_frame, orientation="vertical", command=self.tree.yview, button_color=self.C["border"], button_hover_color=self.C.get("sidebar_hover", "#ddd"), width=8)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<Button-1>", self._on_click)

        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.tag_configure("unreimbursed", foreground="#DC2626")
        self.tree.tag_configure("archived", foreground="#94A3B8")
        self.tree.tag_configure("hover", background="#FFF2E6")
        self.tree.bind("<Motion>", self._on_hover)
        self.tree.bind("<Leave>", self._on_leave)

        # ── 右键菜单 ──
        self._context_menu = tk.Menu(self, tearoff=0,
            bg="#FFFFFF", fg="#4A3728",
            activebackground="#E8D5C4", activeforeground="#4A3728",
            font=("Microsoft YaHei", 10),
        )
        self._context_menu.add_command(label="📝 编辑", command=self._on_edit_selected)
        self._context_menu.add_command(label="📋 复制", command=self._on_copy_selected)
        self._context_menu.add_separator()
        self._context_menu.add_command(label="🗑️ 删除", command=self._on_delete_selected)
        self.tree.bind("<Button-3>", self._on_tree_right_click)

    def _refresh_project_filter(self):
        projects = ["全部"] + self.db.get_projects()
        self.filter_project.configure(values=projects)

    def _load_records(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        records = self.db.get_purchases(archived=1 if self.show_archived else 0)
        ri_filter = self.filter_reimburse.get()
        inv_filter = self.filter_invoice.get()
        proj_filter = self.filter_project.get()

        total_amount = 0
        unreimbursed = 0
        uninvoiced = 0

        for r in records:
            if ri_filter != "全部" and r["reimbursement_status"] != ri_filter:
                continue
            if inv_filter != "全部" and r["invoice_status"] != inv_filter:
                continue
            if proj_filter != "全部" and r["project"] != proj_filter:
                continue

            total_amount += r["total"]
            if r["reimbursement_status"] == "未报销":
                unreimbursed += r["total"]
            if r["invoice_status"] == "未开票":
                uninvoiced += 1

            items_summary = "；".join(
                f"{i['name']}×{i['quantity']}" for i in r["items"] if i.get("name")
            ) or "—"
            if len(items_summary) > 30:
                items_summary = items_summary[:30] + "…"

            tag = "unreimbursed" if r["reimbursement_status"] == "未报销" else ""
            if self.show_archived:
                tag = "archived"

            self.tree.insert("", "end", iid=str(r["id"]),
                             values=(
                                 r["date"], r["project"], r["handler"],
                                 r["payment_method"], items_summary,
                                 f"{r['total']:.2f}",
                                 r["invoice_status"], r["reimbursement_status"],
                                 "编辑  归档  删除" if not self.show_archived else "查看  删除"
                             ), tags=(tag,))

        count = len(self.tree.get_children())
        self.stat_labels["total_count"].configure(text=str(count))
        self.stat_labels["total_amount"].configure(text=f"{total_amount:.2f}")
        self.stat_labels["unreimbursed"].configure(text=f"{unreimbursed:.2f}")
        self.stat_labels["uninvoiced"].configure(text=str(uninvoiced))

    def _toggle_archive(self):
        self.show_archived = not self.show_archived
        if self.show_archived:
            self.archive_btn.configure(text="📋 返回列表", fg_color=self.C["success"])
        else:
            self.archive_btn.configure(text="📁 查看归档", fg_color="#6B7280")
        self._load_records()

    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item:
            return
        pid = int(item)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == 8:
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                if not self.show_archived:
                    if x_rel < col_w * 0.33:
                        self._open_form(pid)
                    elif x_rel < col_w * 0.66:
                        self._archive_record(pid)
                    # 删除已改为单击触发，双击不再处理
                else:
                    if x_rel < col_w * 0.5:
                        self._open_form(pid, view_only=True)
                    # 删除已改为单击触发，双击不再处理
            else:
                self._open_form(pid)
        else:
            self._open_form(pid)

    def _archive_record(self, pid):
        if messagebox.askyesno("归档确认", "确定将此记录归档？归档后不可编辑。"):
            self.db.archive_purchase(pid)
            self._load_records()

    def _delete_record(self, pid):
        """直接删除记录（无确认）"""
        self.db.delete_purchase(pid)
        self._load_records()

    def _open_form(self, pid=None, view_only=False):
        try:
            form = PurchaseForm(self, self.db, self.C, pid=pid,
                                view_only=view_only,
                                on_save=self._on_form_save)
            form.grab_set()
        except Exception as e:
            import traceback
            messagebox.showerror("打开表单失败", f"{e}\n\n{traceback.format_exc()}")

    def _on_form_save(self):
        self._load_records()
        self._refresh_project_filter()

    # ── 导出 Excel ─────────────────────────────────
    def _export_xlsx(self):
        records = self.db.get_purchases(archived=1 if self.show_archived else 0)
        if not records:
            messagebox.showinfo("提示", "没有可导出的数据")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出采购垫付", defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"采购垫付_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not filepath:
            return

        headers = ["日期", "项目归属", "经手人", "支付方式", "物料名称", "规格", "数量",
                    "单价", "供应商", "合计", "开票状态", "报销状态", "备注"]
        rows = []
        for r in records:
            items = r.get("items") or []
            if items:
                for item in items:
                    rows.append({
                        "日期": r.get("date", ""),
                        "项目归属": r.get("project", ""),
                        "经手人": r.get("handler", ""),
                        "支付方式": r.get("payment_method", ""),
                        "物料名称": item.get("name", ""),
                        "规格": item.get("spec", ""),
                        "数量": item.get("quantity", ""),
                        "单价": item.get("unit_price", ""),
                        "供应商": item.get("supplier", ""),
                        "合计": item.get("total", ""),
                        "开票状态": r.get("invoice_status", ""),
                        "报销状态": r.get("reimbursement_status", ""),
                        "备注": r.get("remark", ""),
                    })
            else:
                rows.append({
                    "日期": r.get("date", ""),
                    "项目归属": r.get("project", ""),
                    "经手人": r.get("handler", ""),
                    "支付方式": r.get("payment_method", ""),
                    "物料名称": "", "规格": "", "数量": "", "单价": "", "供应商": "", "合计": "",
                    "开票状态": r.get("invoice_status", ""),
                    "报销状态": r.get("reimbursement_status", ""),
                    "备注": r.get("remark", ""),
                })

        try:
            self.db.export_to_xlsx(filepath, "采购垫付", headers, rows,
                                   col_widths=[12, 12, 8, 10, 16, 12, 8, 8, 14, 10, 10, 10, 16])
            messagebox.showinfo("导出成功", f"已导出 {len(rows)} 条记录到：\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── v1.2 新增：导入xlsx ────────────────────────────
    def _import_xlsx(self):
        filepath = filedialog.askopenfilename(
            title="选择采购垫付xlsx文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            # 列名映射
            col_map = {}
            field_candidates = {
                "date":                  ["日期", "date"],
                "project":               ["项目归属", "项目", "project"],
                "handler":               ["经手人", "handler"],
                "payment_method":        ["支付方式", "payment_method"],
                "invoice_status":        ["开票状态", "invoice_status"],
                "reimbursement_status":  ["报销状态", "reimbursement_status"],
                "remark":                ["备注", "remark"],
                "item_name":             ["物料名称", "名称", "item_name"],
                "item_spec":             ["规格", "spec"],
                "item_quantity":         ["数量", "quantity"],
                "item_unit_price":       ["单价", "unit_price"],
                "item_supplier":         ["供应商", "supplier"],
                "item_total":            ["合计", "total"],
            }
            for idx, h in enumerate(headers):
                for field, candidates in field_candidates.items():
                    if h in candidates and field not in col_map:
                        col_map[field] = idx
                        break

            if "date" not in col_map:
                messagebox.showerror("导入失败", "未找到「日期」列，请检查xlsx格式")
                return

            def _cell(row_vals, key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row_vals):
                    return ""
                v = row_vals[idx]
                return str(v).strip() if v is not None else ""

            # 按日期+项目+经手人分组，每组对应一条采购记录
            groups = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                date_val = _cell(row, "date")
                if not date_val:
                    continue
                project = _cell(row, "project") or "默认项目"
                handler = _cell(row, "handler") or "未知"
                payment = _cell(row, "payment_method") or "微信"
                invoice = _cell(row, "invoice_status") or "未开票"
                reimburse = _cell(row, "reimbursement_status") or "未报销"
                remark = _cell(row, "remark")

                gkey = (date_val, project, handler)
                if gkey not in groups:
                    groups[gkey] = {
                        "date": date_val,
                        "project": project,
                        "handler": handler,
                        "payment_method": payment,
                        "invoice_status": invoice,
                        "reimbursement_status": reimburse,
                        "remark": remark,
                        "items": [],
                    }

                item_name = _cell(row, "item_name")
                if item_name:
                    try:
                        qty = float(_cell(row, "item_quantity") or 0)
                        price = float(_cell(row, "item_unit_price") or 0)
                        total = float(_cell(row, "item_total") or 0) or qty * price
                    except (ValueError, TypeError):
                        qty = price = total = 0
                    groups[gkey]["items"].append({
                        "name": item_name,
                        "spec": _cell(row, "item_spec"),
                        "quantity": qty,
                        "unit_price": price,
                        "supplier": _cell(row, "item_supplier"),
                        "total": total,
                    })

            imported = 0
            for gdata in groups.values():
                data = {k: v for k, v in gdata.items() if k != "items"}
                items = gdata["items"]
                self.db.save_purchase(data, items)
                imported += 1

            self._load_records()
            self._refresh_project_filter()
            messagebox.showinfo("导入成功", f"成功导入 {imported} 条采购垫付记录")

        except Exception as e:
            import traceback
            messagebox.showerror("导入失败", f"{e}\n\n{traceback.format_exc()}")



    # ── 单击选中高亮行 / 操作列"删除"单击即删 ──
    def _on_click(self, event):
        """单击选中行；操作列"删除"区域单击直接删除"""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        self.tree.selection_set(item)
        col = self.tree.identify_column(event.x)
        col_idx = int(col.replace("#", "")) - 1
        # 操作列：检测是否点击了"删除"区域
        if col_idx == 8:
            pid = int(item)
            bbox = self.tree.bbox(item, col)
            if bbox:
                x_rel = event.x - bbox[0]
                col_w = bbox[2]
                if not self.show_archived:
                    # "编辑  归档  删除" — 删除在右 1/3
                    if x_rel >= col_w * 0.66:
                        self._delete_record(pid)
                else:
                    # "查看  删除" — 删除在右 1/2
                    if x_rel >= col_w * 0.5:
                        self._delete_record(pid)
            else:
                # bbox 不可见时，直接删除（兜底）
                self._delete_record(pid)

    # ── 右键菜单回调 ──────────────────────────────
    def _on_tree_right_click(self, event):
        """右键点击表格行→弹出菜单"""
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self._context_menu.post(event.x_root, event.y_root)

    def _on_edit_selected(self):
        """编辑选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        pid = int(selection[0])
        self._open_form(pid, view_only=self.show_archived)

    def _on_copy_selected(self):
        """复制选中行信息到剪贴板"""
        selection = self.tree.selection()
        if not selection:
            return
        pid = int(selection[0])
        records = self.db.get_purchases(archived=1 if self.show_archived else 0)
        rec = next((r for r in records if r["id"] == pid), None)
        if rec:
            text = f"项目:{rec.get('project','')} 经手人:{rec.get('handler','')} 合计:¥{rec.get('total',0):.2f}"
            self.clipboard_clear()
            self.clipboard_append(text)

    def _on_delete_selected(self):
        """删除选中行"""
        selection = self.tree.selection()
        if not selection:
            return
        pid = int(selection[0])
        if messagebox.askyesno("删除确认", "确定删除此采购垫付记录？", icon="warning"):
            self._delete_record(pid)

    def _on_hover(self, event):
        """鼠标悬浮时高亮行背景"""
        item = self.tree.identify_row(event.y)
        if item and item != getattr(self, "_last_hover", None):
            if hasattr(self, "_last_hover") and self._last_hover:
                tags = list(self.tree.item(self._last_hover, "tags")[0])
                if "hover" in tags:
                    tags.remove("hover")
                    self.tree.item(self._last_hover, tags=tags)
            tags = list(self.tree.item(item, "tags")[0])
            if "hover" not in tags:
                tags.append("hover")
                self.tree.item(item, tags=tags)
            self._last_hover = item
        elif not item and hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

    def _on_leave(self, event):
        """鼠标离开时清除悬浮高亮"""
        if hasattr(self, "_last_hover") and self._last_hover:
            tags = list(self.tree.item(self._last_hover, "tags")[0])
            if "hover" in tags:
                tags.remove("hover")
                self.tree.item(self._last_hover, tags=tags)
            self._last_hover = None

# ============================
# 采购垫付表单弹窗
# ============================
class PurchaseForm(ctk.CTkToplevel):
    def __init__(self, parent, db, colors, pid=None, view_only=False, on_save=None):
        super().__init__(parent)
        self.db = db
        self.C = colors
        self.pid = pid
        self.view_only = view_only
        self.on_save = on_save
        self.item_rows = []

        title = "查看垫付记录" if view_only else ("编辑垫付记录" if pid else "新增垫付记录")
        self.title(title)
        self.geometry("860x700")
        self.resizable(True, True)
        self.configure(fg_color=self.C["bg"])

        self.update_idletasks()
        pw, ph = parent.winfo_rootx(), parent.winfo_rooty()
        self.geometry(f"860x700+{pw+100}+{ph+50}")

        self._build()
        if pid:
            self._load_data(pid)

    def _build(self):
        outer = ctk.CTkScrollableFrame(self, fg_color=self.C["bg"])
        outer.pack(fill="both", expand=True, padx=0, pady=0)
        self.form_outer = outer

        card = ctk.CTkFrame(outer, fg_color=self.C["card"], corner_radius=self.C["radius_modal"])
        card.pack(fill="both", expand=True, padx=16, pady=16)

        # === 基础信息 ===
        self._section(card, "📋 基础信息")
        row0 = ctk.CTkFrame(card, fg_color="transparent")
        row0.pack(fill="x", padx=16, pady=(0, 8))

        ctk.CTkLabel(row0, text="日期 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=0, column=0, padx=4, pady=4, sticky="w")
        self.date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        date_entry = ctk.CTkEntry(row0, textvariable=self.date_var, width=130, height=32)
        date_entry.grid(row=0, column=1, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(row0, text="经手人 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=0, column=2, padx=(16, 4), pady=4, sticky="w")
        self.handler_var = tk.StringVar(value="纪委")
        ctk.CTkEntry(row0, textvariable=self.handler_var, width=120, height=32).grid(
            row=0, column=3, padx=4, pady=4, sticky="w")

        ctk.CTkLabel(row0, text="项目归属 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=1, column=0, padx=4, pady=4, sticky="w")
        projects = self.db.get_projects()
        self.project_var = tk.StringVar(value=projects[0] if projects else "默认项目")
        self.project_combo = ctk.CTkComboBox(row0, values=projects,
                                              variable=self.project_var, width=150, height=32)
        self.project_combo.grid(row=1, column=1, padx=4, pady=4, sticky="w")
        ctk.CTkButton(row0, text="＋项目", width=60, height=28,
                      fg_color="#E2E8F0", text_color=self.C["text"],
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=12),
                      command=self._add_project, corner_radius=20).grid(row=1, column=2, padx=4)

        ctk.CTkLabel(row0, text="支付方式 *", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).grid(row=1, column=3, padx=(8, 4), pady=4, sticky="w")
        self.payment_var = tk.StringVar(value="微信")
        ctk.CTkComboBox(row0, values=["微信", "支付宝", "淘宝", "银行卡", "许丹红", "宋总"],
                         variable=self.payment_var, width=120, height=32).grid(
            row=1, column=4, padx=4, pady=4, sticky="w")

        row1 = ctk.CTkFrame(card, fg_color="transparent")
        row1.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(row1, text="开票状态:", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(0, 4))
        self.invoice_var = tk.StringVar(value="未开票")
        for v in ["未开票", "已开票"]:
            ctk.CTkRadioButton(row1, text=v, variable=self.invoice_var, value=v,
                               font=ctk.CTkFont(size=13)).pack(side="left", padx=8)

        ctk.CTkLabel(row1, text="  报销状态:", font=ctk.CTkFont(size=13),
                     text_color=self.C["text_secondary"]).pack(side="left", padx=(16, 4))
        self.reimburse_var = tk.StringVar(value="未报销")
        for v in ["未报销", "已报销"]:
            ctk.CTkRadioButton(row1, text=v, variable=self.reimburse_var, value=v,
                               font=ctk.CTkFont(size=13)).pack(side="left", padx=8)

        # === 物料明细 ===
        self._section(card, "📦 物料明细")
        items_header = ctk.CTkFrame(card, fg_color="#F8FAFC", corner_radius=6)
        items_header.pack(fill="x", padx=16, pady=(0, 4))
        for txt, w in [("名称", 140), ("规格", 100), ("数量", 70), ("单价", 80),
                       ("供应商", 130), ("合计", 90), ("", 40)]:
            ctk.CTkLabel(items_header, text=txt, width=w, font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=self.C["text_secondary"]).pack(side="left", padx=2, pady=6)

        self.items_container = ctk.CTkFrame(card, fg_color="transparent")
        self.items_container.pack(fill="x", padx=16)
        self._add_item_row()

        ctk.CTkButton(card, text="＋ 添加物料行", height=30, width=120,
                      fg_color="#E2E8F0", text_color=self.C["text"],
                      hover_color="#CBD5E1", font=ctk.CTkFont(size=13),
                      command=self._add_item_row, corner_radius=20).pack(anchor="w", padx=16, pady=4)

        total_row = ctk.CTkFrame(card, fg_color="transparent")
        total_row.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(total_row, text="物料合计：",
                     font=ctk.CTkFont(size=14), text_color=self.C["text_secondary"]).pack(side="left")
        self.total_label = ctk.CTkLabel(total_row, text="¥ 0.00",
                                         font=ctk.CTkFont(size=17, weight="bold"),
                                         text_color=self.C["warning"])
        self.total_label.pack(side="left")

        # === 备注 ===
        self._section(card, "📝 备注")
        self.remark_text = ctk.CTkTextbox(card, height=80, font=ctk.CTkFont(size=13))
        self.remark_text.pack(fill="x", padx=16, pady=(0, 12))

        # === 按钮 ===
        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=12)
        if not self.view_only:
            ctk.CTkButton(btn_row, text="✓ 保存", width=120, height=38,
                          fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
                          font=ctk.CTkFont(size=16, weight="bold"),
                          command=self._save, corner_radius=20).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="✕ 关闭", width=80, height=38,
                      fg_color="#6B7280", hover_color="#4B5563",
                      font=ctk.CTkFont(size=14), command=self.destroy, corner_radius=20).pack(side="left", padx=4)

    def _section(self, parent, title):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=self.C["primary"]).pack(side="left")
        ctk.CTkFrame(f, height=1, fg_color=self.C["border"]).pack(
            side="left", fill="x", expand=True, padx=(8, 0))

    def _add_item_row(self, data=None):
        row = ctk.CTkFrame(self.items_container, fg_color="transparent")
        row.pack(fill="x", pady=2)

        vars_ = {}
        fields = [
            ("name", 140, "名称"), ("spec", 100, "规格"),
            ("quantity", 70, "数量"), ("unit_price", 80, "单价"),
            ("supplier", 130, "供应商"),
        ]
        for fkey, w, ph in fields:
            v = tk.StringVar(value=str(data.get(fkey, "")) if data else "")
            vars_[fkey] = v
            e = ctk.CTkEntry(row, textvariable=v, width=w, height=28,
                             placeholder_text=ph, font=ctk.CTkFont(size=12))
            e.pack(side="left", padx=2)
            if fkey in ("quantity", "unit_price"):
                v.trace_add("write", lambda *a, r=row: self._recalc_row(r))

        total_var = tk.StringVar(value="0.00" if not data else f"{data.get('total', 0):.2f}")
        vars_["total"] = total_var
        ctk.CTkEntry(row, textvariable=total_var, width=90, height=28,
                     state="readonly", font=ctk.CTkFont(size=12)).pack(side="left", padx=2)

        del_btn = ctk.CTkButton(row, text="✕", width=32, height=28,
                                 fg_color="#FEE2E2", text_color=self.C["danger"],
                                 hover_color="#FECACA", font=ctk.CTkFont(size=12),
                                 command=lambda r=row: self._remove_item_row(r), corner_radius=20)
        del_btn.pack(side="left", padx=2)

        self.item_rows.append({"frame": row, "vars": vars_})

    def _remove_item_row(self, frame):
        for i, r in enumerate(self.item_rows):
            if r["frame"] == frame:
                self.item_rows.pop(i)
                break
        frame.destroy()
        self._update_total()

    def _recalc_row(self, frame):
        for r in self.item_rows:
            if r["frame"] == frame:
                try:
                    q = float(r["vars"]["quantity"].get() or 0)
                    p = float(r["vars"]["unit_price"].get() or 0)
                    r["vars"]["total"].set(f"{q*p:.2f}")
                except ValueError:
                    pass
                break
        self._update_total()

    def _update_total(self):
        total = 0
        for r in self.item_rows:
            try:
                total += float(r["vars"]["total"].get() or 0)
            except ValueError:
                pass
        self.total_label.configure(text=f"¥ {total:.2f}")

    def _add_project(self):
        name = simpledialog.askstring("添加项目", "请输入项目名称：", parent=self)
        if name and name.strip():
            self.db.add_project(name.strip())
            projects = self.db.get_projects()
            self.project_combo.configure(values=projects)
            self.project_var.set(name.strip())

    def _load_data(self, pid):
        records = self.db.get_purchases()
        arch = self.db.get_purchases(archived=1)
        all_r = records + arch
        r = next((x for x in all_r if x["id"] == pid), None)
        if not r:
            return
        self.date_var.set(r["date"])
        self.handler_var.set(r["handler"])
        self.project_var.set(r["project"])
        self.payment_var.set(r["payment_method"])
        self.invoice_var.set(r["invoice_status"])
        self.reimburse_var.set(r["reimbursement_status"])
        self.remark_text.delete("1.0", "end")
        if r["remark"]:
            self.remark_text.insert("1.0", r["remark"])
        for row_data in self.item_rows:
            row_data["frame"].destroy()
        self.item_rows.clear()
        for item in r["items"]:
            self._add_item_row(item)
        if not r["items"]:
            self._add_item_row()
        self._update_total()

    def _save(self):
        date_val = self.date_var.get().strip()
        handler = self.handler_var.get().strip()
        project = self.project_var.get().strip()
        payment = self.payment_var.get().strip()
        if not date_val:
            messagebox.showerror("验证失败", "请填写日期", parent=self)
            return
        if not handler:
            messagebox.showerror("验证失败", "请填写经手人", parent=self)
            return

        items = []
        for r in self.item_rows:
            v = r["vars"]
            name = v["name"].get().strip()
            if not name:
                continue
            try:
                qty = float(v["quantity"].get() or 0)
                price = float(v["unit_price"].get() or 0)
                total = float(v["total"].get() or 0)
            except ValueError:
                total = qty = price = 0
            items.append({
                "name": name,
                "spec": v["spec"].get().strip(),
                "quantity": qty,
                "unit_price": price,
                "supplier": v["supplier"].get().strip(),
                "total": total,
            })

        data = {
            "date": date_val,
            "project": project,
            "handler": handler,
            "payment_method": payment,
            "invoice_status": self.invoice_var.get(),
            "reimbursement_status": self.reimburse_var.get(),
            "remark": self.remark_text.get("1.0", "end").strip(),
        }

        if self.pid:
            self.db.update_purchase(self.pid, data, items)
            messagebox.showinfo("成功", "记录已更新", parent=self)
        else:
            self.db.save_purchase(data, items)
            messagebox.showinfo("成功", "记录已保存", parent=self)

        if self.on_save:
            self.on_save()
        self.destroy()